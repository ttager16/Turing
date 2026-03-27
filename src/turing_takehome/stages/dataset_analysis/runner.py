from __future__ import annotations

import argparse
import csv
import json
import math
import re
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass
from itertools import combinations
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from turing_takehome.llm import (
    AsyncTaskSpec,
    DEFAULT_LOCAL_EMBEDDING_MODEL,
    embed_texts_cached_for_target,
    run_async_tasks_sync,
)
from turing_takehome.reporting import export_combined_report
from turing_takehome.reporting.notes import (
    generate_notes_for_requests,
    split_cached_note_requests,
    update_note_cache,
)
from turing_takehome.stages.sample_efficacy_analysis.data import load_samples


PROJECT_ROOT = Path(__file__).resolve().parents[4]
ARTIFACTS_DIR = PROJECT_ROOT / "artifacts"
OUTPUT_DIR = PROJECT_ROOT / "outputs" / "dataset_analysis"
STAGE1_WORKBOOK = PROJECT_ROOT / "outputs" / "sample_requirements_analysis" / "guideline_audit.xlsx"
STAGE2_DIR = PROJECT_ROOT / "outputs" / "sample_efficacy_analysis"

SECTION_PREFIXES = ("1.", "2.", "3.", "4_", "6.", "7_")
STATUS_NUMERIC = {
    "PASS": 1.0,
    "PARTIAL": 0.5,
    "UNCLEAR": 0.25,
    "FAIL": 0.0,
}
EXEMPLAR_EFFICACY = {"High Efficacy", "Moderate Efficacy"}
STAGE3_TEST_COLUMNS = [
    "PromptLengthOutlier",
    "StarterCodeOutlier",
    "IdealResponseOutlier",
    "TestCountOutlier",
    "PerformanceOutlier",
    "RedundancyStatus",
    "AttemptVarianceCheck",
    "ModelDisagreementCheck",
    "ThresholdSensitivityCheck",
    "DifficultySignalRegime",
    "ContradictionCheck",
    "BenchmarkDefectCandidate",
    "TrivialityCheck",
    "ExemplarCheck",
]
RELATIONSHIP_MIN_SUPPORT = 5
RELATIONSHIP_MIN_SUPPORT_SHARE = 0.03
RELATIONSHIP_MIN_DATASET_SIZE = 10
THRESHOLD_SENSITIVITY_MARGIN = 0.05
ATTEMPT_VARIANCE_RANGE_FLAG = 0.35
ATTEMPT_VARIANCE_STD_FLAG = 0.15
MODEL_DISAGREEMENT_GAP_FLAG = 0.25
MODEL_DISAGREEMENT_GAP_HIGH = 0.45
STAGE3_AUDIT_TARGETS = ("openai-gpt-5-mini", "local-qwen")
_ILLEGAL_XLSX_CHARS = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")
STAGE3_AUDITOR_SCHEMA = {
    "type": "object",
    "properties": {
        "dataset_utility_label": {
            "type": "string",
            "enum": ["strong", "usable", "caveated", "contradictory", "saturated"],
        },
        "primary_risk": {
            "type": "string",
            "enum": ["none", "redundancy", "instability", "benchmark_defect", "threshold_fragility", "outlier", "contradiction"],
        },
        "audit_priority": {
            "type": "string",
            "enum": ["normal", "medium", "high", "critical"],
        },
    },
    "required": ["dataset_utility_label", "primary_risk", "audit_priority"],
    "additionalProperties": False,
}


@dataclass(frozen=True)
class SimilarityPair:
    left_index: int
    right_index: int
    prompt_similarity: float
    template_similarity: float
    test_similarity: float
    starter_similarity: float
    title_similarity: float
    function_similarity: float
    embedding_similarity: float
    lexical_similarity: float
    structural_similarity: float
    combined_similarity: float
    similarity_label: str


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Run Stage 3 Dataset Analysis over source samples plus Stage 1/2 artifacts."
    )
    parser.add_argument(
        "--jsonl",
        type=Path,
        default=ARTIFACTS_DIR / "provided" / "Samples.jsonl",
    )
    parser.add_argument(
        "--stage1-workbook",
        type=Path,
        default=STAGE1_WORKBOOK,
    )
    parser.add_argument(
        "--stage2-dir",
        type=Path,
        default=STAGE2_DIR,
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=OUTPUT_DIR,
    )
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--offset", type=int, default=0)
    parser.add_argument("--indices", default=None)
    parser.add_argument("--near-duplicate-threshold", type=float, default=0.68)
    parser.add_argument("--template-threshold", type=float, default=0.55)
    parser.add_argument("--cluster-threshold", type=float, default=0.50)
    return parser


def run_cli(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    jsonl_path = _resolve_path(args.jsonl)
    stage1_workbook = _resolve_path(args.stage1_workbook)
    stage2_dir = _resolve_path(args.stage2_dir)
    output_dir = _resolve_path(args.output_dir)

    if not jsonl_path.exists():
        raise FileNotFoundError(f"Missing samples jsonl: {jsonl_path}")
    if not stage1_workbook.exists():
        raise FileNotFoundError(f"Missing Stage 1 workbook: {stage1_workbook}")
    if not stage2_dir.exists():
        raise FileNotFoundError(f"Missing Stage 2 output dir: {stage2_dir}")

    output_dir.mkdir(parents=True, exist_ok=True)

    stage1_rows = _load_stage1(stage1_workbook)
    stage2_context = _load_stage2(stage2_dir)
    joined_rows = _build_joined_rows(
        jsonl_path=jsonl_path,
        stage1_rows=stage1_rows,
        stage2_context=stage2_context,
        output_dir=output_dir,
        limit=args.limit,
        offset=args.offset,
        indices=args.indices,
        near_duplicate_threshold=args.near_duplicate_threshold,
        template_threshold=args.template_threshold,
        cluster_threshold=args.cluster_threshold,
    )

    summary = _build_summary(joined_rows, stage2_context)
    audit_queues = _build_audit_queues(joined_rows)
    relationship_rows = _build_relationship_rows(joined_rows)
    duplicate_rows = _build_duplicate_rows(joined_rows)
    detailed_rows, test_columns = _build_stage3_detailed_rows(joined_rows, stage2_context, output_dir)
    summary_rows = _build_summary_rows(summary, joined_rows, audit_queues)
    queue_rows = _flatten_audit_queues(audit_queues)
    run_manifest = _build_run_manifest(
        jsonl_path=jsonl_path,
        stage1_workbook=stage1_workbook,
        stage2_dir=stage2_dir,
        output_dir=output_dir,
        args=args,
        joined_rows=joined_rows,
    )
    payload = {
        "stage3": {
            "detailed": detailed_rows,
            "summary": summary_rows,
            "duplicate_pairs": duplicate_rows,
            "relationship_rules": relationship_rows,
            "audit_queues": queue_rows,
            "test_columns": test_columns,
            "dataset_summary": summary,
            "run_manifest": run_manifest,
        }
    }

    _write_json(output_dir / "dataset_summary.json", summary)
    _write_json(output_dir / "dataset_analysis.json", payload)
    _write_json(output_dir / "audit_queues.json", audit_queues)
    _write_json(output_dir / "relationship_analysis.json", {"rules": relationship_rows})
    _write_json(output_dir / "run_manifest.json", run_manifest)
    _write_jsonl(output_dir / "enriched_samples.jsonl", detailed_rows)
    _write_csv(output_dir / "enriched_samples.csv", detailed_rows)
    _write_csv(output_dir / "detailed_rows.csv", detailed_rows)
    _write_csv(output_dir / "duplicate_pairs.csv", duplicate_rows)
    _write_markdown(output_dir / "summary.md", summary, audit_queues, relationship_rows)
    _write_workbook(
        output_dir / "dataset_analysis.xlsx",
        summary_rows=summary_rows,
        detailed_rows=detailed_rows,
        test_columns=test_columns,
    )
    workbook_path, json_path = export_combined_report(
        stage1_workbook_path=stage1_workbook,
        stage2_output_root=stage2_dir,
        stage3_output_root=output_dir,
    )

    print(f"Wrote dataset summary to {output_dir / 'dataset_summary.json'}")
    print(f"Wrote Stage 3 payload to {output_dir / 'dataset_analysis.json'}")
    print(f"Wrote enriched samples to {output_dir / 'enriched_samples.csv'}")
    print(f"Wrote audit queues to {output_dir / 'audit_queues.json'}")
    print(f"Wrote narrative summary to {output_dir / 'summary.md'}")
    print(f"Wrote workbook to {output_dir / 'dataset_analysis.xlsx'}")
    print(f"Wrote combined workbook to {workbook_path}")
    print(f"Wrote combined json to {json_path}")
    return 0


def _build_joined_rows(
    *,
    jsonl_path: Path,
    stage1_rows: dict[str, dict[int, dict[str, Any]]],
    stage2_context: dict[str, Any],
    output_dir: Path,
    limit: int | None,
    offset: int,
    indices: str | None,
    near_duplicate_threshold: float,
    template_threshold: float,
    cluster_threshold: float,
) -> list[dict[str, Any]]:
    sample_records = load_samples(jsonl_path, limit=limit, offset=offset, indices=indices)
    sample_indices = [sample.index for sample in sample_records]
    attempt_rows_by_sample = stage2_context["attempt_rows_by_sample"]
    model_rows_by_sample = stage2_context["model_rows_by_sample"]

    joined_rows: list[dict[str, Any]] = []
    for sample in sample_records:
        stage1_detailed = stage1_rows["detailed"].get(sample.index, {})
        stage1_aggregate = stage1_rows["aggregate"].get(sample.index, {})
        stage2_sample = stage2_context["sample_rows"].get(sample.index, {})
        stage2_model = stage2_context["model_rows"].get(sample.index, {})
        stage2_models = model_rows_by_sample.get(sample.index, [])
        attempt_rows = attempt_rows_by_sample.get(sample.index, [])
        per_test_rows = stage2_context["per_test_rows"].get(sample.index, [])

        status_counts = _count_statuses(stage1_detailed)
        section_scores = _stage1_section_scores(stage1_detailed)
        stage1_score = _stage1_score(status_counts)
        failure_counter = Counter(
            str(row.get("failure_type", "")).strip()
            for row in per_test_rows
            if str(row.get("source", "")).startswith("model_candidate") and str(row.get("failure_type", "")).strip()
        )
        dominant_failure_type = failure_counter.most_common(1)[0][0] if failure_counter else ""
        attempt_features = _attempt_variance_features(attempt_rows)
        disagreement_features = _model_disagreement_features(stage2_models)
        threshold_features = _threshold_sensitivity_features(
            combined_pass_rate=_to_float(
                stage2_model.get("BestCombinedPassRate", stage2_sample.get("ModelABestCombinedPassRate"))
            ),
            efficacy_label=str(stage2_model.get("EfficacyLabel", stage2_sample.get("ModelAEfficacyLabel", ""))).strip(),
            suspicious=_to_bool(stage2_sample.get("Suspicious", stage2_model.get("Suspicious"))),
        )
        starter_signature = _starter_signature(sample.starter_code)
        title_text = str(sample.row.get("question_title", ""))
        embedding_text = _embedding_text(sample)

        row = {
            "Index": sample.index,
            "QuestionId": sample.row.get("question_id", ""),
            "QuestionTitle": sample.row.get("question_title", ""),
            "Difficulty": sample.row.get("difficulty", ""),
            "Source": sample.metadata.get("source", ""),
            "Language": sample.metadata.get("lang", ""),
            "FunctionName": sample.function_name,
            "PromptChars": len(sample.question_content),
            "PromptWords": _word_count(sample.question_content),
            "StarterCodeChars": len(sample.starter_code),
            "StarterCodeLines": _line_count(sample.starter_code),
            "IdealResponseChars": len(sample.ideal_response),
            "IdealResponseLines": _line_count(sample.ideal_response),
            "PublicTests": len(sample.public_tests),
            "PrivateTests": len(sample.private_tests),
            "TotalTests": len(sample.public_tests) + len(sample.private_tests),
            "PromptUsability": str(stage1_aggregate.get("Prompt", "")).strip(),
            "IdealResponseUsability": str(stage1_aggregate.get("Ideal_Response", "")).strip(),
            "TestCasesUsability": str(stage1_aggregate.get("Test_Cases", "")).strip(),
            "Stage1PassCount": status_counts["PASS"],
            "Stage1PartialCount": status_counts["PARTIAL"],
            "Stage1FailCount": status_counts["FAIL"],
            "Stage1UnclearCount": status_counts["UNCLEAR"],
            "Stage1NACount": status_counts["NA"],
            "Stage1Score": stage1_score,
            "Stage1CriticalFailCount": _count_critical_fails(stage1_detailed),
            "Stage1Flags": "; ".join(_top_failures(stage1_detailed)),
            "_Stage1FailureFlags": _all_failure_flags(stage1_detailed),
            "Stage1Section1Score": section_scores.get("1", 0.0),
            "Stage1Section2Score": section_scores.get("2", 0.0),
            "Stage1Section3Score": section_scores.get("3", 0.0),
            "Stage1Section4Score": section_scores.get("4", 0.0),
            "Stage1Section6Score": section_scores.get("6", 0.0),
            "Stage1Section7Score": section_scores.get("7", 0.0),
            "OraclePassRate": _to_float(stage2_sample.get("OraclePassRate", stage2_model.get("OraclePassRate"))),
            "WinnerCombinedPassRate": _to_float(
                stage2_model.get("BestCombinedPassRate", stage2_sample.get("ModelABestCombinedPassRate"))
            ),
            "WinnerProvidedPassRate": _to_float(
                stage2_model.get("BestProvidedPassRate", stage2_sample.get("ModelABestProvidedPassRate"))
            ),
            "WinnerGeneratedPassRate": _to_float(
                stage2_model.get("BestGeneratedPassRate", stage2_sample.get("ModelABestGeneratedPassRate"))
            ),
            "GeneratedTests": _to_int(stage2_sample.get("GeneratedTests", stage2_model.get("GeneratedTests"))),
            "EfficacyLabel": str(stage2_model.get("EfficacyLabel", stage2_sample.get("ModelAEfficacyLabel", ""))).strip(),
            "BenchmarkQualitySignal": str(
                stage2_model.get(
                    "BenchmarkQualitySignal",
                    stage2_sample.get("ModelABenchmarkQualitySignal", ""),
                )
            ).strip(),
            "Stage2FailureCategory": str(stage2_model.get("FailureCategory", "")).strip(),
            "DominantFailureType": dominant_failure_type,
            "Suspicious": _to_bool(stage2_sample.get("Suspicious", stage2_model.get("Suspicious"))),
            "NeedsAudit": _to_bool(stage2_sample.get("NeedsAudit", stage2_model.get("NeedsAudit"))),
            "ComparisonNote": str(stage2_sample.get("ComparisonNote", "")).strip(),
            "FailureTypes": json.dumps(dict(failure_counter), ensure_ascii=False, sort_keys=True),
            "AttemptCount": attempt_features["attempt_count"],
            "AttemptCombinedPassAverage": attempt_features["average"],
            "AttemptCombinedPassBest": attempt_features["best"],
            "AttemptCombinedPassWorst": attempt_features["worst"],
            "AttemptCombinedPassRange": attempt_features["range"],
            "AttemptCombinedPassStdDev": attempt_features["std_dev"],
            "AttemptExecutionSuccessRate": attempt_features["execution_success_rate"],
            "AttemptVarianceLabel": attempt_features["label"],
            "ModelCount": disagreement_features["model_count"],
            "ModelBestCombinedPassGap": disagreement_features["best_gap"],
            "ModelEfficacyDisagreement": disagreement_features["efficacy_disagreement"],
            "ModelBenchmarkSignalDisagreement": disagreement_features["benchmark_signal_disagreement"],
            "ModelSuspicionDisagreement": disagreement_features["suspicion_disagreement"],
            "ModelDisagreementLabel": disagreement_features["label"],
            "ThresholdNearestBoundary": threshold_features["nearest_boundary"],
            "ThresholdDistance": threshold_features["distance"],
            "ThresholdSensitivityLabel": threshold_features["label"],
            "ThresholdSensitivityReason": threshold_features["reason"],
            "_prompt_text": sample.question_content,
            "_prompt_ngrams": _char_ngrams(_normalize_text(sample.question_content)),
            "_template_ngrams": _char_ngrams(_template_text(sample.question_content)),
            "_test_signature": _test_signature(sample),
            "_starter_signature": starter_signature,
            "_title_ngrams": _char_ngrams(_normalize_text(title_text)),
            "_function_signature": _function_signature(sample.function_name),
            "_embedding_text": embedding_text,
        }
        joined_rows.append(row)

    embedding_metadata = _attach_embedding_vectors(joined_rows, output_dir=output_dir)
    pair_rows = _similarity_pairs(
        joined_rows,
        near_duplicate_threshold=near_duplicate_threshold,
        template_threshold=template_threshold,
    )
    cluster_assignments, cluster_sizes = _cluster_pairs(pair_rows, sample_indices, cluster_threshold)
    duplicate_details = _duplicate_detail_maps(pair_rows)
    outlier_flags = _compute_outlier_flags(joined_rows)

    final_rows: list[dict[str, Any]] = []
    for row in joined_rows:
        duplicate_detail = duplicate_details.get(row["Index"], {})
        flags = outlier_flags.get(row["Index"], [])
        redundancy_score = float(duplicate_detail.get("max_similarity", 0.0))
        contradiction = _contradiction_label(row)
        final_rows.append(
            {
                **{k: v for k, v in row.items() if not k.startswith("_")},
                "RedundancyScore": round(redundancy_score, 4),
                "DuplicateLabel": duplicate_detail.get("max_label", ""),
                "ClosestDuplicateIndex": duplicate_detail.get("closest_index"),
                "ClosestDuplicateSimilarity": duplicate_detail.get("max_similarity"),
                "ClosestDuplicateLexicalSimilarity": duplicate_detail.get("max_lexical_similarity"),
                "ClosestDuplicateEmbeddingSimilarity": duplicate_detail.get("max_embedding_similarity"),
                "RedundancyClusterId": cluster_assignments.get(row["Index"], row["Index"]),
                "RedundancyClusterSize": cluster_sizes.get(cluster_assignments.get(row["Index"], row["Index"]), 1),
                "OutlierFlags": "; ".join(flags),
                "ContradictionLabel": contradiction,
                "AuditPriority": _audit_priority(row, redundancy_score, flags, contradiction),
                "ExemplarCandidate": _is_exemplar_candidate(row, redundancy_score, flags),
            }
        )

    for row in final_rows:
        row["EmbeddingModel"] = embedding_metadata.get("model", "")
        row["EmbeddingCacheHits"] = embedding_metadata.get("cache_hits", 0)
        row["EmbeddingCacheMisses"] = embedding_metadata.get("cache_misses", 0)
        row["EmbeddingAvailable"] = bool(embedding_metadata.get("available"))

    stage3_judgments = _stage3_auditor_disagreement(final_rows, output_dir=output_dir)
    for row in final_rows:
        judgment = stage3_judgments.get(int(row["Index"]), {})
        if row.get("ModelCount", 0) <= 1 and judgment:
            row["ModelCount"] = judgment.get("model_count", row.get("ModelCount", 1))
            row["ModelBestCombinedPassGap"] = judgment.get("gap_proxy", row.get("ModelBestCombinedPassGap", 0.0))
            row["ModelEfficacyDisagreement"] = judgment.get("utility_disagreement", False)
            row["ModelBenchmarkSignalDisagreement"] = judgment.get("risk_disagreement", False)
            row["ModelSuspicionDisagreement"] = judgment.get("priority_disagreement", False)
            row["ModelDisagreementLabel"] = judgment.get("label", row.get("ModelDisagreementLabel", "single_model"))
            row["ModelDisagreementSource"] = "stage3_auditors"
            row["Stage3AuditorUtilityLabels"] = judgment.get("utility_labels", "")
            row["Stage3AuditorPrimaryRisks"] = judgment.get("primary_risks", "")
            row["Stage3AuditorPriorities"] = judgment.get("audit_priorities", "")
            row["AuditPriority"] = _escalate_priority(row["AuditPriority"], judgment.get("recommended_priority"))
        else:
            row["ModelDisagreementSource"] = "stage2_models" if row.get("ModelCount", 0) > 1 else "single_model"
            row["Stage3AuditorUtilityLabels"] = ""
            row["Stage3AuditorPrimaryRisks"] = ""
            row["Stage3AuditorPriorities"] = ""

    return sorted(final_rows, key=lambda item: item["Index"])


def _build_summary(joined_rows: list[dict[str, Any]], stage2_context: dict[str, Any]) -> dict[str, Any]:
    if not joined_rows:
        return {"dataset_size": 0}

    lengths = {
        "prompt_words": _numeric_summary([row["PromptWords"] for row in joined_rows]),
        "starter_code_lines": _numeric_summary([row["StarterCodeLines"] for row in joined_rows]),
        "ideal_response_lines": _numeric_summary([row["IdealResponseLines"] for row in joined_rows]),
        "public_tests": _numeric_summary([row["PublicTests"] for row in joined_rows]),
        "private_tests": _numeric_summary([row["PrivateTests"] for row in joined_rows]),
        "total_tests": _numeric_summary([row["TotalTests"] for row in joined_rows]),
        "stage1_score": _numeric_summary([row["Stage1Score"] for row in joined_rows]),
        "winner_combined_pass_rate": _numeric_summary([row["WinnerCombinedPassRate"] for row in joined_rows]),
        "oracle_pass_rate": _numeric_summary([row["OraclePassRate"] for row in joined_rows]),
        "attempt_combined_pass_range": _numeric_summary([row["AttemptCombinedPassRange"] for row in joined_rows]),
        "model_disagreement_gap": _numeric_summary([row["ModelBestCombinedPassGap"] for row in joined_rows]),
        "threshold_distance": _numeric_summary([row["ThresholdDistance"] for row in joined_rows]),
    }
    duplicate_counts = Counter(row["DuplicateLabel"] for row in joined_rows if row["DuplicateLabel"])
    contradiction_counts = Counter(row["ContradictionLabel"] for row in joined_rows if row["ContradictionLabel"])
    outlier_counter = Counter()
    for row in joined_rows:
        for flag in _split_flags(row["OutlierFlags"]):
            outlier_counter[flag] += 1

    stage1_status_counter = Counter()
    for field in ("PromptUsability", "IdealResponseUsability", "TestCasesUsability"):
        for row in joined_rows:
            value = str(row[field]).strip()
            if value:
                stage1_status_counter[f"{field}:{value}"] += 1

    stage2_efficacy = Counter(row["EfficacyLabel"] for row in joined_rows if row["EfficacyLabel"])
    benchmark_quality = Counter(
        row["BenchmarkQualitySignal"] for row in joined_rows if row["BenchmarkQualitySignal"]
    )
    attempt_variance = Counter(row["AttemptVarianceLabel"] for row in joined_rows if row["AttemptVarianceLabel"])
    model_disagreement = Counter(row["ModelDisagreementLabel"] for row in joined_rows if row["ModelDisagreementLabel"])
    threshold_sensitivity = Counter(
        row["ThresholdSensitivityLabel"] for row in joined_rows if row["ThresholdSensitivityLabel"]
    )
    failure_categories = Counter(
        row["Stage2FailureCategory"] for row in joined_rows if row["Stage2FailureCategory"]
    )
    difficulty_counter = Counter(row["Difficulty"] for row in joined_rows if row["Difficulty"])
    title_counter = Counter(row["QuestionTitle"] for row in joined_rows if row["QuestionTitle"])

    relationship_rows = _build_relationship_rows(joined_rows)
    top_relationship = relationship_rows[0] if relationship_rows else {}

    return {
        "dataset_size": len(joined_rows),
        "sample_indices": [row["Index"] for row in joined_rows],
        "counts": {
            "by_difficulty": dict(difficulty_counter),
            "by_question_title": dict(title_counter.most_common(10)),
            "stage1_labels": dict(stage1_status_counter),
            "stage2_efficacy": dict(stage2_efficacy),
            "stage2_benchmark_quality": dict(benchmark_quality),
            "stage2_failure_categories": dict(failure_categories),
            "attempt_variance": dict(attempt_variance),
            "model_disagreement": dict(model_disagreement),
            "threshold_sensitivity": dict(threshold_sensitivity),
            "duplicate_labels": dict(duplicate_counts),
            "contradictions": dict(contradiction_counts),
            "outlier_flags": dict(outlier_counter),
        },
        "distributions": lengths,
        "headline_findings": {
            "top_risks": _top_risks(joined_rows, duplicate_counts, contradiction_counts, benchmark_quality),
            "strongest_stage1_to_stage2_signal": top_relationship,
            "recommendations": _recommendations(joined_rows, top_relationship, duplicate_counts),
        },
        "manifest": stage2_context["manifest"],
        "embedding": {
            "available": any(bool(row.get("EmbeddingAvailable")) for row in joined_rows),
            "model": next((str(row.get("EmbeddingModel", "")) for row in joined_rows if row.get("EmbeddingModel")), ""),
        },
    }


def _build_audit_queues(joined_rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    def entry(row: dict[str, Any], reason: str) -> dict[str, Any]:
        return {
            "Index": row["Index"],
            "QuestionId": row["QuestionId"],
            "QuestionTitle": row["QuestionTitle"],
            "Difficulty": row["Difficulty"],
            "Reason": reason,
            "AuditPriority": row["AuditPriority"],
            "DatasetUtilityLabel": _dataset_utility_label(row),
            "RedundancyScore": row["RedundancyScore"],
            "ClosestDuplicateIndex": row["ClosestDuplicateIndex"],
            "ContradictionLabel": row["ContradictionLabel"],
            "OutlierFlags": row["OutlierFlags"],
        }

    benchmark_defects = sorted(
        (
            entry(
                row,
                _benchmark_defect_reason(row),
            )
            for row in joined_rows
            if row["Suspicious"] or row["OraclePassRate"] < 0.9999
        ),
        key=lambda item: (
            item["AuditPriority"] != "critical",
            item["Reason"],
            item["Index"],
        ),
    )[:15]

    redundancy = sorted(
        (
            entry(
                row,
                f"{row['DuplicateLabel']} vs {row['ClosestDuplicateIndex']} at {row['ClosestDuplicateSimilarity']}",
            )
            for row in joined_rows
            if row["RedundancyScore"] >= 0.55
        ),
        key=lambda item: (-item["RedundancyScore"], item["Index"]),
    )[:15]

    contradictions = sorted(
        (
            entry(row, row["ContradictionLabel"])
            for row in joined_rows
            if row["ContradictionLabel"]
        ),
        key=lambda item: (
            item["ContradictionLabel"] != "high_static_low_dynamic",
            item["Reason"],
            item["Index"],
        ),
        reverse=True,
    )[:15]

    trivial = sorted(
        (
            entry(row, "near-saturated pass rate with otherwise clean Stage 1 signals")
            for row in joined_rows
            if row["EfficacyLabel"] == "Low Efficacy" or row["WinnerCombinedPassRate"] >= 0.95
        ),
        key=lambda item: (item["AuditPriority"], item["Index"]),
    )[:15]

    exemplars = sorted(
        (
            entry(row, "clean, interpretable, non-redundant evaluation candidate")
            for row in joined_rows
            if row["ExemplarCandidate"]
        ),
        key=lambda item: (item["AuditPriority"], item["Index"]),
    )[:15]

    return {
        "benchmark_defect_candidates": benchmark_defects,
        "redundancy_candidates": redundancy,
        "contradictory_candidates": contradictions,
        "trivial_candidates": trivial,
        "exemplar_candidates": exemplars,
    }


def _build_relationship_rows(joined_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if len(joined_rows) < RELATIONSHIP_MIN_DATASET_SIZE:
        return []
    relationship_rows: list[dict[str, Any]] = []
    minimum_support = max(RELATIONSHIP_MIN_SUPPORT, math.ceil(len(joined_rows) * RELATIONSHIP_MIN_SUPPORT_SHARE))

    for signal in _stage1_check_names_from_rows(joined_rows):
        fail_group = [row for row in joined_rows if signal in _failure_flags_for_relationships(row)]
        pass_group = [row for row in joined_rows if signal not in _failure_flags_for_relationships(row)]
        if len(fail_group) < minimum_support or not pass_group:
            continue
        average_pass_rate = statistics.mean(row["WinnerCombinedPassRate"] for row in fail_group)
        suspicious_rate = sum(1 for row in fail_group if row["Suspicious"]) / len(fail_group)
        pass_rate_delta = (
            statistics.mean(row["WinnerCombinedPassRate"] for row in fail_group)
            - statistics.mean(row["WinnerCombinedPassRate"] for row in pass_group)
        )
        suspicious_lift = (
            sum(1 for row in fail_group if row["Suspicious"]) / len(fail_group)
        ) - (
            sum(1 for row in pass_group if row["Suspicious"]) / len(pass_group)
        )
        support_share = len(fail_group) / len(joined_rows)
        relationship_rows.append(
            {
                "Signal": signal,
                "Group": "failing_check",
                "SampleCount": len(fail_group),
                "SupportShare": round(support_share, 4),
                "AveragePassRate": round(average_pass_rate, 4),
                "SuspiciousRate": round(suspicious_rate, 4),
                "PassRateDeltaVsOthers": round(pass_rate_delta, 4),
                "SuspiciousLiftVsOthers": round(suspicious_lift, 4),
                "AdjustedStrength": round(_relationship_strength(pass_rate_delta, suspicious_lift, support_share), 4),
                "Reliability": _relationship_reliability(len(fail_group), support_share),
            }
        )

    for signal in (
        "Stage1Section1Score",
        "Stage1Section2Score",
        "Stage1Section3Score",
        "Stage1Section4Score",
        "Stage1Section6Score",
        "Stage1Section7Score",
        "Stage1Score",
        "Stage1CriticalFailCount",
    ):
        values = [row[signal] for row in joined_rows]
        if len({value for value in values}) <= 1:
            continue
        relationship_rows.append(
            {
                "Signal": signal,
                "Group": "section_score",
                "SampleCount": len(values),
                "SupportShare": 1.0,
                "AveragePassRate": round(statistics.mean(row["WinnerCombinedPassRate"] for row in joined_rows), 4),
                "SuspiciousRate": round(
                    sum(1 for row in joined_rows if row["Suspicious"]) / len(joined_rows), 4
                ),
                "PassRateDeltaVsOthers": round(_pearson(values, [row["WinnerCombinedPassRate"] for row in joined_rows]), 4),
                "SuspiciousLiftVsOthers": round(
                    _pearson(values, [1.0 if row["Suspicious"] else 0.0 for row in joined_rows]),
                    4,
                ),
                "AdjustedStrength": round(
                    _relationship_strength(
                        _pearson(values, [row["WinnerCombinedPassRate"] for row in joined_rows]),
                        _pearson(values, [1.0 if row["Suspicious"] else 0.0 for row in joined_rows]),
                        1.0,
                    ),
                    4,
                ),
                "Reliability": "broad",
            }
        )

    relationship_rows.sort(
        key=lambda row: (
            -_to_float(row.get("AdjustedStrength")),
            -row["SampleCount"],
        )
    )
    return relationship_rows[:20]


def _build_duplicate_rows(joined_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for left, right in combinations(joined_rows, 2):
        if (
            left["ClosestDuplicateIndex"] != right["Index"]
            and right["ClosestDuplicateIndex"] != left["Index"]
        ):
            continue
        similarity = round(max(left["RedundancyScore"], right["RedundancyScore"]), 4)
        if similarity < 0.55:
            continue
        rows.append(
            {
                "LeftIndex": left["Index"],
                "RightIndex": right["Index"],
                "LeftQuestionId": left["QuestionId"],
                "RightQuestionId": right["QuestionId"],
                "LeftTitle": left["QuestionTitle"],
                "RightTitle": right["QuestionTitle"],
                "CombinedSimilarity": similarity,
                "LexicalSimilarity": round(
                    max(
                        _to_float(left.get("ClosestDuplicateLexicalSimilarity")),
                        _to_float(right.get("ClosestDuplicateLexicalSimilarity")),
                    ),
                    4,
                ),
                "EmbeddingSimilarity": round(
                    max(
                        _to_float(left.get("ClosestDuplicateEmbeddingSimilarity")),
                        _to_float(right.get("ClosestDuplicateEmbeddingSimilarity")),
                    ),
                    4,
                ),
                "PairLabel": _pair_label_from_rows(left, right),
            }
        )
    return sorted(rows, key=lambda row: (-row["CombinedSimilarity"], row["LeftIndex"], row["RightIndex"]))


def _build_stage3_detailed_rows(
    joined_rows: list[dict[str, Any]],
    stage2_context: dict[str, Any],
    output_dir: Path,
) -> tuple[list[dict[str, Any]], list[str]]:
    best_model_rows = stage2_context["best_model_rows"]
    note_cache_path = output_dir / "detailed_test_notes.json"
    note_requests: list[dict[str, Any]] = []
    for aggregate_row in joined_rows:
        request = _build_stage3_note_request(aggregate_row)
        if request:
            note_requests.append(request)
    notes_by_request, missing_note_requests = split_cached_note_requests(
        "dataset-analysis",
        note_requests,
        note_cache_path,
    )
    if missing_note_requests:
        fresh_notes = generate_notes_for_requests(
            "dataset-analysis",
            missing_note_requests,
            max_concurrency=8,
        )
        notes_by_request.update(fresh_notes)
        update_note_cache(
            "dataset-analysis",
            missing_note_requests,
            fresh_notes,
            note_cache_path,
        )

    detailed_rows: list[dict[str, Any]] = []
    for aggregate_row in joined_rows:
        sample_index = aggregate_row["Index"]
        best_model_row = best_model_rows.get(sample_index, {})
        row_notes = notes_by_request.get(_stage3_note_request_id(sample_index), {})
        detailed_row = {
            "Index": aggregate_row["Index"],
            "QuestionId": aggregate_row["QuestionId"],
            "QuestionTitle": aggregate_row["QuestionTitle"],
            "Difficulty": aggregate_row["Difficulty"],
            "FunctionName": aggregate_row["FunctionName"],
            "PromptLengthBand": _length_band(aggregate_row["PromptWords"], "prompt"),
            "StarterCodeLengthBand": _length_band(aggregate_row["StarterCodeLines"], "starter"),
            "IdealResponseLengthBand": _length_band(aggregate_row["IdealResponseLines"], "ideal"),
            "TestCountBand": _test_count_band(aggregate_row["TotalTests"]),
            "PromptLengthOutlier": _flag_value(aggregate_row["OutlierFlags"], "prompt_length_outlier"),
            "StarterCodeOutlier": _flag_value(aggregate_row["OutlierFlags"], "starter_code_outlier"),
            "IdealResponseOutlier": _flag_value(aggregate_row["OutlierFlags"], "ideal_response_outlier"),
            "TestCountOutlier": _flag_value(aggregate_row["OutlierFlags"], "test_count_outlier"),
            "PerformanceOutlier": _flag_value(aggregate_row["OutlierFlags"], "pass_rate_outlier"),
            "RedundancyStatus": "FLAG" if aggregate_row["RedundancyScore"] >= 0.55 else "PASS",
            "RedundancyLabel": aggregate_row["DuplicateLabel"] or "unique",
            "ClosestNeighborIndex": aggregate_row["ClosestDuplicateIndex"] or "",
            "ClosestNeighborSimilarity": aggregate_row["ClosestDuplicateSimilarity"] or "",
            "ClosestNeighborLexicalSimilarity": aggregate_row["ClosestDuplicateLexicalSimilarity"] or "",
            "ClosestNeighborEmbeddingSimilarity": aggregate_row["ClosestDuplicateEmbeddingSimilarity"] or "",
            "RedundancyClusterId": aggregate_row["RedundancyClusterId"],
            "RedundancyClusterSize": aggregate_row["RedundancyClusterSize"],
            "AttemptVarianceCheck": "FLAG" if aggregate_row["AttemptVarianceLabel"] != "stable" else "PASS",
            "AttemptVarianceLabel": aggregate_row["AttemptVarianceLabel"],
            "AttemptCombinedPassRange": aggregate_row["AttemptCombinedPassRange"],
            "AttemptCombinedPassStdDev": aggregate_row["AttemptCombinedPassStdDev"],
            "AttemptExecutionSuccessRate": aggregate_row["AttemptExecutionSuccessRate"],
            "ModelDisagreementCheck": "FLAG" if aggregate_row["ModelDisagreementLabel"] not in {"aligned", "single_model"} else "PASS",
            "ModelDisagreementLabel": aggregate_row["ModelDisagreementLabel"],
            "ModelDisagreementSource": aggregate_row.get("ModelDisagreementSource", ""),
            "ModelCount": aggregate_row["ModelCount"],
            "ModelBestCombinedPassGap": aggregate_row["ModelBestCombinedPassGap"],
            "ThresholdSensitivityCheck": "FLAG" if aggregate_row["ThresholdSensitivityLabel"] != "stable" else "PASS",
            "ThresholdSensitivityLabel": aggregate_row["ThresholdSensitivityLabel"],
            "ThresholdNearestBoundary": aggregate_row["ThresholdNearestBoundary"],
            "ThresholdDistance": aggregate_row["ThresholdDistance"],
            "DifficultySignalRegime": _difficulty_signal_regime(aggregate_row),
            "ContradictionCheck": aggregate_row["ContradictionLabel"] or "none",
            "BenchmarkDefectCandidate": "FLAG" if aggregate_row["AuditPriority"] == "critical" and _is_benchmark_defect_candidate(aggregate_row) else "PASS",
            "TrivialityCheck": "FLAG" if _is_trivial_candidate(aggregate_row) else "PASS",
            "ExemplarCheck": "FLAG" if aggregate_row["ExemplarCandidate"] else "PASS",
            "AuditPriority": aggregate_row["AuditPriority"],
            "AuditReason": _audit_reason_summary(aggregate_row),
        }
        ordered_row: dict[str, Any] = {}
        leading_columns = [
            "Index",
            "QuestionId",
            "QuestionTitle",
            "Difficulty",
            "FunctionName",
            "PromptLengthBand",
            "StarterCodeLengthBand",
            "IdealResponseLengthBand",
            "TestCountBand",
            "RedundancyLabel",
            "ClosestNeighborIndex",
            "ClosestNeighborSimilarity",
            "ClosestNeighborLexicalSimilarity",
            "ClosestNeighborEmbeddingSimilarity",
            "RedundancyClusterId",
            "RedundancyClusterSize",
        ]
        for column_name in leading_columns:
            ordered_row[column_name] = detailed_row[column_name]
        for column_name in STAGE3_TEST_COLUMNS:
            ordered_row[column_name] = detailed_row[column_name]
            if column_name in row_notes or column_name in _stage3_note_eligible_columns():
                ordered_row[f"Notes-{column_name}"] = row_notes.get(column_name, "")
        ordered_row["AuditPriority"] = detailed_row["AuditPriority"]
        ordered_row["AuditReason"] = detailed_row["AuditReason"]
        ordered_row["ModelDisagreementSource"] = detailed_row["ModelDisagreementSource"]
        detailed_rows.append(ordered_row)
    return detailed_rows, list(STAGE3_TEST_COLUMNS)


def _build_summary_rows(
    summary: dict[str, Any],
    joined_rows: list[dict[str, Any]],
    audit_queues: dict[str, list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    size = summary.get("dataset_size", 0)
    counts = summary.get("counts", {})
    headline = summary.get("headline_findings", {})
    distributions = summary.get("distributions", {})
    stage2_efficacy = counts.get("stage2_efficacy", {})
    attempt_variance = counts.get("attempt_variance", {})
    model_disagreement = counts.get("model_disagreement", {})
    threshold_sensitivity = counts.get("threshold_sensitivity", {})
    top_relationship = headline.get("strongest_stage1_to_stage2_signal", {})
    embedding_info = summary.get("embedding", {})
    redundancy_candidates = len(audit_queues.get("redundancy_candidates", []))
    defect_candidates = len(audit_queues.get("benchmark_defect_candidates", []))
    contradiction_candidates = len(audit_queues.get("contradictory_candidates", []))

    rows.append(
        {
            "AnalysisArea": "dataset_scope",
            "TestName": "dataset_size",
            "Result": "PASS" if size >= 100 else "REVIEW",
            "Evidence": f"{size} samples analyzed.",
            "Interpretation": "Large enough for dataset-level structure checks." if size >= 100 else "Small slice; conclusions may be unstable.",
            "ContextualCaveat": "",
            "Recommendation": "",
        }
    )
    rows.append(
        {
            "AnalysisArea": "difficulty_regime",
            "TestName": "efficacy_distribution",
            "Result": _difficulty_distribution_result(stage2_efficacy),
            "Evidence": _format_counter(stage2_efficacy),
            "Interpretation": _difficulty_distribution_interpretation(stage2_efficacy),
            "ContextualCaveat": f"{defect_candidates} benchmark-defect candidates may distort the apparent difficulty mix." if defect_candidates else "",
            "Recommendation": "Audit suspicious defect candidates before fully trusting the difficulty profile." if defect_candidates else "",
        }
    )
    rows.append(
        {
            "AnalysisArea": "redundancy",
            "TestName": "template_overlap_scan",
            "Result": "REVIEW" if redundancy_candidates else "PASS",
            "Evidence": f"{redundancy_candidates} redundancy candidates; top issue count from summary risk scan.",
            "Interpretation": "Repeated templates reduce information density." if redundancy_candidates else "No strong template recycling surfaced with the current redundancy scan.",
            "ContextualCaveat": (
                f"Embeddings enabled via {embedding_info.get('model')}."
                if embedding_info.get("available")
                else "Embeddings were unavailable, so the scan relied on lexical and structural signals only."
            ),
            "Recommendation": "Deduplicate or downweight repeated templates." if redundancy_candidates else "Consider stronger redundancy checks only if time allows.",
        }
    )
    rows.append(
        {
            "AnalysisArea": "stability",
            "TestName": "attempt_variance_scan",
            "Result": "REVIEW" if any(key != "stable" for key, value in attempt_variance.items() if value) else "PASS",
            "Evidence": _format_counter(attempt_variance),
            "Interpretation": "Large best-vs-worst swings suggest unstable benchmark signal across repeated attempts." if attempt_variance else "No repeated-attempt evidence available.",
            "ContextualCaveat": "Attempt variance is only visible when multiple attempts were run.",
            "Recommendation": "Prioritize unstable samples for Stage 4 and avoid over-trusting best-attempt views." if attempt_variance else "",
        }
    )
    rows.append(
        {
            "AnalysisArea": "stability",
            "TestName": "model_disagreement_scan",
            "Result": "REVIEW" if any(key not in {"aligned", "single_model"} for key, value in model_disagreement.items() if value) else "PASS",
            "Evidence": _format_counter(model_disagreement),
            "Interpretation": "Cross-model disagreement can reveal ambiguous or misleading benchmark items." if model_disagreement else "No model-level comparison evidence available.",
            "ContextualCaveat": "In single-model Stage 2 runs, this comes from a lightweight Stage 3 dual-auditor pass rather than a second full benchmark execution.",
            "Recommendation": "Audit samples with large model gaps before treating them as clean capability signal." if model_disagreement else "",
        }
    )
    rows.append(
        {
            "AnalysisArea": "stability",
            "TestName": "threshold_sensitivity_scan",
            "Result": "REVIEW" if any(key != "stable" for key, value in threshold_sensitivity.items() if value) else "PASS",
            "Evidence": _format_counter(threshold_sensitivity),
            "Interpretation": "Near-threshold items can flip interpretation under small heuristic changes." if threshold_sensitivity else "No near-threshold concentration detected.",
            "ContextualCaveat": "This tests the fragility of heuristic cutoffs, not benchmark correctness itself.",
            "Recommendation": "Send near-threshold items to manual review before treating regime counts as stable." if threshold_sensitivity else "",
        }
    )
    rows.append(
        {
            "AnalysisArea": "structural_health",
            "TestName": "length_and_test_count_outliers",
            "Result": "REVIEW" if counts.get("outlier_flags") else "PASS",
            "Evidence": _format_counter(counts.get("outlier_flags", {})),
            "Interpretation": "Outliers may represent broken, atypical, or unusually informative items." if counts.get("outlier_flags") else "No notable outlier concentration detected.",
            "ContextualCaveat": "",
            "Recommendation": "Inspect the outlier queue during Stage 4." if counts.get("outlier_flags") else "",
        }
    )
    rows.append(
        {
            "AnalysisArea": "cross_stage_caveat",
            "TestName": "high_static_low_dynamic_contradictions",
            "Result": "REVIEW" if contradiction_candidates else "PASS",
            "Evidence": f"{contradiction_candidates} contradiction candidates.",
            "Interpretation": "Some items may look well-formed but still fail to provide trustworthy evaluation signal." if contradiction_candidates else "No major contradiction concentration detected.",
            "ContextualCaveat": "This is contextual evidence from earlier stages, not a standalone Stage 3 structural test.",
            "Recommendation": "Send contradiction candidates to manual audit first." if contradiction_candidates else "",
        }
    )
    rows.append(
        {
            "AnalysisArea": "contextual_signal",
            "TestName": "strongest_upstream_caveat",
            "Result": "INFO",
            "Evidence": (
                f"{top_relationship.get('Signal', 'n/a')} "
                f"(support={top_relationship.get('SampleCount', 0)}, "
                f"support_share={top_relationship.get('SupportShare', 0)}, "
                f"reliability={top_relationship.get('Reliability', 'n/a')}, "
                f"suspicious_lift={top_relationship.get('SuspiciousLiftVsOthers', 0)})"
            ),
            "Interpretation": "Useful as a caveat when reading dataset-level findings, not as a replacement for Stage 3 structure tests.",
            "ContextualCaveat": "Relationship analysis is descriptive and can overemphasize sparse signals.",
            "Recommendation": "",
        }
    )
    for recommendation in headline.get("recommendations", []):
        rows.append(
            {
                "AnalysisArea": "recommendation",
                "TestName": "next_step",
                "Result": "ACTION",
                "Evidence": recommendation,
                "Interpretation": "",
                "ContextualCaveat": "",
                "Recommendation": recommendation,
            }
        )
    rows.append(
        {
            "AnalysisArea": "distribution_context",
            "TestName": "prompt_length_distribution",
            "Result": "INFO",
            "Evidence": _format_distribution(distributions.get("prompt_words", {})),
            "Interpretation": "Prompt length spread is descriptive context for coverage and outlier review.",
            "ContextualCaveat": "",
            "Recommendation": "",
        }
    )
    rows.append(
        {
            "AnalysisArea": "distribution_context",
            "TestName": "test_count_distribution",
            "Result": "INFO",
            "Evidence": _format_distribution(distributions.get("total_tests", {})),
            "Interpretation": "Test-count spread helps explain saturation, brittleness, and audit cost.",
            "ContextualCaveat": "",
            "Recommendation": "",
        }
    )
    return rows


def _flatten_audit_queues(audit_queues: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for queue_name, entries in audit_queues.items():
        for entry in entries:
            rows.append({"Queue": queue_name, **entry})
    return rows


def _stage3_auditor_disagreement(rows: list[dict[str, Any]], *, output_dir: Path) -> dict[int, dict[str, Any]]:
    if not rows:
        return {}
    task_specs: list[AsyncTaskSpec] = []
    for row in rows:
        for target_name in STAGE3_AUDIT_TARGETS:
            request_id = f"stage3-audit:{row['Index']}:{target_name}"
            task_specs.append(
                AsyncTaskSpec(
                    request_id=request_id,
                    task_type="json",
                    kwargs={
                        "target_name": target_name,
                        "schema_name": "stage3_auditor_judgment",
                        "schema": STAGE3_AUDITOR_SCHEMA,
                        "user_prompt": _build_stage3_auditor_prompt(row),
                        "trace_dir": output_dir / "stage3_auditor_traces" / f"sample_{row['Index']}" / target_name,
                    },
                )
            )
    try:
        raw_results = run_async_tasks_sync(task_specs, max_concurrency=4)
    except Exception:
        return {}
    by_sample: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for request_id, payload in raw_results.items():
        _, sample_index, _target_name = request_id.split(":", 2)
        by_sample[int(sample_index)].append(payload)

    summary: dict[int, dict[str, Any]] = {}
    for sample_index, payloads in by_sample.items():
        utility_labels = [str(payload.get("dataset_utility_label", "")).strip() for payload in payloads]
        primary_risks = [str(payload.get("primary_risk", "")).strip() for payload in payloads]
        audit_priorities = [str(payload.get("audit_priority", "")).strip() for payload in payloads]
        utility_disagreement = len(set(utility_labels)) > 1
        risk_disagreement = len(set(primary_risks)) > 1
        priority_disagreement = len(set(audit_priorities)) > 1
        disagreement_count = sum([utility_disagreement, risk_disagreement, priority_disagreement])
        label = "aligned"
        if disagreement_count >= 2 or priority_disagreement:
            label = "strong_disagreement"
        elif disagreement_count == 1:
            label = "moderate_disagreement"
        summary[sample_index] = {
            "model_count": len(payloads),
            "gap_proxy": round(disagreement_count / 3.0, 4),
            "utility_disagreement": utility_disagreement,
            "risk_disagreement": risk_disagreement,
            "priority_disagreement": priority_disagreement,
            "utility_labels": "; ".join(utility_labels),
            "primary_risks": "; ".join(primary_risks),
            "audit_priorities": "; ".join(audit_priorities),
            "recommended_priority": _highest_priority(audit_priorities),
            "label": label,
        }
    return summary


def _build_stage3_auditor_prompt(row: dict[str, Any]) -> str:
    return "\n".join(
        [
            "Judge this sample only as a dataset-audit item within a benchmark.",
            "Focus on benchmark utility, not coding quality.",
            "Return JSON only.",
            "",
            f"Index: {row['Index']}",
            f"Title: {row['QuestionTitle']}",
            f"Difficulty: {row['Difficulty']}",
            f"WinnerCombinedPassRate: {row['WinnerCombinedPassRate']}",
            f"OraclePassRate: {row['OraclePassRate']}",
            f"EfficacyLabel: {row['EfficacyLabel']}",
            f"Suspicious: {row['Suspicious']}",
            f"RedundancyScore: {row['RedundancyScore']}",
            f"DuplicateLabel: {row['DuplicateLabel']}",
            f"OutlierFlags: {row['OutlierFlags']}",
            f"AttemptVarianceLabel: {row['AttemptVarianceLabel']}",
            f"ThresholdSensitivityLabel: {row['ThresholdSensitivityLabel']}",
            f"ContradictionLabel: {row['ContradictionLabel']}",
            f"AuditPriority: {row['AuditPriority']}",
            "",
            "Choose:",
            "- dataset_utility_label: strong / usable / caveated / contradictory / saturated",
            "- primary_risk: none / redundancy / instability / benchmark_defect / threshold_fragility / outlier / contradiction",
            "- audit_priority: normal / medium / high / critical",
        ]
    )


def _highest_priority(priorities: list[str]) -> str:
    ordering = {"normal": 0, "medium": 1, "high": 2, "critical": 3}
    best = "normal"
    for priority in priorities:
        if ordering.get(priority, -1) > ordering.get(best, -1):
            best = priority
    return best


def _escalate_priority(current: str, requested: str | None) -> str:
    ordering = {"normal": 0, "medium": 1, "high": 2, "critical": 3}
    requested_value = ordering.get(str(requested or ""), -1)
    current_value = ordering.get(current, -1)
    if requested_value > current_value:
        return str(requested)
    return current


def _attempt_variance_features(attempt_rows: list[dict[str, Any]]) -> dict[str, Any]:
    if not attempt_rows:
        return {
            "attempt_count": 0,
            "average": 0.0,
            "best": 0.0,
            "worst": 0.0,
            "range": 0.0,
            "std_dev": 0.0,
            "execution_success_rate": 0.0,
            "label": "unknown",
        }
    rates = [_to_float(row.get("CombinedPassRate")) for row in attempt_rows]
    best = max(rates)
    worst = min(rates)
    average = statistics.mean(rates)
    std_dev = statistics.pstdev(rates) if len(rates) > 1 else 0.0
    execution_success_rate = sum(
        1 for row in attempt_rows if str(row.get("ExecutionProbeStatus", "")).strip() == "ok"
    ) / len(attempt_rows)
    label = "stable"
    if len(attempt_rows) <= 1:
        label = "stable"
    elif execution_success_rate < 1.0 or (best - worst) >= ATTEMPT_VARIANCE_RANGE_FLAG or std_dev >= ATTEMPT_VARIANCE_STD_FLAG:
        label = "volatile"
    elif (best - worst) >= 0.2 or std_dev >= 0.08:
        label = "moderate"
    return {
        "attempt_count": len(attempt_rows),
        "average": round(average, 4),
        "best": round(best, 4),
        "worst": round(worst, 4),
        "range": round(best - worst, 4),
        "std_dev": round(std_dev, 4),
        "execution_success_rate": round(execution_success_rate, 4),
        "label": label,
    }


def _model_disagreement_features(model_rows: list[dict[str, Any]]) -> dict[str, Any]:
    if not model_rows:
        return {
            "model_count": 0,
            "best_gap": 0.0,
            "efficacy_disagreement": False,
            "benchmark_signal_disagreement": False,
            "suspicion_disagreement": False,
            "label": "unknown",
        }
    if len(model_rows) == 1:
        return {
            "model_count": 1,
            "best_gap": 0.0,
            "efficacy_disagreement": False,
            "benchmark_signal_disagreement": False,
            "suspicion_disagreement": False,
            "label": "single_model",
        }
    best_rates = [_to_float(row.get("BestCombinedPassRate")) for row in model_rows]
    gap = max(best_rates) - min(best_rates)
    efficacy_values = {str(row.get("EfficacyLabel", "")).strip() for row in model_rows if str(row.get("EfficacyLabel", "")).strip()}
    benchmark_values = {
        str(row.get("BenchmarkQualitySignal", "")).strip()
        for row in model_rows
        if str(row.get("BenchmarkQualitySignal", "")).strip()
    }
    suspicion_values = {bool(_to_bool(row.get("Suspicious"))) for row in model_rows}
    efficacy_disagreement = len(efficacy_values) > 1
    benchmark_disagreement = len(benchmark_values) > 1
    suspicion_disagreement = len(suspicion_values) > 1
    label = "aligned"
    if gap >= MODEL_DISAGREEMENT_GAP_HIGH or suspicion_disagreement or (efficacy_disagreement and gap >= MODEL_DISAGREEMENT_GAP_FLAG):
        label = "strong_disagreement"
    elif gap >= MODEL_DISAGREEMENT_GAP_FLAG or efficacy_disagreement or benchmark_disagreement:
        label = "moderate_disagreement"
    return {
        "model_count": len(model_rows),
        "best_gap": round(gap, 4),
        "efficacy_disagreement": efficacy_disagreement,
        "benchmark_signal_disagreement": benchmark_disagreement,
        "suspicion_disagreement": suspicion_disagreement,
        "label": label,
    }


def _threshold_sensitivity_features(
    *,
    combined_pass_rate: float,
    efficacy_label: str,
    suspicious: bool,
) -> dict[str, Any]:
    boundaries = {
        "high_efficacy_floor": 0.25,
        "moderate_high_split": 0.8,
        "saturation_boundary": 0.95,
    }
    nearest_boundary = min(boundaries.items(), key=lambda item: abs(combined_pass_rate - item[1]))
    distance = abs(combined_pass_rate - nearest_boundary[1])
    label = "stable"
    reason = ""
    if suspicious:
        label = "caveated"
        reason = "suspicious benchmark behavior already dominates interpretation"
    elif distance <= THRESHOLD_SENSITIVITY_MARGIN / 2:
        label = "high"
        reason = f"pass rate sits very close to {nearest_boundary[0]}"
    elif distance <= THRESHOLD_SENSITIVITY_MARGIN:
        label = "moderate"
        reason = f"pass rate sits near {nearest_boundary[0]}"
    if efficacy_label == "Inconclusive":
        label = "caveated"
        reason = "upstream efficacy is already inconclusive"
    return {
        "nearest_boundary": nearest_boundary[0],
        "distance": round(distance, 4),
        "label": label,
        "reason": reason,
    }


def _starter_signature(text: str) -> set[str]:
    normalized = _normalize_text(text)
    return _char_ngrams(normalized) | set(_signature_tokens(normalized))


def _function_signature(function_name: str) -> set[str]:
    normalized = _normalize_text(function_name)
    return set(_signature_tokens(normalized)) | _char_ngrams(normalized)


def _signature_tokens(text: str) -> list[str]:
    return [token for token in re.split(r"[^a-z0-9_]+", text) if token]


def _embedding_text(sample) -> str:
    prompt = sample.question_content.strip()
    starter = sample.starter_code.strip()
    title = str(sample.row.get("question_title", "")).strip()
    function_name = sample.function_name.strip()
    parts = [part for part in [title, function_name, prompt[:2500], starter[:1000]] if part]
    return "\n".join(parts)


def _attach_embedding_vectors(rows: list[dict[str, Any]], *, output_dir: Path) -> dict[str, Any]:
    if not rows:
        return {"available": False, "model": DEFAULT_LOCAL_EMBEDDING_MODEL, "cache_hits": 0, "cache_misses": 0}
    cache_path = output_dir / "embedding_cache.json"
    texts = [str(row.get("_embedding_text", "")) for row in rows]
    try:
        response = embed_texts_cached_for_target(
            "local-qwen",
            texts,
            cache_path=cache_path,
            model_name=DEFAULT_LOCAL_EMBEDDING_MODEL,
            trace_dir=output_dir / "embedding_traces",
        )
    except Exception:
        for row in rows:
            row["_embedding_vector"] = []
        return {"available": False, "model": DEFAULT_LOCAL_EMBEDDING_MODEL, "cache_hits": 0, "cache_misses": len(rows)}
    vectors = response.get("vectors", [])
    for row, vector in zip(rows, vectors):
        row["_embedding_vector"] = vector
    return {
        "available": True,
        "model": str(response.get("model", DEFAULT_LOCAL_EMBEDDING_MODEL)),
        "cache_hits": int(response.get("cache_hits", 0)),
        "cache_misses": int(response.get("cache_misses", 0)),
    }


def _cosine_similarity(left: list[float], right: list[float]) -> float:
    if not left or not right or len(left) != len(right):
        return 0.0
    numerator = sum(left_value * right_value for left_value, right_value in zip(left, right))
    left_norm = math.sqrt(sum(value * value for value in left))
    right_norm = math.sqrt(sum(value * value for value in right))
    if left_norm == 0.0 or right_norm == 0.0:
        return 0.0
    return numerator / (left_norm * right_norm)


def _build_run_manifest(
    *,
    jsonl_path: Path,
    stage1_workbook: Path,
    stage2_dir: Path,
    output_dir: Path,
    args,
    joined_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "stage": "dataset-analysis",
        "jsonl_path": str(jsonl_path),
        "stage1_workbook": str(stage1_workbook),
        "stage2_dir": str(stage2_dir),
        "output_dir": str(output_dir),
        "sample_indices": [int(row["Index"]) for row in joined_rows],
        "question_ids": [str(row["QuestionId"]) for row in joined_rows],
        "limit": args.limit,
        "offset": args.offset,
        "indices": args.indices,
        "near_duplicate_threshold": args.near_duplicate_threshold,
        "template_threshold": args.template_threshold,
        "cluster_threshold": args.cluster_threshold,
        "threshold_sensitivity_margin": THRESHOLD_SENSITIVITY_MARGIN,
        "attempt_variance_range_flag": ATTEMPT_VARIANCE_RANGE_FLAG,
        "attempt_variance_std_flag": ATTEMPT_VARIANCE_STD_FLAG,
        "model_disagreement_gap_flag": MODEL_DISAGREEMENT_GAP_FLAG,
        "embedding_model": DEFAULT_LOCAL_EMBEDDING_MODEL,
    }


def _relationship_strength(pass_rate_delta: float, suspicious_lift: float, support_share: float) -> float:
    return (abs(pass_rate_delta) + 0.5 * abs(suspicious_lift)) * max(support_share, 0.05)


def _relationship_reliability(sample_count: int, support_share: float) -> str:
    if sample_count >= 20 or support_share >= 0.15:
        return "broad"
    if sample_count >= 10 or support_share >= 0.08:
        return "moderate"
    return "narrow"


def _stage3_note_eligible_columns() -> set[str]:
    return {
        "PromptLengthOutlier",
        "StarterCodeOutlier",
        "IdealResponseOutlier",
        "TestCountOutlier",
        "PerformanceOutlier",
        "RedundancyStatus",
        "AttemptVarianceCheck",
        "ModelDisagreementCheck",
        "ThresholdSensitivityCheck",
        "DifficultySignalRegime",
        "ContradictionCheck",
        "BenchmarkDefectCandidate",
        "TrivialityCheck",
        "ExemplarCheck",
    }


def _stage3_note_request_id(sample_index: int) -> str:
    return f"stage3:{sample_index}"


def _build_stage3_note_request(aggregate_row: dict[str, Any]) -> dict[str, Any] | None:
    candidate_lines: list[str] = []
    if _flag_value(aggregate_row["OutlierFlags"], "prompt_length_outlier") == "FLAG":
        candidate_lines.append(f"- PromptLengthOutlier: prompt_words={aggregate_row['PromptWords']}")
    if _flag_value(aggregate_row["OutlierFlags"], "starter_code_outlier") == "FLAG":
        candidate_lines.append(f"- StarterCodeOutlier: starter_code_lines={aggregate_row['StarterCodeLines']}")
    if _flag_value(aggregate_row["OutlierFlags"], "ideal_response_outlier") == "FLAG":
        candidate_lines.append(f"- IdealResponseOutlier: ideal_response_lines={aggregate_row['IdealResponseLines']}")
    if _flag_value(aggregate_row["OutlierFlags"], "test_count_outlier") == "FLAG":
        candidate_lines.append(
            f"- TestCountOutlier: public_tests={aggregate_row['PublicTests']}, private_tests={aggregate_row['PrivateTests']}, total_tests={aggregate_row['TotalTests']}"
        )
    if _flag_value(aggregate_row["OutlierFlags"], "pass_rate_outlier") == "FLAG":
        candidate_lines.append(
            f"- PerformanceOutlier: combined_pass_rate={aggregate_row['WinnerCombinedPassRate']}, efficacy_label={aggregate_row['EfficacyLabel']}"
        )
    if aggregate_row["RedundancyScore"] >= 0.55:
        candidate_lines.append(
            f"- RedundancyStatus: label={aggregate_row['DuplicateLabel']}, closest_index={aggregate_row['ClosestDuplicateIndex']}, similarity={aggregate_row['ClosestDuplicateSimilarity']}, lexical={aggregate_row['ClosestDuplicateLexicalSimilarity']}, embedding={aggregate_row['ClosestDuplicateEmbeddingSimilarity']}, cluster_size={aggregate_row['RedundancyClusterSize']}"
        )
    if aggregate_row["AttemptVarianceLabel"] != "stable":
        candidate_lines.append(
            f"- AttemptVarianceCheck: label={aggregate_row['AttemptVarianceLabel']}, attempt_count={aggregate_row['AttemptCount']}, pass_range={aggregate_row['AttemptCombinedPassRange']}, pass_std={aggregate_row['AttemptCombinedPassStdDev']}, execution_success_rate={aggregate_row['AttemptExecutionSuccessRate']}"
        )
    if aggregate_row["ModelDisagreementLabel"] not in {"aligned", "single_model"}:
        candidate_lines.append(
            f"- ModelDisagreementCheck: label={aggregate_row['ModelDisagreementLabel']}, model_count={aggregate_row['ModelCount']}, best_gap={aggregate_row['ModelBestCombinedPassGap']}, efficacy_disagreement={aggregate_row['ModelEfficacyDisagreement']}, suspicion_disagreement={aggregate_row['ModelSuspicionDisagreement']}"
        )
    if aggregate_row["ThresholdSensitivityLabel"] != "stable":
        candidate_lines.append(
            f"- ThresholdSensitivityCheck: label={aggregate_row['ThresholdSensitivityLabel']}, boundary={aggregate_row['ThresholdNearestBoundary']}, distance={aggregate_row['ThresholdDistance']}, reason={aggregate_row['ThresholdSensitivityReason']}"
        )
    difficulty_regime = _difficulty_signal_regime(aggregate_row)
    if difficulty_regime != "discriminative":
        candidate_lines.append(
            f"- DifficultySignalRegime: regime={difficulty_regime}, combined_pass_rate={aggregate_row['WinnerCombinedPassRate']}, efficacy_label={aggregate_row['EfficacyLabel']}, oracle_pass_rate={aggregate_row['OraclePassRate']}"
        )
    if aggregate_row["ContradictionLabel"]:
        candidate_lines.append(
            f"- ContradictionCheck: label={aggregate_row['ContradictionLabel']}, stage1_score={aggregate_row['Stage1Score']}, combined_pass_rate={aggregate_row['WinnerCombinedPassRate']}, suspicious={aggregate_row['Suspicious']}"
        )
    if _is_benchmark_defect_candidate(aggregate_row):
        candidate_lines.append(
            f"- BenchmarkDefectCandidate: oracle_pass_rate={aggregate_row['OraclePassRate']}, suspicious={aggregate_row['Suspicious']}, benchmark_quality={aggregate_row['BenchmarkQualitySignal']}"
        )
    if _is_trivial_candidate(aggregate_row):
        candidate_lines.append(
            f"- TrivialityCheck: combined_pass_rate={aggregate_row['WinnerCombinedPassRate']}, efficacy_label={aggregate_row['EfficacyLabel']}"
        )
    if aggregate_row["ExemplarCandidate"]:
        candidate_lines.append(
            f"- ExemplarCheck: stage1_score={aggregate_row['Stage1Score']}, efficacy_label={aggregate_row['EfficacyLabel']}, redundancy_score={aggregate_row['RedundancyScore']}"
        )
    if not candidate_lines:
        return None
    prompt = "\n".join(
        [
            "Write terse per-test notes for a Stage 3 dataset-audit row.",
            "Return notes only for the allowed Stage 3 columns listed below.",
            "Allowed columns: " + ", ".join(line.split(":")[0].replace("- ", "") for line in candidate_lines),
            "Explain just enough for a human auditor to know where to inspect next.",
            "",
            f"Sample index: {aggregate_row['Index']}",
            f"Question title: {aggregate_row['QuestionTitle']}",
            f"Difficulty: {aggregate_row['Difficulty']}",
            "",
            "Flagged Stage 3 checks:",
            *candidate_lines,
        ]
    )
    return {
        "request_id": _stage3_note_request_id(int(aggregate_row["Index"])),
        "prompt": prompt,
        "allowed_columns": [line.split(":")[0].replace("- ", "") for line in candidate_lines],
    }


def _length_band(value: int, dimension: str) -> str:
    if dimension == "prompt":
        if value < 150:
            return "short"
        if value <= 500:
            return "medium"
        return "long"
    if dimension == "starter":
        if value <= 5:
            return "minimal"
        if value <= 20:
            return "moderate"
        return "large"
    if value <= 80:
        return "compact"
    if value <= 250:
        return "moderate"
    return "large"


def _test_count_band(value: int) -> str:
    if value < 10:
        return "sparse"
    if value <= 20:
        return "recommended"
    return "heavy"


def _flag_value(raw_flags: str, flag: str) -> str:
    return "FLAG" if flag in _split_flags(raw_flags) else "PASS"


def _difficulty_signal_regime(row: dict[str, Any]) -> str:
    if row["Suspicious"]:
        return "caveated"
    if row.get("ThresholdSensitivityLabel") != "stable":
        return "threshold_fragile"
    if row.get("AttemptVarianceLabel") == "volatile" or row.get("ModelDisagreementLabel") == "strong_disagreement":
        return "unstable"
    if row["WinnerCombinedPassRate"] >= 0.95:
        return "saturated"
    if row["WinnerCombinedPassRate"] <= 0.1:
        return "extreme_or_blocked"
    return "discriminative"


def _is_benchmark_defect_candidate(row: dict[str, Any]) -> bool:
    return bool(row["Suspicious"] or row["OraclePassRate"] < 0.9999)


def _is_trivial_candidate(row: dict[str, Any]) -> bool:
    return bool(row["EfficacyLabel"] == "Low Efficacy" or row["WinnerCombinedPassRate"] >= 0.95)


def _dataset_utility_label(row: dict[str, Any]) -> str:
    if _is_benchmark_defect_candidate(row):
        return "caveated"
    if _is_trivial_candidate(row):
        return "saturated"
    if row["ContradictionLabel"]:
        return "contradictory"
    if row["ExemplarCandidate"]:
        return "strong"
    return "usable"


def _benchmark_defect_reason(row: dict[str, Any]) -> str:
    if row["OraclePassRate"] < 0.9999:
        return "oracle or benchmark artifact appears misaligned"
    signal = row["BenchmarkQualitySignal"] or row["ComparisonNote"]
    if signal:
        return str(signal)
    return "suspicious benchmark behavior"


def _audit_reason_summary(row: dict[str, Any]) -> str:
    reasons: list[str] = []
    if row["RedundancyScore"] >= 0.55:
        reasons.append("redundancy")
    if row.get("AttemptVarianceLabel") not in {"", "stable"}:
        reasons.append("attempt_variance")
    if row.get("ModelDisagreementLabel") not in {"", "aligned", "single_model"}:
        reasons.append("model_disagreement")
    if row.get("ThresholdSensitivityLabel") not in {"", "stable"}:
        reasons.append("threshold_sensitivity")
    if row["ContradictionLabel"]:
        reasons.append(row["ContradictionLabel"])
    if _is_benchmark_defect_candidate(row):
        reasons.append("benchmark_defect")
    if _is_trivial_candidate(row):
        reasons.append("triviality")
    flags = _split_flags(row["OutlierFlags"])
    if flags:
        reasons.append("outlier")
    return ", ".join(reasons) if reasons else "no elevated audit signal"


def _difficulty_distribution_result(counter: dict[str, int]) -> str:
    nonzero = sum(1 for value in counter.values() if value)
    return "PASS" if nonzero >= 3 else "REVIEW"


def _difficulty_distribution_interpretation(counter: dict[str, int]) -> str:
    if _difficulty_distribution_result(counter) == "PASS":
        return "The dataset spans multiple observed difficulty regimes."
    return "The observed difficulty regime looks narrow or unstable."


def _format_counter(counter: dict[str, int]) -> str:
    if not counter:
        return "none"
    return "; ".join(f"{key}={value}" for key, value in sorted(counter.items()))


def _format_distribution(distribution: dict[str, Any]) -> str:
    if not distribution:
        return "none"
    return (
        f"count={distribution.get('count')}, "
        f"mean={distribution.get('mean')}, "
        f"median={distribution.get('median')}, "
        f"q1={distribution.get('q1')}, "
        f"q3={distribution.get('q3')}"
    )


def _pair_label_from_rows(left: dict[str, Any], right: dict[str, Any]) -> str:
    labels = {left["DuplicateLabel"], right["DuplicateLabel"]}
    if "exact_duplicate" in labels:
        return "exact_duplicate"
    if "semantic_duplicate" in labels:
        return "semantic_duplicate"
    if "near_duplicate" in labels:
        return "near_duplicate"
    return "template_overlap"


def _load_stage1(workbook_path: Path) -> dict[str, dict[int, dict[str, Any]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    detailed_rows = _read_sheet_rows(workbook, "Detailed")
    aggregate_rows = _read_sheet_rows(workbook, "Summary")
    return {
        "detailed": {int(row["Index"]): row for row in detailed_rows},
        "aggregate": {int(row["Index"]): row for row in aggregate_rows},
    }


def _load_stage2(stage2_dir: Path) -> dict[str, Any]:
    sample_rows = {
        int(row["Index"]): row
        for row in _read_csv_rows(stage2_dir / "sample_results.csv")
    }
    model_rows_all = _read_csv_rows(stage2_dir / "sample_model_results.csv")
    model_rows: dict[int, dict[str, Any]] = {}
    for row in model_rows_all:
        index = int(row["Index"])
        existing = model_rows.get(index)
        if existing is None or _to_float(row.get("BestCombinedPassRate")) > _to_float(
            existing.get("BestCombinedPassRate")
        ):
            model_rows[index] = row

    attempt_rows = _read_jsonl_rows(stage2_dir / "model_attempts.jsonl")
    attempt_rows_by_sample: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in attempt_rows:
        attempt_rows_by_sample[int(row["Index"])].append(row)
    model_rows_by_sample: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in model_rows_all:
        model_rows_by_sample[int(row["Index"])].append(row)
    per_test_rows_all = _read_jsonl_rows(stage2_dir / "per_test_results.jsonl")
    per_test_rows_by_sample: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in per_test_rows_all:
        per_test_rows_by_sample[int(row["sample_index"])].append(row)

    manifest = json.loads((stage2_dir / "run_manifest.json").read_text(encoding="utf-8"))
    return {
        "sample_rows": sample_rows,
        "model_rows": model_rows,
        "best_model_rows": model_rows,
        "model_rows_all": model_rows_all,
        "attempt_rows": attempt_rows,
        "attempt_rows_by_sample": attempt_rows_by_sample,
        "model_rows_by_sample": model_rows_by_sample,
        "per_test_rows_all": per_test_rows_all,
        "per_test_rows": per_test_rows_by_sample,
        "manifest": manifest,
        "output_root": stage2_dir,
    }


def _similarity_pairs(
    rows: list[dict[str, Any]],
    *,
    near_duplicate_threshold: float,
    template_threshold: float,
) -> list[SimilarityPair]:
    pairs: list[SimilarityPair] = []
    for left, right in combinations(rows, 2):
        prompt_similarity = _jaccard(left["_prompt_ngrams"], right["_prompt_ngrams"])
        template_similarity = _jaccard(left["_template_ngrams"], right["_template_ngrams"])
        test_similarity = _jaccard(left["_test_signature"], right["_test_signature"])
        starter_similarity = _jaccard(left["_starter_signature"], right["_starter_signature"])
        title_similarity = _jaccard(left["_title_ngrams"], right["_title_ngrams"])
        function_similarity = _jaccard(left["_function_signature"], right["_function_signature"])
        lexical_similarity = max(
            prompt_similarity,
            template_similarity,
            (0.8 * prompt_similarity) + (0.2 * title_similarity),
        )
        structural_similarity = max(
            (0.55 * template_similarity) + (0.25 * test_similarity) + (0.20 * starter_similarity),
            (0.55 * starter_similarity) + (0.30 * function_similarity) + (0.15 * title_similarity),
        )
        embedding_similarity = _cosine_similarity(
            left.get("_embedding_vector", []),
            right.get("_embedding_vector", []),
        )
        combined_similarity = round(
            max(
                lexical_similarity,
                structural_similarity,
                (0.45 * lexical_similarity) + (0.30 * structural_similarity) + (0.25 * embedding_similarity),
                (0.55 * embedding_similarity) + (0.25 * structural_similarity) + (0.20 * lexical_similarity),
            ),
            4,
        )
        if prompt_similarity >= 0.98 and left["FunctionName"] == right["FunctionName"]:
            label = "exact_duplicate"
        elif embedding_similarity >= 0.9 and structural_similarity >= 0.55:
            label = "semantic_duplicate"
        elif combined_similarity >= near_duplicate_threshold:
            label = "near_duplicate"
        elif template_similarity >= template_threshold or structural_similarity >= template_threshold:
            label = "template_overlap"
        else:
            continue
        pairs.append(
            SimilarityPair(
                left_index=left["Index"],
                right_index=right["Index"],
                prompt_similarity=round(prompt_similarity, 4),
                template_similarity=round(template_similarity, 4),
                test_similarity=round(test_similarity, 4),
                starter_similarity=round(starter_similarity, 4),
                title_similarity=round(title_similarity, 4),
                function_similarity=round(function_similarity, 4),
                embedding_similarity=round(embedding_similarity, 4),
                lexical_similarity=round(lexical_similarity, 4),
                structural_similarity=round(structural_similarity, 4),
                combined_similarity=combined_similarity,
                similarity_label=label,
            )
        )
    return pairs


def _cluster_pairs(
    pairs: list[SimilarityPair],
    sample_indices: list[int],
    cluster_threshold: float,
) -> tuple[dict[int, int], dict[int, int]]:
    parents = {index: index for index in sample_indices}

    def find(index: int) -> int:
        while parents[index] != index:
            parents[index] = parents[parents[index]]
            index = parents[index]
        return index

    def union(left: int, right: int) -> None:
        root_left = find(left)
        root_right = find(right)
        if root_left == root_right:
            return
        if root_left < root_right:
            parents[root_right] = root_left
        else:
            parents[root_left] = root_right

    for pair in pairs:
        if pair.combined_similarity >= cluster_threshold:
            union(pair.left_index, pair.right_index)

    cluster_assignments = {index: find(index) for index in sample_indices}
    cluster_sizes = Counter(cluster_assignments.values())
    return cluster_assignments, dict(cluster_sizes)


def _duplicate_detail_maps(pairs: list[SimilarityPair]) -> dict[int, dict[str, Any]]:
    detail_map: dict[int, dict[str, Any]] = {}
    for pair in pairs:
        for sample_index, other_index in (
            (pair.left_index, pair.right_index),
            (pair.right_index, pair.left_index),
        ):
            current = detail_map.get(sample_index)
            if current is None or pair.combined_similarity > current["max_similarity"]:
                detail_map[sample_index] = {
                    "closest_index": other_index,
                    "max_similarity": pair.combined_similarity,
                    "max_label": pair.similarity_label,
                    "max_lexical_similarity": pair.lexical_similarity,
                    "max_embedding_similarity": pair.embedding_similarity,
                }
    return detail_map


def _compute_outlier_flags(rows: list[dict[str, Any]]) -> dict[int, list[str]]:
    metrics = {
        "prompt_length_outlier": {row["Index"] for row in rows if _iqr_flag(row["PromptWords"], [item["PromptWords"] for item in rows])},
        "starter_code_outlier": {
            row["Index"] for row in rows if _iqr_flag(row["StarterCodeLines"], [item["StarterCodeLines"] for item in rows])
        },
        "ideal_response_outlier": {
            row["Index"] for row in rows if _iqr_flag(row["IdealResponseLines"], [item["IdealResponseLines"] for item in rows])
        },
        "test_count_outlier": {row["Index"] for row in rows if _iqr_flag(row["TotalTests"], [item["TotalTests"] for item in rows])},
        "pass_rate_outlier": {
            row["Index"] for row in rows if _iqr_flag(row["WinnerCombinedPassRate"], [item["WinnerCombinedPassRate"] for item in rows])
        },
    }
    flags_by_index: dict[int, list[str]] = defaultdict(list)
    for flag, indices in metrics.items():
        for index in indices:
            flags_by_index[index].append(flag)
    return flags_by_index


def _audit_priority(
    row: dict[str, Any],
    redundancy_score: float,
    outlier_flags: list[str],
    contradiction: str,
) -> str:
    if row["Suspicious"] or row["OraclePassRate"] < 0.9999 or contradiction == "high_static_low_dynamic":
        return "critical"
    if (
        redundancy_score >= 0.82
        or len(outlier_flags) >= 2
        or contradiction
        or row.get("ModelDisagreementLabel") == "strong_disagreement"
    ):
        return "high"
    if (
        row["NeedsAudit"]
        or redundancy_score >= 0.72
        or outlier_flags
        or row.get("AttemptVarianceLabel") == "volatile"
        or row.get("ThresholdSensitivityLabel") != "stable"
        or row.get("ModelDisagreementLabel") == "moderate_disagreement"
    ):
        return "medium"
    return "normal"


def _contradiction_label(row: dict[str, Any]) -> str:
    if row["Stage1Score"] >= 0.82 and (
        row["Suspicious"] or row["WinnerCombinedPassRate"] <= 0.2 or row["EfficacyLabel"] == "Inconclusive"
    ):
        return "high_static_low_dynamic"
    if row["Stage1Score"] <= 0.45 and row["WinnerCombinedPassRate"] >= 0.9:
        return "low_static_high_dynamic"
    return ""


def _is_exemplar_candidate(row: dict[str, Any], redundancy_score: float, outlier_flags: list[str]) -> bool:
    return (
        row["Stage1Score"] >= 0.78
        and row["EfficacyLabel"] in EXEMPLAR_EFFICACY
        and not row["Suspicious"]
        and not _contradiction_label(row)
        and row.get("AttemptVarianceLabel") == "stable"
        and row.get("ModelDisagreementLabel") in {"aligned", "single_model"}
        and row.get("ThresholdSensitivityLabel") == "stable"
        and redundancy_score < 0.55
        and not outlier_flags
    )


def _top_risks(
    joined_rows: list[dict[str, Any]],
    duplicate_counts: Counter[str],
    contradiction_counts: Counter[str],
    benchmark_quality: Counter[str],
) -> list[str]:
    risks: list[str] = []
    suspicious_share = sum(1 for row in joined_rows if row["Suspicious"]) / len(joined_rows)
    if suspicious_share >= 0.08:
        risks.append(f"{suspicious_share:.1%} of samples are already flagged as suspicious or broken.")
    if duplicate_counts:
        duplicate_total = sum(1 for row in joined_rows if row["RedundancyScore"] >= 0.55)
        risks.append(f"{duplicate_total} samples appear materially redundant or template-recycled.")
    if contradiction_counts.get("high_static_low_dynamic", 0):
        risks.append(
            f"{contradiction_counts['high_static_low_dynamic']} samples look strong statically but fail dynamically."
        )
    unstable_attempts = sum(1 for row in joined_rows if row.get("AttemptVarianceLabel") == "volatile")
    if unstable_attempts:
        risks.append(f"{unstable_attempts} samples show volatile repeated-attempt behavior.")
    disagreement_rows = sum(
        1 for row in joined_rows if row.get("ModelDisagreementLabel") not in {"", "aligned", "single_model"}
    )
    if disagreement_rows:
        risks.append(f"{disagreement_rows} samples show material cross-model disagreement.")
    if benchmark_quality:
        label, count = benchmark_quality.most_common(1)[0]
        if label and label != "clean_evaluation":
            risks.append(f"Most common non-clean benchmark quality signal: {label} ({count} samples).")
    return risks[:4]


def _recommendations(
    joined_rows: list[dict[str, Any]],
    top_relationship: dict[str, Any],
    duplicate_counts: Counter[str],
) -> list[str]:
    recommendations: list[str] = []
    if duplicate_counts:
        recommendations.append("Deduplicate or downweight near-identical prompt templates before using aggregate scores.")
    if top_relationship:
        recommendations.append(
            "Prioritize manual review for samples failing "
            f"{top_relationship.get('Signal', 'the strongest Stage 1 signal')}."
        )
    if any(row.get("AttemptVarianceLabel") == "volatile" for row in joined_rows):
        recommendations.append("Treat unstable repeated-attempt samples as caveated benchmark signal, not clean difficulty evidence.")
    if any(row.get("ModelDisagreementLabel") not in {"", "aligned", "single_model"} for row in joined_rows):
        recommendations.append("Use cross-model disagreement to triage ambiguous benchmark items before trusting aggregate conclusions.")
    low_efficacy_share = sum(1 for row in joined_rows if row["EfficacyLabel"] == "Low Efficacy") / len(joined_rows)
    if low_efficacy_share >= 0.2:
        recommendations.append("Refresh the easiest items: the benchmark shows noticeable saturation.")
    suspicious_rows = [row for row in joined_rows if row["Suspicious"]]
    if suspicious_rows:
        recommendations.append("Stage 4 should start with suspicious benchmark-defect candidates, not random sampling.")
    return recommendations[:4]


def _write_markdown(
    path: Path,
    summary: dict[str, Any],
    audit_queues: dict[str, list[dict[str, Any]]],
    relationship_rows: list[dict[str, Any]],
) -> None:
    findings = summary.get("headline_findings", {})
    strongest_signal = findings.get("strongest_stage1_to_stage2_signal", {})
    lines = [
        "# Stage 3 Dataset Analysis",
        "",
        f"- Samples analyzed: {summary.get('dataset_size', 0)}",
        f"- Biggest structural risks: {'; '.join(findings.get('top_risks', [])) or 'None detected.'}",
        f"- Strongest Stage 1 to Stage 2 signal: {strongest_signal.get('Signal', 'n/a')}",
        f"- Recommended next actions: {'; '.join(findings.get('recommendations', [])) or 'None recorded.'}",
        "",
        "## Audit Queues",
    ]
    for queue_name, entries in audit_queues.items():
        lines.append(f"### {queue_name}")
        if not entries:
            lines.append("- No items.")
            continue
        for item in entries[:5]:
            lines.append(
                f"- Index {item['Index']} ({item['QuestionId']}): {item['Reason']} "
                f"[priority={item['AuditPriority']}, utility={item['DatasetUtilityLabel']}, redundancy={item['RedundancyScore']}]"
            )
    lines.extend(["", "## Stage 1 to Stage 2 Relationships"])
    if relationship_rows:
        for row in relationship_rows[:5]:
            lines.append(
                f"- {row['Signal']}: sample_count={row['SampleCount']}, "
                f"pass_delta={row['PassRateDeltaVsOthers']}, suspicious_lift={row['SuspiciousLiftVsOthers']}"
            )
    else:
        lines.append("- No relationship rows generated.")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_workbook(
    path: Path,
    *,
    summary_rows: list[dict[str, Any]],
    detailed_rows: list[dict[str, Any]],
    test_columns: list[str],
) -> None:
    workbook = Workbook()
    ws_summary = workbook.active
    ws_summary.title = "Summary"
    _write_sheet(ws_summary, summary_rows)

    ws_detailed = workbook.create_sheet("Detailed")
    _write_sheet(
        ws_detailed,
        detailed_rows,
        percentage_columns={
            "ClosestNeighborSimilarity",
        },
        test_columns=set(test_columns),
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _write_sheet(
    ws,
    rows: list[dict[str, Any]],
    percentage_columns: set[str] | None = None,
    test_columns: set[str] | None = None,
) -> None:
    if not rows:
        ws.append(["Empty"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
    for row in rows:
        ws.append([_excel_safe(row.get(header, "")) for header in headers])
    ws.freeze_panes = "A2"
    percentage_columns = percentage_columns or set()
    for column_index, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=column_index).column_letter].width = min(
            max(len(str(header)) + 2, 14),
            40,
        )
        if test_columns and header in test_columns:
            for cell in ws.iter_cols(min_col=column_index, max_col=column_index, min_row=2):
                for value_cell in cell:
                    _apply_test_fill(value_cell)
            ws.column_dimensions[ws.cell(row=1, column=column_index).column_letter].width = 14
        if header in percentage_columns:
            for cell in ws.iter_cols(min_col=column_index, max_col=column_index, min_row=2):
                for value_cell in cell:
                    if isinstance(value_cell.value, (int, float)):
                        value_cell.number_format = "0.0%"


def _apply_test_fill(cell) -> None:
    value = str(cell.value or "")
    if value == "PASS":
        cell.fill = PatternFill("solid", fgColor="C6E0B4")
    elif value.startswith("FAIL"):
        cell.fill = PatternFill("solid", fgColor="F4CCCC")
    elif value.startswith("ERROR") or value.startswith("TIMEOUT"):
        cell.fill = PatternFill("solid", fgColor="FCE4D6")


def _excel_safe(value: Any) -> Any:
    if isinstance(value, str):
        return _ILLEGAL_XLSX_CHARS.sub("", value)
    return value


def _test_column_name(visibility: str, case_index: int) -> str:
    label = visibility.capitalize() if visibility else "Test"
    return f"{label}_{case_index + 1:02d}"


def _test_column_sort_key(label: str) -> tuple[int, int]:
    visibility, _, suffix = label.partition("_")
    order = {"Public": 0, "Private": 1, "Generated": 2}.get(visibility, 9)
    try:
        index = int(suffix)
    except ValueError:
        index = 999
    return (order, index)


def _test_cell_value(row: dict[str, Any]) -> str:
    status = str(row.get("status", "")).lower()
    if status == "pass":
        return "PASS"
    if status == "fail":
        failure_type = str(row.get("failure_type", "")).strip() or "incorrect_output"
        return f"FAIL:{failure_type}"
    if status == "timeout":
        return "TIMEOUT"
    if status == "error":
        failure_type = str(row.get("failure_type", "")).strip()
        exception_type = str(row.get("exception_type", "")).strip()
        detail = failure_type or exception_type or "error"
        return f"ERROR:{detail}"
    return status.upper()


def _resolve_path(path: Path) -> Path:
    return path if path.is_absolute() else (PROJECT_ROOT / path).resolve()


def _read_sheet_rows(workbook, sheet_name: str) -> list[dict[str, Any]]:
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value).strip().lstrip("\ufeff") for value in rows[0]]
    return [
        {headers[index]: row[index] for index in range(len(headers))}
        for row in rows[1:]
        if any(value is not None and str(value).strip() for value in row)
    ]


def _read_csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [
            {key.lstrip("\ufeff"): value for key, value in row.items()}
            for row in csv.DictReader(handle)
        ]


def _read_jsonl_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as handle:
        return [json.loads(line) for line in handle if line.strip()]


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def _count_statuses(stage1_detailed: dict[str, Any]) -> Counter[str]:
    counter: Counter[str] = Counter()
    for key, value in stage1_detailed.items():
        if key in {"Index", "Question_Id", "Question_Title", "Difficulty", "Function_Name", "Runtime_Pass_Rate"}:
            continue
        if str(key).startswith("Notes-"):
            continue
        status = str(value).strip()
        if status:
            counter[status] += 1
    return counter


def _stage1_score(status_counts: Counter[str]) -> float:
    applicable = sum(status_counts[status] for status in ("PASS", "PARTIAL", "UNCLEAR", "FAIL"))
    if not applicable:
        return 0.0
    weighted = (
        status_counts["PASS"] * 1.0
        + status_counts["PARTIAL"] * 0.5
        + status_counts["UNCLEAR"] * 0.25
    )
    return round(weighted / applicable, 4)


def _stage1_section_scores(stage1_detailed: dict[str, Any]) -> dict[str, float]:
    section_values: dict[str, list[float]] = defaultdict(list)
    for key, value in stage1_detailed.items():
        status = str(value).strip()
        if status not in STATUS_NUMERIC:
            continue
        for prefix in SECTION_PREFIXES:
            if key.startswith(prefix):
                section_values[prefix[0]].append(STATUS_NUMERIC[status])
                break
    return {
        section: round(sum(values) / len(values), 4)
        for section, values in section_values.items()
        if values
    }


def _count_critical_fails(stage1_detailed: dict[str, Any]) -> int:
    critical_prefixes = ("3.", "4_", "6.")
    return sum(
        1
        for key, value in stage1_detailed.items()
        if any(key.startswith(prefix) for prefix in critical_prefixes) and str(value).strip() == "FAIL"
    )


def _top_failures(stage1_detailed: dict[str, Any]) -> list[str]:
    return [
        key
        for key, value in stage1_detailed.items()
        if str(value).strip() in {"FAIL", "PARTIAL", "UNCLEAR"}
    ][:8]


def _all_failure_flags(stage1_detailed: dict[str, Any]) -> list[str]:
    return [
        key
        for key, value in stage1_detailed.items()
        if str(value).strip() in {"FAIL", "PARTIAL", "UNCLEAR"}
    ]


def _word_count(text: str) -> int:
    return len(re.findall(r"\w+", text))


def _line_count(text: str) -> int:
    stripped = text.strip("\n")
    return 0 if not stripped else len(stripped.splitlines())


def _normalize_text(text: str) -> str:
    text = text.lower()
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _template_text(text: str) -> str:
    text = _normalize_text(text)
    text = re.sub(r"\d+", " <num> ", text)
    text = re.sub(r"'[^']*'|\"[^\"]*\"", " <str> ", text)
    text = re.sub(r"\b[a-z_][a-z0-9_]{2,}\b", lambda match: _token_class(match.group(0)), text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _token_class(token: str) -> str:
    if token in {"input", "output", "return", "function", "problem", "statement", "constraints"}:
        return token
    if token.startswith("__") and token.endswith("__"):
        return "<magic>"
    return "<id>"


def _char_ngrams(text: str, size: int = 5) -> set[str]:
    compact = re.sub(r"\s+", " ", text)
    if len(compact) <= size:
        return {compact} if compact else set()
    return {compact[index : index + size] for index in range(len(compact) - size + 1)}


def _test_signature(sample) -> set[str]:
    features: set[str] = set()
    for test in sample.public_tests + sample.private_tests:
        payload = _normalize_text(f"{test.visibility}:{test.testtype}:{test.input_text}:{test.output_text}")
        features.update(_char_ngrams(payload, size=4))
    return features


def _jaccard(left: set[str], right: set[str]) -> float:
    if not left and not right:
        return 1.0
    if not left or not right:
        return 0.0
    union = left | right
    if not union:
        return 0.0
    return len(left & right) / len(union)


def _iqr_flag(value: float, population: list[float]) -> bool:
    if len(population) < 4:
        return False
    q1, _, q3 = statistics.quantiles(population, n=4, method="inclusive")
    iqr = q3 - q1
    if iqr == 0:
        return False
    return value < (q1 - 1.5 * iqr) or value > (q3 + 1.5 * iqr)


def _numeric_summary(values: list[float]) -> dict[str, float | int | None]:
    if not values:
        return {"count": 0, "mean": None, "median": None, "min": None, "q1": None, "q3": None, "max": None}
    sorted_values = sorted(values)
    quartiles = (
        statistics.quantiles(sorted_values, n=4, method="inclusive")
        if len(sorted_values) >= 2
        else [sorted_values[0]] * 3
    )
    return {
        "count": len(values),
        "mean": round(statistics.mean(values), 4),
        "median": round(statistics.median(values), 4),
        "min": round(min(values), 4),
        "q1": round(quartiles[0], 4),
        "q3": round(quartiles[2], 4),
        "max": round(max(values), 4),
    }


def _pearson(left: list[float], right: list[float]) -> float:
    if len(left) != len(right) or len(left) < 2:
        return 0.0
    mean_left = statistics.mean(left)
    mean_right = statistics.mean(right)
    numerator = sum((a - mean_left) * (b - mean_right) for a, b in zip(left, right, strict=False))
    denom_left = math.sqrt(sum((a - mean_left) ** 2 for a in left))
    denom_right = math.sqrt(sum((b - mean_right) ** 2 for b in right))
    if denom_left == 0 or denom_right == 0:
        return 0.0
    return numerator / (denom_left * denom_right)


def _to_float(value: Any) -> float:
    if value in {None, ""}:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _to_int(value: Any) -> int:
    if value in {None, ""}:
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() == "true"


def _split_flags(raw: str) -> set[str]:
    return {piece.strip() for piece in str(raw).split(";") if piece.strip()}


def _failure_flags_for_relationships(row: dict[str, Any]) -> set[str]:
    raw = row.get("_Stage1FailureFlags")
    if isinstance(raw, list):
        return {str(piece).strip() for piece in raw if str(piece).strip()}
    return _split_flags(row.get("Stage1Flags", ""))


def _stage1_check_names_from_rows(joined_rows: list[dict[str, Any]]) -> list[str]:
    names: set[str] = set()
    for row in joined_rows:
        names.update(_failure_flags_for_relationships(row))
    return sorted(names)
