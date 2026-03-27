from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from . import schema
from .requirements import FAIL, NA, PARTIAL, PASS, UNCLEAR


_ILLEGAL_XLSX_CHARS = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


def merge_section_rows(section_rows: list[list[dict[str, Any]]]) -> list[dict[str, Any]]:
    merged_by_index: dict[int, dict[str, Any]] = {}
    for rows in section_rows:
        for row in rows:
            target = merged_by_index.setdefault(row["Index"], {})
            target.update(row)
    return [merged_by_index[index] for index in sorted(merged_by_index)]


def count_statuses(row: dict[str, Any], keys: set[str]) -> tuple[int, int, int]:
    fail_count = sum(1 for key in keys if row.get(key) == FAIL)
    partial_count = sum(1 for key in keys if row.get(key) == PARTIAL)
    unclear_count = sum(1 for key in keys if row.get(key) == UNCLEAR)
    return fail_count, partial_count, unclear_count


def parse_runtime_pass_rate(value: Any) -> tuple[int, int]:
    if not isinstance(value, str) or "/" not in value:
        return (0, 0)
    left, right = value.split("/", 1)
    try:
        return int(left), int(right)
    except ValueError:
        return (0, 0)


def pass_fraction(value: Any) -> float:
    passed, total = parse_runtime_pass_rate(value)
    return (passed / total) if total else 0.0


def requirement_keys_for_category(category: str) -> list[str]:
    return [key for key in schema.DETAILED_KEYS if schema.SECTION_MAP[key][2] == category]


def classify_prompt(row: dict[str, Any]) -> str:
    prompt_keys = set(requirement_keys_for_category("Prompt")) | {
        "v_prompt_test_solution_aligned",
        "v_model_breaking_prompt_defined_only",
        "v_no_unmentioned_internal_logic",
        "v_output_schema_aligned",
    }
    fail_count, partial_count, unclear_count = count_statuses(row, prompt_keys)
    if row.get("p_function_signature_present") == FAIL:
        return "Unusable"
    if (
        row.get("p_practical_algorithmic_problem") == FAIL
        and row.get("p_not_vague") == FAIL
        and row.get("p_no_conflicting_objectives") == FAIL
    ):
        return "Unusable"
    if (
        row.get("v_prompt_test_solution_aligned") == FAIL
        and row.get("v_model_breaking_prompt_defined_only") == FAIL
        and row.get("v_no_unmentioned_internal_logic") == FAIL
    ):
        return "Unusable"
    if (
        row.get("p_not_vague") == FAIL
        and row.get("p_no_conflicting_objectives") == FAIL
        and row.get("p_no_unrealistic_constraints") == FAIL
    ):
        return "Unusable"
    if fail_count or partial_count or unclear_count:
        return "Needs Fixing"
    return "Usable"


def classify_ideal(row: dict[str, Any]) -> str:
    ideal_keys = set(requirement_keys_for_category("Ideal Response")) | {
        "v_no_unmentioned_internal_logic",
        "v_prompt_test_solution_aligned",
        "v_model_breaking_prompt_defined_only",
    }
    fail_count, partial_count, unclear_count = count_statuses(row, ideal_keys)
    runtime_fraction = pass_fraction(row.get("Runtime_Pass_Rate"))
    if row.get("i_executes_without_error") == FAIL:
        return "Unusable"
    if row.get("i_passes_internal_tests") == FAIL or runtime_fraction < 0.5:
        return "Unusable"
    if (
        row.get("v_no_unmentioned_internal_logic") == FAIL
        and row.get("v_prompt_test_solution_aligned") == FAIL
        and row.get("v_model_breaking_prompt_defined_only") == FAIL
    ):
        return "Unusable"
    if fail_count or partial_count or unclear_count or runtime_fraction < 1.0:
        return "Needs Fixing"
    return "Usable"


def classify_tests(row: dict[str, Any]) -> str:
    test_keys = set(requirement_keys_for_category("Test Cases")) | {
        "v_model_breaking_prompt_defined_only",
        "v_no_unmentioned_internal_logic",
        "v_prompt_test_solution_aligned",
        "v_no_extra_parameters",
        "v_cross_verified_dry_run",
    }
    fail_count, partial_count, unclear_count = count_statuses(row, test_keys)
    if row.get("t_json_encoded") == FAIL or row.get("t_string_fields") == FAIL:
        return "Unusable"
    if row.get("v_no_extra_parameters") == FAIL:
        return "Unusable"
    if (
        row.get("v_prompt_test_solution_aligned") == FAIL
        and row.get("v_model_breaking_prompt_defined_only") == FAIL
    ):
        return "Unusable"
    if (
        row.get("t_json_escaping_valid") == FAIL
        and (row.get("t_input_json_object") == FAIL or row.get("t_output_json_object") == FAIL)
    ):
        return "Unusable"
    if (
        row.get("v_cross_verified_dry_run") == FAIL
        and row.get("v_prompt_test_solution_aligned") == FAIL
    ):
        return "Unusable"
    if fail_count or partial_count or unclear_count:
        return "Needs Fixing"
    return "Usable"


def build_summary_rows(detailed_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    summary_rows: list[dict[str, Any]] = []
    metadata_names = [name for name, _ in schema.METADATA_COLUMNS]
    for row in detailed_rows:
        legacy_view = dict(row)
        for key in schema.DETAILED_KEYS:
            legacy_view[key] = row.get(schema.display_name(key))
        summary_rows.append(
            {
                **{name: row[name] for name in metadata_names},
                "Prompt": classify_prompt(legacy_view),
                "Ideal_Response": classify_ideal(legacy_view),
                "Test_Cases": classify_tests(legacy_view),
            }
        )
    return summary_rows


def write_sheet(ws, rows: list[dict[str, Any]], percentage_columns: set[str] | None = None) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in rows:
        ws.append([_excel_safe(row.get(header)) for header in headers])
    ws.freeze_panes = "A2"
    percentage_columns = percentage_columns or set()
    for idx, header in enumerate(headers, start=1):
        if header in percentage_columns:
            for col in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for cell in col:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0%"
        if header in {"Prompt", "Ideal_Response", "Test_Cases"}:
            for col in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for cell in col:
                    if cell.value == "Usable":
                        cell.fill = PatternFill("solid", fgColor="C6E0B4")
                    elif cell.value == "Needs Fixing":
                        cell.fill = PatternFill("solid", fgColor="FFF2CC")
                    elif cell.value == "Unusable":
                        cell.fill = PatternFill("solid", fgColor="F4CCCC")
        ws.column_dimensions[get_column_letter(idx)].width = max(14, min(44, len(header) + 2))


def write_workbook(detailed_rows: list[dict[str, Any]], subsection_rows: list[dict[str, Any]], summary_rows: list[dict[str, Any]], output_path: Path) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Detailed"
    write_sheet(ws1, detailed_rows)
    ws2 = wb.create_sheet("Summary")
    write_sheet(ws2, summary_rows)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _excel_safe(value: Any) -> Any:
    if isinstance(value, str):
        return _ILLEGAL_XLSX_CHARS.sub("", value)
    return value
