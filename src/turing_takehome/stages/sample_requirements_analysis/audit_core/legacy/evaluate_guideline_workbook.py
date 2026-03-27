from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import evaluate_guideline as base


KEEP_METADATA = [
    ("Index", "Index"),
    ("Question_Id", "QuestionId"),
    ("Question_Title", "QuestionTitle"),
    ("Difficulty", "Difficulty"),
    ("Function_Name", "FunctionName"),
    ("Runtime_Pass_Rate", "RuntimePassRate"),
]
DROP_REQUIREMENTS = {
    "s_in_markdown",
    "i_structured_classes_or_functions",
    "i_consistent_structure",
    "t_no_markdown_json_fence",
    "t_one_to_one_json_cases",
    "t_exact_output_check",
}

VERDICT_SCORE = {base.PASS: 1.0, base.PARTIAL: 0.5, base.FAIL: 0.0}


SECTION_INFO: dict[str, tuple[str, str, str]] = {}
REQUIREMENT_DISPLAY_ORDER: list[str] = []


def assign(keys: list[str], code: str, subsection: str, category: str) -> None:
    for key in keys:
        SECTION_INFO[key] = (code, subsection, category)
        REQUIREMENT_DISPLAY_ORDER.append(key)


assign(
    [
        "p_structured_layout",
        "p_realistic_context",
        "p_input_format_explicit",
        "p_output_format_explicit",
        "p_constraints_defined",
        "p_computational_limits_defined",
        "p_edge_cases_defined",
        "p_return_conditions_defined",
        "p_function_signature_present",
        "p_no_external_libs_stated",
        "p_metadata_alignment",
        "p_not_verbose",
    ],
    "1.1",
    "1.1_Structural_Expectations",
    "Prompt",
)
assign(
    [
        "p_practical_algorithmic_problem",
        "p_measurable_objective",
        "p_difficulty_balanced",
        "p_example_present",
        "p_no_unrealistic_constraints",
        "p_not_vague",
        "p_no_conflicting_objectives",
        "p_no_buzzwords",
        "p_no_time_window_constraints",
        "p_no_random_requirement",
    ],
    "1.2",
    "1.2_Content_Guidelines",
    "Prompt",
)
assign(["p_json_compatible_signature"], "1.3", "1.3_Quality_Checklist", "Prompt")
assign(["s_necessary_imports", "s_only_entry_signature", "s_no_classes", "s_no_helpers", "s_no_logic"], "7", "7_Starter_Code", "Prompt")
assign(
    [
        "i_no_globals",
        "i_state_encapsulated",
        "i_consistent_naming_docs",
        "i_no_arbitrary_limits",
        "i_single_entry_aligned",
        "i_helpers_for_repeated_logic",
        "i_no_sample_io_in_main",
        "i_no_parallelism",
        "i_deterministic_solution",
        "i_mp_module_level_functions",
        "i_mp_context_manager",
        "i_mp_sequential_fallback",
    ],
    "2.1",
    "2.1_Code_Quality_Expectations",
    "Ideal Response",
)
assign(
    [
        "i_no_keyword_only_args",
        "i_no_future_import",
        "i_no_redundant_memoization",
        "i_clear_variable_names",
        "i_no_nested_helpers",
        "i_stdlib_only",
    ],
    "2.2",
    "2.2_Common_Errors_to_Avoid",
    "Ideal Response",
)
assign(["i_executes_without_error", "i_passes_internal_tests"], "2.3", "2.3_Verification_Checklist", "Ideal Response")
assign(["t_single_call_per_test", "t_public_test1_matches_prompt_example"], "3.1", "3.1_Structure_and_Purpose", "Test Cases")
assign(
    ["t_min_5_public", "t_min_10_private", "t_recommended_15_20_total", "t_deterministic", "t_entry_function_only"],
    "3.2",
    "3.2_Test_Requirements",
    "Test Cases",
)
assign(
    [
        "t_json_encoded",
        "t_string_fields",
        "t_input_json_object",
        "t_output_json_object",
        "t_json_escaping_valid",
        "t_optional_values_included",
        "t_no_python_literals",
        "t_no_nonstring_keys",
    ],
    "3.3",
    "3.3_JSON_Format_Compliance",
    "Test Cases",
)
assign(["t_exception_tests_aligned"], "3.3.1", "3.3.1_Exception_and_Error_Handling", "Test Cases")
assign(["t_not_large_or_redundant"], "3.4", "3.4_Common_Test_Mistakes", "Test Cases")
assign(["v_coverage_confidence", "v_cross_verified_dry_run"], "4", "4_Final_Validation_Workflow", "Test Cases")
assign(["v_prompt_test_solution_aligned", "v_entry_name_consistent", "v_signature_arity_consistent", "v_output_schema_aligned"], "6.1", "6.1_Model_Breaking_Definition", "Test Cases")
assign(["v_model_breaking_prompt_defined_only", "v_no_extra_parameters", "v_no_unmentioned_internal_logic"], "6.2", "6.2_Invalid_Model_Breaking", "Test Cases")
assign(["v_prompt_edge_cases_tested", "v_prompt_constraints_tested"], "6.3", "6.3_Valid_Model_Breaking", "Test Cases")


def display_name(key: str) -> str:
    if key not in SECTION_INFO:
        return key
    code, _, _ = SECTION_INFO[key]
    base_name = key.split("_", 1)[1]
    return f"{code}_{base_name}"


DETAILED_KEYS = [key for key in REQUIREMENT_DISPLAY_ORDER if key not in DROP_REQUIREMENTS]
DETAILED_HEADERS = [name for name, _ in KEEP_METADATA] + [display_name(key) for key in DETAILED_KEYS] + ["Notes"]
DISPLAY_TO_KEY = {display_name(key): key for key in DETAILED_KEYS}


def section_sort_key(label: str) -> tuple[int, ...]:
    code = label.split("_", 1)[0]
    return tuple(int(part) for part in code.split("."))


def parse_indices_arg(raw: str | None) -> list[int] | None:
    return base.parse_indices_arg(raw)


def collect_rows(args: argparse.Namespace) -> list[dict[str, Any]]:
    rows_text = [line for line in args.jsonl.read_text(encoding="utf-8").splitlines() if line.strip()]
    explicit_indices = parse_indices_arg(args.indices)
    if explicit_indices is not None:
        selected_pairs = [(index, rows_text[index]) for index in explicit_indices]
    else:
        selected = rows_text[args.offset:]
        if args.limit is not None:
            selected = selected[: args.limit]
        selected_pairs = list(enumerate(selected, start=args.offset))

    output_rows: list[dict[str, Any]] = []
    for index, row_text in selected_pairs:
        sample = base.parse_sample(row_text, index)
        sample_trace_dir = None if args.trace_dir is None else args.trace_dir / f"sample_{index}"
        output_rows.append(base.evaluate_sample(sample, not args.no_llm, trace_dir=sample_trace_dir))
    return output_rows


def make_detailed_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    detailed = []
    for row in rows:
        new_row = {output_key: row[source_key] for output_key, source_key in KEEP_METADATA}
        for key in DETAILED_KEYS:
            new_row[display_name(key)] = row[key]
        new_row["Notes"] = row["Notes"]
        detailed.append(new_row)
    return detailed


def subsection_order() -> list[str]:
    seen = []
    for key in DETAILED_KEYS:
        subsection = SECTION_INFO[key][1]
        if subsection not in seen:
            seen.append(subsection)
    return sorted(seen, key=section_sort_key)


def requirement_keys_for_category(category: str) -> list[str]:
    return [key for key in DETAILED_KEYS if SECTION_INFO[key][2] == category]


def make_subsection_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[str]] = defaultdict(list)
    for key in DETAILED_KEYS:
        grouped[SECTION_INFO[key][1]].append(key)
    out_rows = []
    for row in rows:
        out = {output_key: row[source_key] for output_key, source_key in KEEP_METADATA}
        for subsection in subsection_order():
            keys = grouped[subsection]
            values = [VERDICT_SCORE[row[k]] for k in keys if row.get(k) in VERDICT_SCORE]
            if values:
                out[subsection] = sum(values) / len(values)
            elif all(row.get(k) == base.NA for k in keys):
                out[subsection] = base.NA
            elif any(row.get(k) == base.UNCLEAR for k in keys):
                out[subsection] = base.UNCLEAR
            else:
                out[subsection] = None
        out_rows.append(out)
    return out_rows


def count_statuses(row: dict[str, Any], keys: set[str]) -> tuple[int, int, int]:
    fail_count = sum(1 for key in keys if row.get(key) == base.FAIL)
    partial_count = sum(1 for key in keys if row.get(key) == base.PARTIAL)
    pass_count = sum(1 for key in keys if row.get(key) == base.PASS)
    return fail_count, partial_count, pass_count


def parse_runtime_pass_rate(value: Any) -> tuple[int, int]:
    if not isinstance(value, str) or "/" not in value:
        return (0, 0)
    left, right = value.split("/", 1)
    try:
        return int(left), int(right)
    except ValueError:
        return (0, 0)


def pass_fraction(value: Any) -> float:
    passed, total = parse_runtime_pass_rate(value)
    return (passed / total) if total else 0.0


def classify_prompt(row: dict[str, Any]) -> str:
    decisive_unusable = {
        "p_function_signature_present",
    }
    major_contract = {
        "v_prompt_test_solution_aligned",
        "v_model_breaking_prompt_defined_only",
        "v_no_unmentioned_internal_logic",
    }
    major_quality = {
        "p_no_unrealistic_constraints",
        "p_not_vague",
        "p_no_conflicting_objectives",
        "v_output_schema_aligned",
    }
    prompt_keys = set(requirement_keys_for_category("Prompt")) | major_contract | major_quality
    fail_count, partial_count, _ = count_statuses(row, prompt_keys)
    if any(row.get(key) == base.FAIL for key in decisive_unusable):
        return "Unusable"
    if row.get("p_practical_algorithmic_problem") == base.FAIL and row.get("p_not_vague") == base.FAIL:
        return "Unusable"
    if row.get("p_no_unrealistic_constraints") == base.FAIL and row.get("p_not_vague") == base.FAIL and row.get("p_no_conflicting_objectives") == base.FAIL:
        return "Unusable"
    if fail_count >= 7 and sum(1 for key in major_quality if row.get(key) == base.FAIL) >= 3:
        return "Unusable"
    if fail_count or partial_count:
        return "Needs Fixing"
    return "Usable"


def classify_ideal(row: dict[str, Any]) -> str:
    severe = {
        "i_executes_without_error",
        "i_passes_internal_tests",
        "i_stdlib_only",
    }
    major = {
        "i_no_parallelism",
        "i_deterministic_solution",
        "i_single_entry_aligned",
    }
    ideal_keys = set(requirement_keys_for_category("Ideal Response")) | severe | major
    fail_count, partial_count, _ = count_statuses(row, ideal_keys)
    runtime_fraction = pass_fraction(row.get("RuntimePassRate"))
    if row.get("i_executes_without_error") == base.FAIL:
        return "Unusable"
    if row.get("i_passes_internal_tests") == base.FAIL or runtime_fraction < 0.5:
        return "Unusable"
    if row.get("i_stdlib_only") == base.FAIL and row.get("i_passes_internal_tests") != base.PASS:
        return "Unusable"
    if fail_count >= 5 and sum(1 for key in major if row.get(key) == base.FAIL) >= 2:
        return "Unusable"
    if fail_count or partial_count or runtime_fraction < 1.0:
        return "Needs Fixing"
    return "Usable"


def classify_tests(row: dict[str, Any]) -> str:
    decisive_unusable = {
        "t_json_encoded",
        "t_string_fields",
    }
    major_alignment = {
        "v_model_breaking_prompt_defined_only",
        "v_no_unmentioned_internal_logic",
        "v_prompt_test_solution_aligned",
        "v_no_extra_parameters",
    }
    test_keys = set(requirement_keys_for_category("Test Cases")) | major_alignment
    fail_count, partial_count, _ = count_statuses(row, test_keys)
    if any(row.get(key) == base.FAIL for key in decisive_unusable):
        return "Unusable"
    if sum(1 for key in major_alignment if row.get(key) == base.FAIL) >= 3:
        return "Unusable"
    if row.get("v_prompt_test_solution_aligned") == base.FAIL and row.get("v_no_extra_parameters") == base.FAIL:
        return "Unusable"
    if fail_count >= 8 and sum(1 for key in {"t_input_json_object", "t_output_json_object", "t_json_escaping_valid"} if row.get(key) == base.FAIL) >= 2 and row.get("v_model_breaking_prompt_defined_only") == base.FAIL:
        return "Unusable"
    if fail_count or partial_count:
        return "Needs Fixing"
    return "Usable"


def make_category_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out_rows = []
    for row in rows:
        out_rows.append(
            {
                **{output_key: row[source_key] for output_key, source_key in KEEP_METADATA},
                "Prompt": classify_prompt(row),
                "Ideal_Response": classify_ideal(row),
                "Test_Cases": classify_tests(row),
            }
        )
    return out_rows


def write_sheet(ws, rows: list[dict[str, Any]], percentage_columns: set[str] | None = None) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in rows:
        ws.append([row.get(header) for header in headers])
    ws.freeze_panes = "A2"
    percentage_columns = percentage_columns or set()
    for idx, header in enumerate(headers, start=1):
        if header in percentage_columns:
            for col in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for cell in col:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0%"
        if header in {"Prompt", "Ideal_Response", "Test_Cases"}:
            for col in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for cell in col:
                    if cell.value == "Usable":
                        cell.fill = PatternFill("solid", fgColor="C6E0B4")
                    elif cell.value == "Needs Fixing":
                        cell.fill = PatternFill("solid", fgColor="FFF2CC")
                    elif cell.value == "Unusable":
                        cell.fill = PatternFill("solid", fgColor="F4CCCC")
        ws.column_dimensions[get_column_letter(idx)].width = max(14, min(40, len(header) + 2))


def write_workbook(detailed_rows: list[dict[str, Any]], subsection_rows: list[dict[str, Any]], category_rows: list[dict[str, Any]], output_path: Path) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Detailed"
    write_sheet(ws1, detailed_rows)
    ws2 = wb.create_sheet("Subsections")
    write_sheet(ws2, subsection_rows, percentage_columns=set(subsection_order()))
    ws3 = wb.create_sheet("Summary")
    write_sheet(ws3, category_rows)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate the multi-tab guideline audit workbook.")
    parser.add_argument("--jsonl", type=Path, default=Path("Samples.jsonl"))
    parser.add_argument("--output", type=Path, default=Path("guideline_audit.xlsx"))
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--offset", type=int, default=0)
    parser.add_argument("--indices", default=None)
    parser.add_argument("--trace-dir", type=Path, default=None)
    parser.add_argument("--no-llm", action="store_true")
    args = parser.parse_args()

    raw_rows = collect_rows(args)
    detailed_rows = make_detailed_rows(raw_rows)
    subsection_rows = make_subsection_rows(raw_rows)
    category_rows = make_category_rows(raw_rows)
    write_workbook(detailed_rows, subsection_rows, category_rows, args.output)
    print(f"Wrote workbook to {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
