from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from turing_takehome.llm import override_stage_targets

from .audit_core import schema
from .audit_core.workbook import build_summary_rows, write_workbook
from .runner import ARTIFACTS_DIR, OUTPUTS_DIR, PROJECT_ROOT, run_cli


DEFAULT_BATCH_ROOT = PROJECT_ROOT / "outputs" / "sample_requirements_analysis_batches"
DEFAULT_AGGREGATED_OUTPUT = OUTPUTS_DIR / "guideline_audit_full.xlsx"


def build_batch_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Run Stage 1 in resumable sample batches."
    )
    parser.add_argument(
        "--jsonl",
        type=Path,
        default=ARTIFACTS_DIR / "provided" / "Samples.jsonl",
    )
    parser.add_argument("--batch-size", type=int, default=10)
    parser.add_argument("--output-root", type=Path, default=DEFAULT_BATCH_ROOT)
    parser.add_argument("--target-name", default=None)
    parser.add_argument("--start-index", type=int, default=0)
    parser.add_argument("--end-index", type=int, default=None)
    return parser


def build_aggregate_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Aggregate Stage 1 batch workbooks into a single workbook."
    )
    parser.add_argument("--batch-root", type=Path, default=DEFAULT_BATCH_ROOT)
    parser.add_argument("--output-workbook", type=Path, default=DEFAULT_AGGREGATED_OUTPUT)
    return parser


def run_batch_cli(argv: list[str] | None = None) -> int:
    args = build_batch_parser().parse_args(argv)
    output_root = _resolve_path(args.output_root)
    output_root.mkdir(parents=True, exist_ok=True)
    sample_count = _count_samples(args.jsonl)
    start_index = max(0, args.start_index)
    end_index = sample_count if args.end_index is None else min(args.end_index, sample_count)
    if args.target_name:
        override_stage_targets(
            "sample-requirements-analysis",
            primary_target=args.target_name,
            comparison_targets=(args.target_name,),
        )
    for batch_start in range(start_index, end_index, args.batch_size):
        batch_end = min(batch_start + args.batch_size, end_index)
        batch_indices = list(range(batch_start, batch_end))
        batch_dir = output_root / f"batch_{batch_start:03d}_{batch_end - 1:03d}"
        workbook_path = batch_dir / schema.OUTPUT_WORKBOOK_NAME
        if workbook_path.exists():
            print(f"Skipping existing batch {batch_dir.name}")
            continue
        batch_dir.mkdir(parents=True, exist_ok=True)
        batch_args = [
            "--jsonl",
            str(args.jsonl),
            "--indices",
            ",".join(str(index) for index in batch_indices),
            "--output-dir",
            str(batch_dir),
        ]
        run_cli(batch_args)
    return 0


def run_aggregate_batches_cli(argv: list[str] | None = None) -> int:
    args = build_aggregate_parser().parse_args(argv)
    batch_root = _resolve_path(args.batch_root)
    batch_dirs = sorted(path for path in batch_root.iterdir() if path.is_dir())
    detailed_rows: list[dict[str, Any]] = []
    for batch_dir in batch_dirs:
        workbook_path = batch_dir / schema.OUTPUT_WORKBOOK_NAME
        if not workbook_path.exists():
            continue
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        detailed_rows.extend(_read_sheet_rows(wb, "Detailed"))
    detailed_rows = _dedupe_rows(detailed_rows, key="Index")
    summary_rows = build_summary_rows(detailed_rows)
    output_workbook = _resolve_path(args.output_workbook)
    output_workbook.parent.mkdir(parents=True, exist_ok=True)
    write_workbook(detailed_rows, [], summary_rows, output_workbook)
    print(f"Wrote aggregated workbook to {output_workbook}")
    return 0


def _read_sheet_rows(workbook, sheet_name: str) -> list[dict[str, Any]]:
    ws = workbook[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    return [dict(zip(headers, row)) for row in rows[1:]]


def _dedupe_rows(rows: list[dict[str, Any]], *, key: str) -> list[dict[str, Any]]:
    by_key: dict[Any, dict[str, Any]] = {}
    for row in rows:
        by_key[row[key]] = row
    return [by_key[index] for index in sorted(by_key)]


def _count_samples(jsonl_path: Path) -> int:
    return len([line for line in jsonl_path.read_text(encoding="utf-8").splitlines() if line.strip()])


def _resolve_path(path: Path) -> Path:
    return path if path.is_absolute() else (PROJECT_ROOT / path).resolve()
