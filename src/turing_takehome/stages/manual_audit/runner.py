from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from turing_takehome.reporting import export_combined_report
from turing_takehome.stages.sample_efficacy_analysis.data import load_samples


PROJECT_ROOT = Path(__file__).resolve().parents[4]
ARTIFACTS_DIR = PROJECT_ROOT / "artifacts"
OUTPUT_DIR = PROJECT_ROOT / "outputs" / "manual_audit"
STAGE1_WORKBOOK = PROJECT_ROOT / "outputs" / "sample_requirements_analysis" / "guideline_audit.xlsx"
STAGE2_DIR = PROJECT_ROOT / "outputs" / "sample_efficacy_analysis"
STAGE3_DIR = PROJECT_ROOT / "outputs" / "dataset_analysis"

STAGE4_TEST_COLUMNS = [
    "BenchmarkTrustCheck",
    "FailureAttribution",
    "PipelineCalibrationCheck",
    "FinalAction",
]

FINDING_SLOTS = (1, 2, 3)
_ILLEGAL_XLSX_CHARS = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Run Stage 4 Manual Audit preparation or finalize completed reviews."
    )
    parser.add_argument(
        "--jsonl",
        type=Path,
        default=ARTIFACTS_DIR / "provided" / "Samples.jsonl",
    )
    parser.add_argument("--stage1-workbook", type=Path, default=STAGE1_WORKBOOK)
    parser.add_argument("--stage2-dir", type=Path, default=STAGE2_DIR)
    parser.add_argument("--stage3-dir", type=Path, default=STAGE3_DIR)
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    parser.add_argument("--review-input", type=Path, default=None)
    parser.add_argument("--max-total", type=int, default=26)
    parser.add_argument("--contradictions", type=int, default=8)
    parser.add_argument("--disagreement", type=int, default=8)
    parser.add_argument("--redundancy", type=int, default=6)
    parser.add_argument("--baseline", type=int, default=4)
    parser.add_argument("--indices", default=None)
    return parser


def run_cli(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    jsonl_path = _resolve_path(args.jsonl)
    stage1_workbook = _resolve_path(args.stage1_workbook)
    stage2_dir = _resolve_path(args.stage2_dir)
    stage3_dir = _resolve_path(args.stage3_dir)
    output_dir = _resolve_path(args.output_dir)
    review_input = _resolve_path(args.review_input) if args.review_input else None

    output_dir.mkdir(parents=True, exist_ok=True)

    stage1_rows = _load_stage1(stage1_workbook)
    stage2_rows = _load_stage2(stage2_dir)
    stage2_test_evidence = _load_stage2_test_evidence(stage2_dir, stage2_rows)
    stage3_context = _load_stage3(stage3_dir)
    sample_records = {
        sample.index: sample
        for sample in load_samples(jsonl_path, indices=args.indices, limit=None, offset=0)
    }

    selected_rows = _select_review_candidates(
        sample_records=sample_records,
        stage1_rows=stage1_rows,
        stage2_rows=stage2_rows,
        stage2_test_evidence=stage2_test_evidence,
        stage3_context=stage3_context,
        max_total=args.max_total,
        contradictions=args.contradictions,
        disagreement=args.disagreement,
        redundancy=args.redundancy,
        baseline=args.baseline,
    )
    reviews_by_index = _load_review_input(review_input) if review_input and review_input.exists() else {}
    selected_rows = _augment_selected_rows_with_reviewed_indices(
        selected_rows=selected_rows,
        reviewed_indices=sorted(reviews_by_index),
        sample_records=sample_records,
        stage1_rows=stage1_rows,
        stage2_rows=stage2_rows,
        stage2_test_evidence=stage2_test_evidence,
        stage3_context=stage3_context,
    )
    detailed_rows = _build_detailed_rows(selected_rows, reviews_by_index)
    summary_rows = _build_summary_rows(selected_rows, detailed_rows)
    review_packet = _build_review_packet(selected_rows)
    run_manifest = _build_run_manifest(
        args=args,
        jsonl_path=jsonl_path,
        stage1_workbook=stage1_workbook,
        stage2_dir=stage2_dir,
        stage3_dir=stage3_dir,
        output_dir=output_dir,
        selected_rows=selected_rows,
        review_input=review_input,
    )

    payload = {
        "stage4": {
            "detailed": detailed_rows,
            "summary": summary_rows,
            "test_columns": STAGE4_TEST_COLUMNS,
            "run_manifest": run_manifest,
        }
    }

    _write_json(output_dir / "manual_audit.json", payload)
    _write_json(output_dir / "review_packet.json", {"reviews": review_packet})
    _write_csv(output_dir / "review_template.csv", review_packet)
    _write_csv(output_dir / "detailed_rows.csv", detailed_rows)
    _write_json(output_dir / "run_manifest.json", run_manifest)
    _write_markdown(output_dir / "summary.md", summary_rows, selected_rows)
    _write_workbook(
        output_dir / "manual_audit.xlsx",
        summary_rows=summary_rows,
        detailed_rows=detailed_rows,
    )
    workbook_path, json_path = export_combined_report(
        stage1_workbook_path=stage1_workbook,
        stage2_output_root=stage2_dir,
        stage3_output_root=stage3_dir,
        stage4_output_root=output_dir,
    )

    print(f"Wrote Stage 4 payload to {output_dir / 'manual_audit.json'}")
    print(f"Wrote review packet to {output_dir / 'review_packet.json'}")
    print(f"Wrote workbook to {output_dir / 'manual_audit.xlsx'}")
    print(f"Wrote combined workbook to {workbook_path}")
    print(f"Wrote combined json to {json_path}")
    return 0


def _select_review_candidates(
    *,
    sample_records: dict[int, Any],
    stage1_rows: dict[int, dict[str, Any]],
    stage2_rows: dict[int, dict[str, Any]],
    stage2_test_evidence: dict[int, dict[str, Any]],
    stage3_context: dict[str, Any],
    max_total: int,
    contradictions: int,
    disagreement: int,
    redundancy: int,
    baseline: int,
) -> list[dict[str, Any]]:
    detailed_rows = {int(row["Index"]): row for row in stage3_context.get("detailed", [])}
    queue_rows = stage3_context.get("audit_queues", [])
    selected_indices: set[int] = set()
    selected: list[dict[str, Any]] = []

    def add_row(index: int, bucket: str, reason: str) -> None:
        if index in selected_indices or index not in sample_records or index not in detailed_rows:
            return
        stage3_row = detailed_rows[index]
        stage2_row = stage2_rows.get(index, {})
        stage1_row = stage1_rows.get(index, {})
        stage2_evidence = stage2_test_evidence.get(index, {})
        sample = sample_records[index]
        selected_indices.add(index)
        selected.append(
            {
                "Index": index,
                "QuestionId": sample.row.get("question_id", ""),
                "QuestionTitle": sample.row.get("question_title", ""),
                "Difficulty": sample.row.get("difficulty", ""),
                "SelectionBucket": bucket,
                "SelectionReason": reason,
                "Stage3AuditPriority": stage3_row.get("AuditPriority", ""),
                "Stage3AuditReason": stage3_row.get("AuditReason", ""),
                "Stage3ModelDisagreementSource": stage3_row.get("ModelDisagreementSource", ""),
                "PipelineUtilityLabel": _pipeline_utility_label(stage3_row),
                "Stage1Prompt": stage1_row.get("Prompt", ""),
                "Stage1IdealResponse": stage1_row.get("Ideal_Response", ""),
                "Stage1TestCases": stage1_row.get("Test_Cases", ""),
                "Stage2EfficacyLabel": stage2_row.get("EfficacyLabel", ""),
                "Stage2BenchmarkQualitySignal": stage2_row.get("BenchmarkQualitySignal", ""),
                "Stage2FailureCategory": stage2_row.get("FailureCategory", ""),
                "WinnerCombinedPassRate": stage2_row.get("BestCombinedPassRate", ""),
                "OraclePassRate": stage2_row.get("OraclePassRate", ""),
                "ReviewContext": _review_context(stage3_row, stage2_row),
                "ObservedTestEvidence": stage2_evidence.get("summary", ""),
                "ObservedFailedTests": stage2_evidence.get("failed_tests", []),
                "PromptExcerpt": _excerpt(sample.question_content, 700),
                "StarterCodeExcerpt": _excerpt(sample.starter_code, 400),
                "IdealResponseExcerpt": _excerpt(sample.ideal_response, 500),
                "_full_prompt": sample.question_content,
                "_full_starter_code": sample.starter_code,
                "_full_ideal_response": sample.ideal_response,
            }
        )

    contradictory = sorted(
        (
            int(row["Index"])
            for row in queue_rows
            if str(row.get("Queue", "")) == "contradictory_candidates"
        )
    )
    for index in contradictory[:contradictions]:
        add_row(index, "contradiction", "Stage 3 contradiction candidate")

    disagreement_candidates = sorted(
        (
            row for row in detailed_rows.values()
            if row.get("ModelDisagreementCheck") == "FLAG" or row.get("AttemptVarianceCheck") == "FLAG"
        ),
        key=lambda row: (
            _priority_rank(str(row.get("AuditPriority", ""))),
            str(row.get("ModelDisagreementCheck", "")) != "FLAG",
            str(row.get("AttemptVarianceCheck", "")) != "FLAG",
            int(row["Index"]),
        ),
    )
    for row in disagreement_candidates:
        if sum(1 for item in selected if item["SelectionBucket"] == "disagreement") >= disagreement:
            break
        reason_parts: list[str] = []
        if row.get("ModelDisagreementCheck") == "FLAG":
            reason_parts.append("Stage 3 disagreement")
        if row.get("AttemptVarianceCheck") == "FLAG":
            reason_parts.append("attempt variance")
        add_row(int(row["Index"]), "disagreement", " + ".join(reason_parts))

    redundancy_candidates = sorted(
        (
            row for row in detailed_rows.values()
            if row.get("RedundancyStatus") == "FLAG"
        ),
        key=lambda row: (
            -(_to_int(row.get("RedundancyClusterSize")) or 0),
            _priority_rank(str(row.get("AuditPriority", ""))),
            int(row["Index"]),
        ),
    )
    seen_clusters: set[Any] = set()
    for row in redundancy_candidates:
        if sum(1 for item in selected if item["SelectionBucket"] == "redundancy") >= redundancy:
            break
        cluster_id = row.get("RedundancyClusterId")
        if cluster_id in seen_clusters:
            continue
        seen_clusters.add(cluster_id)
        add_row(
            int(row["Index"]),
            "redundancy",
            f"cluster {cluster_id} representative; {row.get('AuditReason', 'redundancy')}",
        )

    clean_candidates = sorted(
        (
            row for row in detailed_rows.values()
            if row.get("AuditPriority") == "normal"
        ),
        key=lambda row: (
            str(row.get("ExemplarCheck", "")) != "FLAG",
            int(row["Index"]),
        ),
    )
    for row in clean_candidates:
        if sum(1 for item in selected if item["SelectionBucket"] == "baseline") >= baseline:
            break
        add_row(int(row["Index"]), "baseline", "low-risk baseline sanity check")

    if len(selected) < max_total:
        backfill = sorted(
            detailed_rows.values(),
            key=lambda row: (_priority_rank(str(row.get("AuditPriority", ""))), int(row["Index"])),
        )
        for row in backfill:
            if len(selected) >= max_total:
                break
            add_row(int(row["Index"]), "backfill", str(row.get("AuditReason", "")) or "priority backfill")

    return sorted(selected[:max_total], key=lambda row: (row["SelectionBucket"], row["Index"]))


def _build_detailed_rows(
    selected_rows: list[dict[str, Any]],
    reviews_by_index: dict[int, dict[str, Any]],
) -> list[dict[str, Any]]:
    detailed_rows: list[dict[str, Any]] = []
    for row in selected_rows:
        review = reviews_by_index.get(int(row["Index"]), {})
        detailed_row = {
            "Index": row["Index"],
            "QuestionId": row["QuestionId"],
            "QuestionTitle": row["QuestionTitle"],
            "Difficulty": row["Difficulty"],
            "SelectionBucket": row["SelectionBucket"],
            "SelectionReason": row["SelectionReason"],
            "Stage3AuditPriority": row["Stage3AuditPriority"],
            "PipelineUtilityLabel": row["PipelineUtilityLabel"],
            "Stage1Prompt": row["Stage1Prompt"],
            "Stage1IdealResponse": row["Stage1IdealResponse"],
            "Stage1TestCases": row["Stage1TestCases"],
            "Stage2EfficacyLabel": row["Stage2EfficacyLabel"],
            "Stage2BenchmarkQualitySignal": row["Stage2BenchmarkQualitySignal"],
            "Stage2FailureCategory": row["Stage2FailureCategory"],
            "WinnerCombinedPassRate": row["WinnerCombinedPassRate"],
            "OraclePassRate": row["OraclePassRate"],
            "Stage3ModelDisagreementSource": row["Stage3ModelDisagreementSource"],
            "ReviewContext": row["ReviewContext"],
            "ObservedTestEvidence": row.get("ObservedTestEvidence", ""),
            "PromptExcerpt": row["PromptExcerpt"],
            "BenchmarkTrustCheck": review.get("BenchmarkTrustCheck", ""),
            "FailureAttribution": review.get("FailureAttribution", ""),
            "PipelineCalibrationCheck": review.get("PipelineCalibrationCheck", ""),
            "FinalAction": review.get("FinalAction", ""),
            "SummaryConfidence": review.get("SummaryConfidence", ""),
            "ReviewStatus": "completed" if review else "pending",
        }
        for slot in FINDING_SLOTS:
            detailed_row[f"Finding{slot}DefectType"] = review.get(f"Finding{slot}DefectType", "")
            detailed_row[f"Finding{slot}Severity"] = review.get(f"Finding{slot}Severity", "")
            detailed_row[f"Finding{slot}Confidence"] = review.get(f"Finding{slot}Confidence", "")
            detailed_row[f"Notes-Finding{slot}"] = review.get(f"Notes-Finding{slot}", "")
        ordered_row: dict[str, Any] = {}
        leading_columns = [
            "Index",
            "QuestionId",
            "QuestionTitle",
            "Difficulty",
            "SelectionBucket",
            "SelectionReason",
            "Stage3AuditPriority",
            "PipelineUtilityLabel",
            "Stage1Prompt",
            "Stage1IdealResponse",
            "Stage1TestCases",
            "Stage2EfficacyLabel",
            "Stage2BenchmarkQualitySignal",
            "Stage2FailureCategory",
            "WinnerCombinedPassRate",
            "OraclePassRate",
            "Stage3ModelDisagreementSource",
            "ReviewContext",
            "ObservedTestEvidence",
            "PromptExcerpt",
        ]
        for column_name in leading_columns:
            ordered_row[column_name] = detailed_row[column_name]
        for column_name in STAGE4_TEST_COLUMNS:
            ordered_row[column_name] = detailed_row[column_name]
            ordered_row[f"Notes-{column_name}"] = review.get(f"Notes-{column_name}", "")
        ordered_row["SummaryConfidence"] = detailed_row["SummaryConfidence"]
        for slot in FINDING_SLOTS:
            ordered_row[f"Finding{slot}DefectType"] = detailed_row[f"Finding{slot}DefectType"]
            ordered_row[f"Finding{slot}Severity"] = detailed_row[f"Finding{slot}Severity"]
            ordered_row[f"Finding{slot}Confidence"] = detailed_row[f"Finding{slot}Confidence"]
            ordered_row[f"Notes-Finding{slot}"] = detailed_row[f"Notes-Finding{slot}"]
        ordered_row["ReviewStatus"] = detailed_row["ReviewStatus"]
        detailed_rows.append(ordered_row)
    return detailed_rows


def _build_summary_rows(
    selected_rows: list[dict[str, Any]],
    detailed_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    completed_rows = [row for row in detailed_rows if row["ReviewStatus"] == "completed"]
    bucket_counts = Counter(row["SelectionBucket"] for row in selected_rows)
    rows.append(
        {
            "AnalysisArea": "selection",
            "TestName": "pipeline_guided_sampling",
            "Result": "PASS",
            "Evidence": "; ".join(f"{bucket}={count}" for bucket, count in sorted(bucket_counts.items())),
            "Interpretation": "Stage 4 samples are queue-driven rather than randomly sampled.",
            "ContextualCaveat": "This is a precision-oriented audit slice, not an unbiased population estimate.",
            "Recommendation": "",
        }
    )
    rows.append(
        {
            "AnalysisArea": "completion",
            "TestName": "review_completion",
            "Result": "PASS" if len(completed_rows) == len(detailed_rows) and detailed_rows else "REVIEW",
            "Evidence": f"{len(completed_rows)} of {len(detailed_rows)} selected rows contain completed review judgments.",
            "Interpretation": "Stage 4 is ready for calibration analysis." if completed_rows else "Templates have been generated but not yet filled.",
            "ContextualCaveat": "",
            "Recommendation": "Fill the review template and rerun Stage 4." if not completed_rows else "",
        }
    )
    if not completed_rows:
        return rows

    agreement_counter = Counter(
        row["PipelineCalibrationCheck"] for row in completed_rows if row["PipelineCalibrationCheck"]
    )
    rows.append(
        {
            "AnalysisArea": "calibration",
            "TestName": "pipeline_agreement",
            "Result": "INFO",
            "Evidence": _format_counter(agreement_counter),
            "Interpretation": "Human audit estimates where the automated pipeline is aligned, overstated, or missing defects.",
            "ContextualCaveat": "Agreement is measured only on the targeted audit slice.",
            "Recommendation": "",
        }
    )

    action_counter = Counter(row["FinalAction"] for row in completed_rows if row["FinalAction"])
    rows.append(
        {
            "AnalysisArea": "actionability",
            "TestName": "final_actions",
            "Result": "INFO",
            "Evidence": _format_counter(action_counter),
            "Interpretation": "Human review translates audit findings into keep/fix/remove decisions.",
            "ContextualCaveat": "",
            "Recommendation": _top_action_recommendation(action_counter),
        }
    )

    defect_counter: Counter[str] = Counter()
    severity_counter: Counter[str] = Counter()
    finding_confidence_counter: Counter[str] = Counter()
    for row in completed_rows:
        for slot in FINDING_SLOTS:
            defect = str(row.get(f"Finding{slot}DefectType", "")).strip()
            severity = str(row.get(f"Finding{slot}Severity", "")).strip()
            confidence = str(row.get(f"Finding{slot}Confidence", "")).strip()
            if defect:
                defect_counter[defect] += 1
            if severity:
                severity_counter[severity] += 1
            if confidence:
                finding_confidence_counter[confidence] += 1
    rows.append(
        {
            "AnalysisArea": "failure_taxonomy",
            "TestName": "defect_types",
            "Result": "INFO",
            "Evidence": _format_counter(defect_counter),
            "Interpretation": "This is the current human-confirmed benchmark-defect taxonomy over the reviewed slice.",
            "ContextualCaveat": "",
            "Recommendation": "",
        }
    )
    rows.append(
        {
            "AnalysisArea": "severity",
            "TestName": "finding_severities",
            "Result": "INFO",
            "Evidence": _format_counter(severity_counter),
            "Interpretation": "This summarizes the severities attached to structured manual findings.",
            "ContextualCaveat": "",
            "Recommendation": "",
        }
    )

    attribution_counter = Counter(
        row["FailureAttribution"] for row in completed_rows if row["FailureAttribution"]
    )
    rows.append(
        {
            "AnalysisArea": "attribution",
            "TestName": "failure_attribution",
            "Result": "INFO",
            "Evidence": _format_counter(attribution_counter),
            "Interpretation": "Stage 4 separates likely dataset defects from true model failures where the distinction matters.",
            "ContextualCaveat": "",
            "Recommendation": "",
        }
    )

    trust_counter = Counter(row["BenchmarkTrustCheck"] for row in completed_rows if row["BenchmarkTrustCheck"])
    rows.append(
        {
            "AnalysisArea": "trustworthiness",
            "TestName": "benchmark_trust_check",
            "Result": "INFO",
            "Evidence": _format_counter(trust_counter),
            "Interpretation": "This estimates whether the reviewed samples look trustworthy as benchmark items.",
            "ContextualCaveat": "This is measured on a targeted audit slice, not the whole dataset.",
            "Recommendation": "",
        }
    )

    highest_priority = []
    for row in completed_rows:
        high_finding = any(
            row.get(f"Finding{slot}Severity") == "high" for slot in FINDING_SLOTS
        )
        if high_finding or row.get("FinalAction") in {"remove", "fix"}:
            highest_priority.append(row)
    rows.append(
        {
            "AnalysisArea": "policy",
            "TestName": "highest_priority_cases",
            "Result": "REVIEW" if highest_priority else "PASS",
            "Evidence": f"{len(highest_priority)} reviewed samples currently look high-severity or likely need fix/remove action.",
            "Interpretation": "These are the most decision-relevant manual audit outcomes.",
            "ContextualCaveat": "",
            "Recommendation": "Start final adjudication with the high-severity fix/remove set." if highest_priority else "",
        }
    )

    confidence_counter = Counter(
        row.get("SummaryConfidence", "") for row in completed_rows if row.get("SummaryConfidence", "")
    )
    rows.append(
        {
            "AnalysisArea": "confidence",
            "TestName": "summary_confidence",
            "Result": "INFO",
            "Evidence": _format_counter(confidence_counter),
            "Interpretation": "This estimates how decisive the current sample-level judgments are.",
            "ContextualCaveat": "Low-confidence judgments should be treated more as calibration hints than final truth.",
            "Recommendation": "",
        }
    )
    rows.append(
        {
            "AnalysisArea": "confidence",
            "TestName": "finding_confidence",
            "Result": "INFO",
            "Evidence": _format_counter(finding_confidence_counter),
            "Interpretation": "This estimates how decisive the individual manual findings are.",
            "ContextualCaveat": "Low-confidence findings are useful for triage but should not dominate recalibration policy.",
            "Recommendation": "",
        }
    )
    return rows


def _build_review_packet(selected_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    packet: list[dict[str, Any]] = []
    for row in selected_rows:
        packet.append(
            {
                "Index": row["Index"],
                "QuestionId": row["QuestionId"],
                "QuestionTitle": row["QuestionTitle"],
                "SelectionBucket": row["SelectionBucket"],
                "SelectionReason": row["SelectionReason"],
                "ReviewContext": row["ReviewContext"],
                "ObservedTestEvidence": row.get("ObservedTestEvidence", ""),
                "ObservedFailedTests": row.get("ObservedFailedTests", []),
                "Prompt": row["_full_prompt"],
                "StarterCode": row["_full_starter_code"],
                "IdealResponse": row["_full_ideal_response"],
                "BenchmarkTrustCheck": "",
                "Notes-BenchmarkTrustCheck": "",
                "FailureAttribution": "",
                "Notes-FailureAttribution": "",
                "PipelineCalibrationCheck": "",
                "Notes-PipelineCalibrationCheck": "",
                "FinalAction": "",
                "Notes-FinalAction": "",
                "SummaryConfidence": "",
            }
        )
        for slot in FINDING_SLOTS:
            packet[-1][f"Finding{slot}DefectType"] = ""
            packet[-1][f"Finding{slot}Severity"] = ""
            packet[-1][f"Finding{slot}Confidence"] = ""
            packet[-1][f"Notes-Finding{slot}"] = ""
    return packet


def _load_review_input(path: Path) -> dict[int, dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        with path.open(encoding="utf-8", newline="") as handle:
            rows = list(csv.DictReader(handle))
    else:
        payload = json.loads(path.read_text(encoding="utf-8"))
        rows = payload.get("reviews", payload) if isinstance(payload, dict) else payload
    result: dict[int, dict[str, Any]] = {}
    for row in rows:
        index = _to_int(row.get("Index", row.get("index")))
        if index is None:
            continue
        mapped = {
            "BenchmarkTrustCheck": str(
                row.get("BenchmarkTrustCheck", row.get("GroundTruthValidity", ""))
            ).strip(),
            "Notes-BenchmarkTrustCheck": str(
                row.get("Notes-BenchmarkTrustCheck", row.get("Notes-GroundTruthValidity", ""))
            ).strip(),
            "FailureAttribution": str(row.get("FailureAttribution", "")).strip(),
            "Notes-FailureAttribution": str(row.get("Notes-FailureAttribution", "")).strip(),
            "PipelineCalibrationCheck": str(
                row.get("PipelineCalibrationCheck", row.get("LabelCalibrationCheck", ""))
            ).strip(),
            "Notes-PipelineCalibrationCheck": str(
                row.get("Notes-PipelineCalibrationCheck", row.get("Notes-LabelCalibrationCheck", ""))
            ).strip(),
            "FinalAction": str(row.get("FinalAction", "")).strip(),
            "Notes-FinalAction": str(row.get("Notes-FinalAction", "")).strip(),
            "SummaryConfidence": str(
                row.get("SummaryConfidence", row.get("ReviewConfidence", ""))
            ).strip(),
        }
        legacy_defect = str(row.get("DefectType", row.get("ErrorType", ""))).strip()
        legacy_defect_note = str(row.get("Notes-DefectType", "")).strip()
        legacy_severity = str(row.get("Severity", "")).strip()
        legacy_confidence = str(row.get("ReviewConfidence", "")).strip()
        for slot in FINDING_SLOTS:
            mapped[f"Finding{slot}DefectType"] = str(row.get(f"Finding{slot}DefectType", "")).strip()
            mapped[f"Finding{slot}Severity"] = str(row.get(f"Finding{slot}Severity", "")).strip()
            mapped[f"Finding{slot}Confidence"] = str(row.get(f"Finding{slot}Confidence", "")).strip()
            mapped[f"Notes-Finding{slot}"] = str(row.get(f"Notes-Finding{slot}", "")).strip()
        if legacy_defect and not mapped["Finding1DefectType"]:
            mapped["Finding1DefectType"] = legacy_defect
            mapped["Finding1Severity"] = legacy_severity
            mapped["Finding1Confidence"] = legacy_confidence
            mapped["Notes-Finding1"] = legacy_defect_note
        result[index] = mapped
    return result


def _build_run_manifest(
    *,
    args: argparse.Namespace,
    jsonl_path: Path,
    stage1_workbook: Path,
    stage2_dir: Path,
    stage3_dir: Path,
    output_dir: Path,
    selected_rows: list[dict[str, Any]],
    review_input: Path | None,
) -> dict[str, Any]:
    return {
        "stage": "manual-audit",
        "jsonl_path": str(jsonl_path),
        "stage1_workbook": str(stage1_workbook),
        "stage2_dir": str(stage2_dir),
        "stage3_dir": str(stage3_dir),
        "output_dir": str(output_dir),
        "review_input": str(review_input) if review_input else "",
        "max_total": args.max_total,
        "bucket_targets": {
            "contradictions": args.contradictions,
            "disagreement": args.disagreement,
            "redundancy": args.redundancy,
            "baseline": args.baseline,
        },
        "selected_indices": [int(row["Index"]) for row in selected_rows],
        "selected_buckets": Counter(row["SelectionBucket"] for row in selected_rows),
    }


def _augment_selected_rows_with_reviewed_indices(
    *,
    selected_rows: list[dict[str, Any]],
    reviewed_indices: list[int],
    sample_records: dict[int, Any],
    stage1_rows: dict[int, dict[str, Any]],
    stage2_rows: dict[int, dict[str, Any]],
    stage2_test_evidence: dict[int, dict[str, Any]],
    stage3_context: dict[str, Any],
) -> list[dict[str, Any]]:
    existing = {int(row["Index"]) for row in selected_rows}
    detailed_rows = {int(row["Index"]): row for row in stage3_context.get("detailed", [])}
    augmented = list(selected_rows)
    for index in reviewed_indices:
        if index in existing or index not in sample_records:
            continue
        sample = sample_records[index]
        stage3_row = detailed_rows.get(index, {})
        stage2_row = stage2_rows.get(index, {})
        stage1_row = stage1_rows.get(index, {})
        stage2_evidence = stage2_test_evidence.get(index, {})
        augmented.append(
            {
                "Index": index,
                "QuestionId": sample.row.get("question_id", ""),
                "QuestionTitle": sample.row.get("question_title", ""),
                "Difficulty": sample.row.get("difficulty", ""),
                "SelectionBucket": "manual_lookup",
                "SelectionReason": "reviewed outside default packet",
                "Stage3AuditPriority": stage3_row.get("AuditPriority", ""),
                "Stage3AuditReason": stage3_row.get("AuditReason", ""),
                "Stage3ModelDisagreementSource": stage3_row.get("ModelDisagreementSource", ""),
                "PipelineUtilityLabel": _pipeline_utility_label(stage3_row),
                "Stage1Prompt": stage1_row.get("Prompt", ""),
                "Stage1IdealResponse": stage1_row.get("Ideal_Response", ""),
                "Stage1TestCases": stage1_row.get("Test_Cases", ""),
                "Stage2EfficacyLabel": stage2_row.get("EfficacyLabel", ""),
                "Stage2BenchmarkQualitySignal": stage2_row.get("BenchmarkQualitySignal", ""),
                "Stage2FailureCategory": stage2_row.get("FailureCategory", ""),
                "WinnerCombinedPassRate": stage2_row.get("BestCombinedPassRate", ""),
                "OraclePassRate": stage2_row.get("OraclePassRate", ""),
                "ReviewContext": _review_context(stage3_row, stage2_row),
                "ObservedTestEvidence": stage2_evidence.get("summary", ""),
                "ObservedFailedTests": stage2_evidence.get("failed_tests", []),
                "PromptExcerpt": _excerpt(sample.question_content, 700),
                "StarterCodeExcerpt": _excerpt(sample.starter_code, 400),
                "IdealResponseExcerpt": _excerpt(sample.ideal_response, 500),
                "_full_prompt": sample.question_content,
                "_full_starter_code": sample.starter_code,
                "_full_ideal_response": sample.ideal_response,
            }
        )
    return sorted(augmented, key=lambda row: (str(row.get("SelectionBucket", "")), int(row["Index"])))


def _pipeline_utility_label(stage3_row: dict[str, Any]) -> str:
    if stage3_row.get("BenchmarkDefectCandidate") == "FLAG":
        return "caveated"
    if stage3_row.get("TrivialityCheck") == "FLAG":
        return "saturated"
    if str(stage3_row.get("ContradictionCheck", "")) not in {"", "none"}:
        return "contradictory"
    if stage3_row.get("ExemplarCheck") == "FLAG":
        return "strong"
    return "usable"


def _review_context(stage3_row: dict[str, Any], stage2_row: dict[str, Any]) -> str:
    parts = [
        f"Stage 2: {stage2_row.get('EfficacyLabel', '') or 'n/a'}",
        f"benchmark_quality={stage2_row.get('BenchmarkQualitySignal', '') or 'n/a'}",
        f"failure_category={stage2_row.get('FailureCategory', '') or 'n/a'}",
        f"Stage 3 reason={stage3_row.get('AuditReason', '') or 'n/a'}",
    ]
    return "; ".join(parts)


def _load_stage2_test_evidence(
    stage2_dir: Path,
    stage2_rows: dict[int, dict[str, Any]],
) -> dict[int, dict[str, Any]]:
    path = stage2_dir / "per_test_results.jsonl"
    if not path.exists():
        return {}

    grouped: dict[tuple[int, str, int], list[dict[str, Any]]] = defaultdict(list)
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            raw = line.strip()
            if not raw:
                continue
            row = json.loads(raw)
            sample_index = _to_int(row.get("sample_index"))
            attempt_index = _to_int(row.get("attempt_index"))
            target_name = str(row.get("target_name", "")).strip()
            if sample_index is None or attempt_index is None or not target_name:
                continue
            grouped[(sample_index, target_name, attempt_index)].append(row)

    evidence: dict[int, dict[str, Any]] = {}
    for index, stage2_row in stage2_rows.items():
        target_name = str(stage2_row.get("TargetName", "")).strip()
        best_attempt = _to_int(stage2_row.get("BestAttemptIndex"))
        candidate_attempts: list[int] = []
        if best_attempt is not None:
            candidate_attempts.extend([best_attempt, best_attempt - 1])
        selected_rows: list[dict[str, Any]] = []
        for attempt in candidate_attempts:
            if attempt is None or attempt < 0:
                continue
            selected_rows = grouped.get((index, target_name, attempt), [])
            if selected_rows:
                break
        if not selected_rows:
            fallback = [
                rows
                for (sample_index, row_target, _attempt), rows in grouped.items()
                if sample_index == index and row_target == target_name
            ]
            if fallback:
                selected_rows = max(fallback, key=len)
        evidence[index] = _summarize_test_rows(selected_rows)
    return evidence


def _summarize_test_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    if not rows:
        return {"summary": "No per-test evidence was found for the selected Stage 2 attempt.", "failed_tests": []}

    total = len(rows)
    failed_rows = [
        row
        for row in rows
        if str(row.get("status", "")).strip().lower() not in {"ok", "pass", "passed", "success"}
    ]
    visibility_counts = Counter(str(row.get("visibility", "")).strip() or "unknown" for row in rows)
    failure_counts = Counter(str(row.get("failure_type", "")).strip() or "mismatch" for row in failed_rows)
    summary_parts = [
        f"{len(failed_rows)} of {total} tests failed on the selected Stage 2 attempt.",
        "Test mix: " + _format_counter(visibility_counts) if visibility_counts else "",
        "Failure types: " + _format_counter(failure_counts) if failure_counts else "No failing tests were recorded.",
    ]
    failed_tests: list[str] = []
    for row in failed_rows[:8]:
        visibility = str(row.get("visibility", "")).strip() or "unknown"
        case_index = _to_int(row.get("case_index"))
        label = f"{visibility} test {case_index}" if case_index is not None else f"{visibility} test"
        failure_type = str(row.get("failure_type", "")).strip() or "mismatch"
        focus = str(row.get("focus", "")).strip()
        exception_message = str(row.get("exception_message", "")).strip()
        detail_bits = [failure_type]
        if focus:
            detail_bits.append(focus)
        if exception_message:
            detail_bits.append(exception_message[:180])
        failed_tests.append(f"{label}: " + " | ".join(detail_bits))
    return {
        "summary": " ".join(part for part in summary_parts if part),
        "failed_tests": failed_tests,
    }


def _top_action_recommendation(action_counter: Counter[str]) -> str:
    if action_counter.get("remove", 0):
        return "Validate the remove candidates first; they have the largest benchmark-quality impact."
    if action_counter.get("fix", 0):
        return "Prioritize fix candidates before broadening manual audit coverage."
    if action_counter.get("keep", 0):
        return "Use keep decisions as calibration exemplars for earlier stages."
    return ""


def _load_stage1(workbook_path: Path) -> dict[int, dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows = _read_sheet_rows(workbook, "Summary")
    return {int(row["Index"]): row for row in rows}


def _load_stage2(stage2_dir: Path) -> dict[int, dict[str, Any]]:
    rows = _read_csv_rows(stage2_dir / "sample_model_results.csv")
    best_by_index: dict[int, dict[str, Any]] = {}
    for row in rows:
        index = int(row["Index"])
        existing = best_by_index.get(index)
        if existing is None or _to_float(row.get("BestCombinedPassRate")) > _to_float(existing.get("BestCombinedPassRate")):
            best_by_index[index] = row
    return best_by_index


def _load_stage3(stage3_dir: Path) -> dict[str, Any]:
    payload = json.loads((stage3_dir / "dataset_analysis.json").read_text(encoding="utf-8"))
    return payload.get("stage3", {})


def _write_workbook(
    path: Path,
    *,
    summary_rows: list[dict[str, Any]],
    detailed_rows: list[dict[str, Any]],
) -> None:
    workbook = Workbook()
    ws_summary = workbook.active
    ws_summary.title = "Summary"
    _write_sheet(ws_summary, summary_rows, percentage_columns=set())

    ws_detailed = workbook.create_sheet("Detailed")
    _write_sheet(ws_detailed, detailed_rows, percentage_columns={"WinnerCombinedPassRate", "OraclePassRate"})

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _write_markdown(path: Path, summary_rows: list[dict[str, Any]], selected_rows: list[dict[str, Any]]) -> None:
    lines = [
        "# Stage 4 Manual Audit",
        "",
        f"- Selected samples: {len(selected_rows)}",
        f"- Buckets: {_format_counter(Counter(row['SelectionBucket'] for row in selected_rows))}",
        "",
        "## Summary",
    ]
    for row in summary_rows:
        lines.append(
            f"- {row['AnalysisArea']} / {row['TestName']}: {row['Evidence']} "
            f"[result={row['Result']}]"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_sheet(ws, rows: list[dict[str, Any]], *, percentage_columns: set[str]) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in rows:
        ws.append([_excel_safe(row.get(header)) for header in headers])
    ws.freeze_panes = "A2"
    for column_index, header in enumerate(headers, start=1):
        if header in percentage_columns:
            for column_cells in ws.iter_cols(min_col=column_index, max_col=column_index, min_row=2):
                for cell in column_cells:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.0%"
        ws.column_dimensions[_column_letter(column_index)].width = max(16, min(46, len(header) + 4))


def _read_sheet_rows(workbook, sheet_name: str) -> list[dict[str, Any]]:
    worksheet = workbook[sheet_name]
    values = list(worksheet.iter_rows(values_only=True))
    if not values:
        return []
    headers = [str(value) for value in values[0]]
    rows: list[dict[str, Any]] = []
    for row in values[1:]:
        payload = {header: value for header, value in zip(headers, row)}
        if all(value in (None, "") for value in payload.values()):
            continue
        rows.append(payload)
    return rows


def _read_csv_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [
            {str(key).lstrip("\ufeff"): value for key, value in row.items()}
            for row in csv.DictReader(handle)
        ]


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def _resolve_path(path: Path) -> Path:
    return path if path.is_absolute() else (PROJECT_ROOT / path).resolve()


def _excerpt(text: str, limit: int) -> str:
    stripped = text.strip()
    if len(stripped) <= limit:
        return stripped
    return stripped[: limit - 3].rstrip() + "..."


def _format_counter(counter: Counter[str] | dict[str, int]) -> str:
    if not counter:
        return "none"
    items = counter.items() if isinstance(counter, dict) else counter.items()
    return "; ".join(f"{key}={value}" for key, value in sorted(items))


def _priority_rank(value: str) -> int:
    order = {"critical": 0, "high": 1, "medium": 2, "normal": 3}
    return order.get(value, 4)


def _to_float(value: Any) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _to_int(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _column_letter(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _excel_safe(value: Any) -> Any:
    if isinstance(value, str):
        return _ILLEGAL_XLSX_CHARS.sub("", value)
    return value
