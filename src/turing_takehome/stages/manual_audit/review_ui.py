from __future__ import annotations

import argparse
import csv
import html
import json
import threading
import urllib.parse
import webbrowser
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from turing_takehome.stages.sample_efficacy_analysis.data import load_samples

from .runner import (
    ARTIFACTS_DIR,
    OUTPUT_DIR,
    STAGE1_WORKBOOK,
    STAGE2_DIR,
    STAGE3_DIR,
    _excerpt,
    _load_stage1,
    _load_stage2,
    _load_stage2_test_evidence,
    _load_stage3,
    _pipeline_utility_label,
    _resolve_path,
    _review_context,
)


SUMMARY_FIELDS = [
    "BenchmarkTrustCheck",
    "Notes-BenchmarkTrustCheck",
    "FailureAttribution",
    "Notes-FailureAttribution",
    "PipelineCalibrationCheck",
    "Notes-PipelineCalibrationCheck",
    "FinalAction",
    "Notes-FinalAction",
    "SummaryConfidence",
]

FINDING_SLOTS = (1, 2, 3)
FINDING_FIELDS: list[str] = []
for _slot in FINDING_SLOTS:
    FINDING_FIELDS.extend(
        [
            f"Finding{_slot}DefectType",
            f"Finding{_slot}Severity",
            f"Finding{_slot}Confidence",
            f"Notes-Finding{_slot}",
        ]
    )
REVIEW_FIELDS = SUMMARY_FIELDS + FINDING_FIELDS

EXEMPLAR_LIST = [104, 141, 137, 145, 11, 17, 87]

LIST_METADATA = {
    "exemplars": {
        "label": "Exemplar Set",
        "description": "Curated set of report-ready exemplar samples spanning the main benchmark failure modes plus one strong positive example.",
        "readonly": False,
    },
}

FIELD_LABELS = {
    "BenchmarkTrustCheck": "Does this sample look trustworthy as a benchmark item?",
    "Notes-BenchmarkTrustCheck": "Why?",
    "FailureAttribution": "If there is a real problem here, is it mostly the dataset or the model?",
    "Notes-FailureAttribution": "Why?",
    "PipelineCalibrationCheck": "Did the automated pipeline diagnose this sample correctly?",
    "Notes-PipelineCalibrationCheck": "Why?",
    "FinalAction": "What should happen to this sample?",
    "Notes-FinalAction": "Why?",
    "SummaryConfidence": "How confident are you in your overall judgment?",
}
for _slot in FINDING_SLOTS:
    FIELD_LABELS[f"Finding{_slot}DefectType"] = f"Finding {_slot} Defect Type"
    FIELD_LABELS[f"Finding{_slot}Severity"] = f"Finding {_slot} Severity"
    FIELD_LABELS[f"Finding{_slot}Confidence"] = f"Finding {_slot} Confidence"
    FIELD_LABELS[f"Notes-Finding{_slot}"] = f"Notes-Finding {_slot}"

FIELD_HELP = {
    "BenchmarkTrustCheck": "Quick expert judgment about whether this looks like a usable evaluation item.",
    "Notes-BenchmarkTrustCheck": "Point to the specific mismatch, weakness, or reassurance that drove your judgment.",
    "FailureAttribution": "Only answer this if you think there is a meaningful issue.",
    "Notes-FailureAttribution": "Name the concrete reason you think the issue is mostly in the dataset, mostly in the model, or unclear.",
    "PipelineCalibrationCheck": "Compare your judgment to the pipeline evidence shown on the page.",
    "Notes-PipelineCalibrationCheck": "Say what the pipeline got right, missed, or overstated.",
    "FinalAction": "Choose the action you would recommend for benchmark use.",
    "Notes-FinalAction": "Short rationale for the action.",
    "SummaryConfidence": "Use low confidence if you would want more evidence before being comfortable.",
}
for _slot in FINDING_SLOTS:
    FIELD_HELP[f"Finding{_slot}DefectType"] = "Optional structured finding. Add one when you see a specific benchmark issue worth recording."
    FIELD_HELP[f"Finding{_slot}Severity"] = "How serious is this specific finding for benchmark quality?"
    FIELD_HELP[f"Finding{_slot}Confidence"] = "How confident are you in this specific finding?"
    FIELD_HELP[f"Notes-Finding{_slot}"] = "Short note for this finding."

FIELD_OPTION_HELP = {
    "BenchmarkTrustCheck": {
        "trustworthy": "Looks coherent enough to treat as a usable benchmark item.",
        "defective": "Likely contains a benchmark flaw, mismatch, or invalid evaluation contract.",
        "ambiguous": "Not clearly trustworthy or clearly defective from a quick review.",
    },
    "FindingDefectType": {
        "none": "No clear defect stands out.",
        "prompt_ambiguity": "The prompt leaves important requirements unclear or inconsistent.",
        "ideal_response_issue": "The ideal response looks wrong or mismatched to the prompt.",
        "test_issue": "The tests seem wrong, incomplete, brittle, or misaligned.",
        "alignment_issue": "Prompt, ideal response, and tests do not seem to evaluate the same thing.",
        "difficulty_misalignment": "The sample seems mis-targeted in difficulty for the intended benchmark use.",
        "redundancy": "This sample looks too similar to others to add much benchmark information.",
        "other": "A different defect type is present.",
    },
    "FailureAttribution": {
        "model_fail": "The sample seems valid; the issue looks more like a true model weakness.",
        "dataset_fail": "The issue looks more like a benchmark or labeling defect.",
        "ambiguous": "The source of the issue is unclear from quick review.",
    },
    "PipelineCalibrationCheck": {
        "agree": "The pipeline's diagnosis mostly matches your judgment.",
        "overstates_problem": "The pipeline is making the sample look worse than your review suggests.",
        "misses_problem": "The pipeline is missing or underweighting a real problem you see.",
        "partially_agree": "The pipeline is partly right but is missing something important.",
    },
    "FinalAction": {
        "keep": "Retain this sample as-is.",
        "fix": "Keep only after fixing the identified issue.",
        "remove": "Remove or exclude from benchmark use.",
    },
    "FindingSeverity": {"low": "Small issue.", "medium": "Meaningful concern.", "high": "Serious benchmark-quality problem."},
    "Confidence": {"low": "Substantial uncertainty.", "medium": "Reasonable confidence.", "high": "Clear case with little uncertainty."},
}

SELECT_OPTIONS = {
    "BenchmarkTrustCheck": ["", "trustworthy", "defective", "ambiguous"],
    "FailureAttribution": ["", "model_fail", "dataset_fail", "ambiguous"],
    "PipelineCalibrationCheck": ["", "agree", "overstates_problem", "misses_problem", "partially_agree"],
    "FinalAction": ["", "keep", "fix", "remove"],
    "SummaryConfidence": ["", "low", "medium", "high"],
}
for _slot in FINDING_SLOTS:
    SELECT_OPTIONS[f"Finding{_slot}DefectType"] = ["", "none", "prompt_ambiguity", "ideal_response_issue", "test_issue", "alignment_issue", "difficulty_misalignment", "redundancy", "other"]
    SELECT_OPTIONS[f"Finding{_slot}Severity"] = ["", "low", "medium", "high"]
    SELECT_OPTIONS[f"Finding{_slot}Confidence"] = ["", "low", "medium", "high"]


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Launch a lightweight local web UI for Stage 4 manual review.")
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8765)
    parser.add_argument("--open-browser", action="store_true", default=True)
    return parser


class ReviewStore:
    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.packet_path = output_dir / "review_packet.json"
        self.review_csv_path = output_dir / "human_review.csv"
        self.legacy_csv_path = output_dir / "review_progress.csv"
        payload = json.loads(self.packet_path.read_text(encoding="utf-8")) if self.packet_path.exists() else {"reviews": []}
        self.rows: list[dict[str, Any]] = payload.get("reviews", [])
        self.rows_by_index = {int(row["Index"]): row for row in self.rows if str(row.get("Index", "")).strip()}
        self.jsonl_path = ARTIFACTS_DIR / "provided" / "Samples.jsonl"
        self.stage1_rows = _load_stage1(STAGE1_WORKBOOK)
        self.stage1_detailed_rows = _load_stage1_detailed(STAGE1_WORKBOOK)
        self.stage2_rows = _load_stage2(STAGE2_DIR)
        self.stage2_test_evidence = _load_stage2_test_evidence(STAGE2_DIR, self.stage2_rows)
        self.stage3_context = _load_stage3(STAGE3_DIR)
        self.stage3_detailed_rows = {int(row["Index"]): row for row in self.stage3_context.get("detailed", [])}
        self.progress = self._load_progress()
        self.lock = threading.Lock()

    def _load_progress(self) -> dict[int, dict[str, str]]:
        for path in (self.review_csv_path, self.legacy_csv_path):
            if path.exists():
                with path.open("r", encoding="utf-8-sig", newline="") as handle:
                    return {int(row["Index"]): _extract_review_fields(row) for row in csv.DictReader(handle) if str(row.get("Index", "")).strip()}
        return {}

    def _build_context_row(self, index: int, sample) -> dict[str, Any]:
        stage3_row = self.stage3_detailed_rows.get(index, {})
        stage2_row = self.stage2_rows.get(index, {})
        stage1_row = self.stage1_rows.get(index, {})
        stage1_detailed_row = self.stage1_detailed_rows.get(index, {})
        stage2_evidence = self.stage2_test_evidence.get(index, {})
        return {
            "Index": index,
            "QuestionId": sample.row.get("question_id", ""),
            "QuestionTitle": sample.row.get("question_title", ""),
            "SelectionBucket": "manual_lookup",
            "SelectionReason": "manual sample lookup",
            "ReviewContext": _review_context(stage3_row, stage2_row),
            "ObservedTestEvidence": stage2_evidence.get("summary", ""),
            "ObservedFailedTests": stage2_evidence.get("failed_tests", []),
            "Prompt": sample.question_content,
            "StarterCode": sample.starter_code,
            "IdealResponse": sample.ideal_response,
            "Difficulty": sample.row.get("difficulty", ""),
            "Stage3AuditPriority": stage3_row.get("AuditPriority", ""),
            "PipelineUtilityLabel": _pipeline_utility_label(stage3_row) if stage3_row else "usable",
            "Stage1Prompt": stage1_row.get("Prompt", ""),
            "Stage1IdealResponse": stage1_row.get("Ideal_Response", ""),
            "Stage1TestCases": stage1_row.get("Test_Cases", ""),
            "Stage1ImperfectChecks": _stage1_imperfect_checks(stage1_detailed_row),
            "Stage2EfficacyLabel": stage2_row.get("EfficacyLabel", ""),
            "Stage2BenchmarkQualitySignal": stage2_row.get("BenchmarkQualitySignal", ""),
            "Stage2FailureCategory": stage2_row.get("FailureCategory", ""),
            "WinnerCombinedPassRate": stage2_row.get("BestCombinedPassRate", ""),
            "OraclePassRate": stage2_row.get("OraclePassRate", ""),
            "Stage3ModelDisagreementSource": stage3_row.get("ModelDisagreementSource", ""),
            "PromptExcerpt": _excerpt(sample.question_content, 700),
        }

    def _enrich_row(self, index: int, row: dict[str, Any]) -> dict[str, Any]:
        samples = load_samples(self.jsonl_path, indices=str(index), limit=None, offset=0)
        if not samples:
            raise KeyError(index)
        enriched = self._build_context_row(index, samples[0])
        enriched.update(row)
        for field in REVIEW_FIELDS:
            enriched.setdefault(field, "")
        return enriched

    def ensure_index_loaded(self, index: int) -> None:
        if index in self.rows_by_index:
            self.rows_by_index[index] = self._enrich_row(index, self.rows_by_index[index])
            return
        samples = load_samples(self.jsonl_path, indices=str(index), limit=None, offset=0)
        if not samples:
            raise KeyError(index)
        sample = samples[0]
        row = self._build_context_row(index, sample)
        for field in REVIEW_FIELDS:
            row.setdefault(field, "")
        self.rows.append(row)
        self.rows_by_index[index] = row

    def merged_row(self, index: int) -> dict[str, Any]:
        self.ensure_index_loaded(index)
        base = dict(self.rows_by_index[index])
        base.setdefault("Stage1ImperfectChecks", _stage1_imperfect_checks(self.stage1_detailed_rows.get(index, {})))
        base.update(self.progress.get(index, {}))
        return base

    def reviewed_indices(self) -> list[int]:
        return sorted(
            index
            for index, row in self.progress.items()
            if any(str(row.get(field, "")).strip() for field in REVIEW_FIELDS)
        )

    def ordered_indices(self, list_name: str = "exemplars", current_index: int | None = None) -> list[int]:
        if list_name == "exemplars":
            indices = list(EXEMPLAR_LIST)
        else:
            bucket_order = {"contradiction": 0, "disagreement": 1, "redundancy": 2, "baseline": 3, "backfill": 4, "manual_lookup": 5}
            indices = [
                int(row["Index"])
                for row in sorted(
                    self.rows,
                    key=lambda row: (bucket_order.get(str(row.get("SelectionBucket", "")), 9), int(row["Index"])),
                )
            ]
        if current_index is not None and current_index not in indices:
            self.ensure_index_loaded(current_index)
            indices.append(current_index)
        deduped: list[int] = []
        for index in indices:
            try:
                self.ensure_index_loaded(index)
            except Exception:
                continue
            if index not in deduped:
                deduped.append(index)
        return deduped

    def completion_count(self, list_name: str = "exemplars") -> int:
        target = set(self.ordered_indices(list_name))
        return sum(1 for index in self.reviewed_indices() if index in target)

    def first_index(self, list_name: str = "exemplars") -> int:
        indices = self.ordered_indices(list_name)
        if not indices:
            raise KeyError(list_name)
        return indices[0]

    def first_incomplete_index(self, list_name: str = "exemplars") -> int:
        indices = self.ordered_indices(list_name)
        for index in indices:
            row = self.progress.get(index, {})
            if not any(str(row.get(field, "")).strip() for field in REVIEW_FIELDS):
                return index
        return indices[0]

    def has_saved_progress(self) -> bool:
        return bool(self.reviewed_indices())

    def save_review(self, index: int, payload: dict[str, str]) -> None:
        with self.lock:
            self.ensure_index_loaded(index)
            self.progress[index] = {field: payload.get(field, "").strip() for field in REVIEW_FIELDS}
            self._write_progress_files()

    def clear_progress(self) -> None:
        with self.lock:
            self.progress = {}
            for path in (self.review_csv_path, self.legacy_csv_path):
                if path.exists():
                    path.unlink()

    def _write_progress_files(self) -> None:
        reviewed_rows = [dict(self.merged_row(index)) for index in self.reviewed_indices()]
        if reviewed_rows:
            with self.review_csv_path.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=list(reviewed_rows[0].keys()))
                writer.writeheader()
                writer.writerows(reviewed_rows)
        else:
            self.review_csv_path.write_text("", encoding="utf-8")


def _load_stage1_detailed(workbook_path: Path) -> dict[int, dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if "Detailed" not in workbook.sheetnames:
            return {}
        worksheet = workbook["Detailed"]
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value) if value is not None else "" for value in next(rows, [])]
        result: dict[int, dict[str, Any]] = {}
        for values in rows:
            if not values or all(value is None for value in values):
                continue
            row = {headers[idx]: values[idx] for idx in range(min(len(headers), len(values))) if headers[idx]}
            try:
                index = int(row.get("Index"))
            except (TypeError, ValueError):
                continue
            result[index] = row
        return result
    finally:
        workbook.close()


def _stage1_imperfect_checks(row: dict[str, Any]) -> list[dict[str, str]]:
    checks: list[dict[str, str]] = []
    if not row:
        return checks
    for key, value in row.items():
        if not isinstance(key, str) or not key[:1].isdigit() or key.startswith("Notes-"):
            continue
        status = str(value or "").strip()
        if status not in {"FAIL", "PARTIAL", "UNCLEAR"}:
            continue
        note = str(row.get(f"Notes-{key}", "") or "").strip()
        checks.append({"name": key, "status": status, "note": note})
    return checks


def _extract_review_fields(row: dict[str, Any]) -> dict[str, str]:
    return {field: str(row.get(field, "") or "") for field in REVIEW_FIELDS}


def run_review_ui_cli(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    output_dir = _resolve_path(args.output_dir)
    store = ReviewStore(output_dir)

    class ReviewHandler(BaseHTTPRequestHandler):
        def do_GET(self) -> None:
            parsed = urllib.parse.urlparse(self.path)
            query = urllib.parse.parse_qs(parsed.query)
            list_name = normalize_list_name((query.get("list") or ["exemplars"])[0])
            readonly = LIST_METADATA[list_name]["readonly"] or (query.get("readonly") or ["0"])[0] == "1"
            if parsed.path == "/":
                self._html_response(render_home_page(store))
                return
            if parsed.path == "/resume":
                self._redirect(f"/review/{store.first_incomplete_index(list_name)}?list={list_name}")
                return
            if parsed.path.startswith("/list/"):
                list_name = normalize_list_name(parsed.path.rsplit("/", 1)[-1])
                try:
                    index = store.first_index(list_name)
                except KeyError:
                    self._redirect("/")
                    return
                suffix = "&readonly=1" if LIST_METADATA[list_name]["readonly"] else ""
                self._redirect(f"/review/{index}?list={list_name}{suffix}")
                return
            if parsed.path == "/jump":
                raw_index = (query.get("index") or [""])[0].strip()
                try:
                    target_index = int(raw_index)
                    store.ensure_index_loaded(target_index)
                    suffix = "&readonly=1" if readonly else ""
                    self._redirect(f"/review/{target_index}?list={list_name}{suffix}")
                except Exception:
                    self._redirect("/")
                return
            if parsed.path.startswith("/review/"):
                try:
                    index = int(parsed.path.rsplit("/", 1)[-1])
                except ValueError:
                    self.send_error(HTTPStatus.NOT_FOUND)
                    return
                try:
                    store.ensure_index_loaded(index)
                except Exception:
                    self.send_error(HTTPStatus.NOT_FOUND)
                    return
                self._html_response(render_review_page(store, index, list_name=list_name, readonly=readonly))
                return
            self.send_error(HTTPStatus.NOT_FOUND)

        def do_POST(self) -> None:
            parsed = urllib.parse.urlparse(self.path)
            query = urllib.parse.parse_qs(parsed.query)
            list_name = normalize_list_name((query.get("list") or ["exemplars"])[0])
            readonly = LIST_METADATA[list_name]["readonly"] or (query.get("readonly") or ["0"])[0] == "1"
            if not parsed.path.startswith("/save/"):
                self.send_error(HTTPStatus.NOT_FOUND)
                return
            if readonly:
                self._redirect(f"/review/{parsed.path.rsplit('/', 1)[-1]}?list={list_name}&readonly=1")
                return
            try:
                index = int(parsed.path.rsplit("/", 1)[-1])
            except ValueError:
                self.send_error(HTTPStatus.NOT_FOUND)
                return
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length).decode("utf-8")
            form = urllib.parse.parse_qs(body, keep_blank_values=True)
            payload = {field: form.get(field, [""])[0] for field in REVIEW_FIELDS}
            action = form.get("_action", ["save"])[0]
            store.save_review(index, payload)
            target = index
            if action == "prev":
                target = previous_index(store, index, list_name)
            elif action == "next":
                target = next_index(store, index, list_name)
            self._redirect(f"/review/{target}?list={list_name}")

        def log_message(self, _format: str, *_args) -> None:
            return

        def _html_response(self, content: str) -> None:
            body = content.encode("utf-8")
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def _json_response(self, payload: dict[str, Any]) -> None:
            body = (json.dumps(payload, ensure_ascii=False, indent=2) + "\n").encode("utf-8")
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def _redirect(self, location: str) -> None:
            self.send_response(HTTPStatus.SEE_OTHER)
            self.send_header("Location", location)
            self.end_headers()

    server = ThreadingHTTPServer((args.host, args.port), ReviewHandler)
    url = f"http://{args.host}:{args.port}/"
    print(f"Stage 4 review UI running at {url}")
    print(f"Human review autosaves to {store.review_csv_path}")
    print("When you are done, finalize Stage 4 with: " f"python main.py --stage manual-audit --review-input {store.review_csv_path}")
    if args.open_browser:
        try:
            webbrowser.open(url)
        except Exception:
            pass
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nReview UI stopped.")
    finally:
        server.server_close()
    return 0


def normalize_list_name(value: str) -> str:
    return value if value in LIST_METADATA else "exemplars"


def previous_index(store: ReviewStore, current: int, list_name: str) -> int:
    indices = store.ordered_indices(list_name, current_index=current)
    position = indices.index(current)
    return indices[max(0, position - 1)]


def next_index(store: ReviewStore, current: int, list_name: str) -> int:
    indices = store.ordered_indices(list_name, current_index=current)
    position = indices.index(current)
    return indices[min(len(indices) - 1, position + 1)]


def render_review_page(store: ReviewStore, index: int, *, list_name: str, readonly: bool) -> str:
    row = store.merged_row(index)
    indices = store.ordered_indices(list_name, current_index=index)
    position = indices.index(index)
    completed = store.completion_count(list_name)
    progress_percent = completed / max(1, len(indices)) * 100.0
    save_controls = (
        '<button type="submit" form="review-form" name="_action" value="prev" class="alt">Save + Previous</button>'
        '<button type="submit" form="review-form" name="_action" value="save" class="alt">Save</button>'
        '<button type="submit" form="review-form" name="_action" value="next">Save + Next</button>'
        if not readonly
        else '<span class="small"><strong>Read-only mode.</strong> Editing and saving are disabled.</span>'
    )
    form_open = f'<form id="review-form" method="post" action="/save/{index}?list={list_name}">' if not readonly else '<div id="review-form">'
    form_close = "</form>" if not readonly else "</div>"
    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><title>Stage 4 Review - {index}</title>
<style>
body{{margin:0;font-family:Georgia,'Times New Roman',serif;background:linear-gradient(180deg,#f8f4ec 0%,#f2efe8 100%);color:#1f2937;}}
.shell{{max-width:1040px;margin:0 auto;padding:104px 24px 72px;}}
.card{{background:rgba(255,255,255,0.88);border:1px solid #dccfb9;border-radius:18px;padding:18px 20px;box-shadow:0 8px 30px rgba(80,58,28,0.08);margin-bottom:18px;}}
.actionbar{{position:fixed;top:0;left:0;right:0;z-index:10;background:rgba(248,244,236,0.96);border-bottom:1px solid #dccfb9;box-shadow:0 8px 24px rgba(80,58,28,0.08);}}
.actionbar-inner{{max-width:1040px;margin:0 auto;padding:14px 24px;display:flex;justify-content:space-between;align-items:center;gap:12px;flex-wrap:wrap;}}
.actions{{display:flex;gap:10px;flex-wrap:wrap;align-items:center;}}
.small{{font-size:13px;color:#5b6470;line-height:1.5;}}
.meta{{display:flex;flex-wrap:wrap;gap:10px;margin-top:10px;}}
.pill{{padding:6px 10px;border-radius:999px;background:#efe5d1;font-size:13px;}}
.progress{{width:100%;height:10px;background:#e8dcc4;border-radius:999px;overflow:hidden;margin:10px 0 6px;}} .progress>div{{height:100%;width:{progress_percent:.2f}%;background:linear-gradient(90deg,#9b6b2f 0%,#d4a15d 100%);}}
details{{margin-bottom:14px;border:1px solid #e0d3bf;border-radius:14px;background:#fffdf8;overflow:hidden;}} summary{{cursor:pointer;padding:14px 16px;font-weight:600;background:#f7f0e1;}}
pre{{white-space:pre-wrap;word-break:break-word;margin:0;padding:16px;font-family:Consolas,'Courier New',monospace;font-size:13px;line-height:1.5;max-height:520px;overflow:auto;}}
label{{display:block;font-weight:600;margin-bottom:6px;}} .field{{margin-bottom:14px;padding:12px 12px 10px;border-radius:12px;border:1px solid #eadfce;background:#fff8ef;}} .field.readonly-filled{{background:#fff8ef;border-color:#d7b98a;}} .field.readonly-empty{{background:#f8f5ef;border-color:#e7dece;}} .label-row{{display:flex;align-items:center;gap:8px;margin-bottom:6px;}}
select,textarea,input[type=number]{{width:100%;border:1px solid #cbb89a;border-radius:10px;padding:10px 12px;box-sizing:border-box;font:inherit;background:white;}} textarea{{min-height:76px;resize:vertical;}}
.section-title{{margin:18px 0 10px;padding-top:14px;border-top:1px solid #e3d7c3;font-size:14px;letter-spacing:.02em;text-transform:uppercase;color:#6b7280;}}
.finding-block{{margin-top:14px;padding-top:14px;border-top:1px dashed #d8c9b2;}} .finding-grid{{display:grid;grid-template-columns:1.4fr .8fr .8fr;gap:10px;align-items:start;}}
.tooltip{{position:relative;display:inline-flex;align-items:center;justify-content:center;width:18px;height:18px;border-radius:999px;background:#e8dcc4;color:#244c5a;font-size:12px;cursor:help;}}
.tooltip .tooltip-text{{visibility:hidden;opacity:0;transition:opacity .12s ease;position:absolute;left:24px;top:-6px;width:320px;z-index:3;background:#213540;color:#f8fafc;padding:10px 12px;border-radius:10px;box-shadow:0 10px 24px rgba(0,0,0,.22);line-height:1.45;text-align:left;}} .tooltip:hover .tooltip-text{{visibility:visible;opacity:1;}}
.evidence-list{{margin:8px 0 0;padding-left:18px;font-size:13px;line-height:1.45;color:#374151;}} a{{color:#244c5a;text-decoration:none;}}
button,.button-link{{border:0;border-radius:999px;padding:10px 16px;font:inherit;cursor:pointer;background:#244c5a;color:white;text-decoration:none;display:inline-block;}} button.alt,.button-link.alt{{background:#7b5b36;}}
.readonly-note{{padding:12px 14px;border-radius:12px;background:#f6ead7;color:#6b4c22;font-size:13px;margin-bottom:14px;}} .auto-section{{margin-top:10px;}} .auto-divider{{border:0;border-top:1px solid #e3d7c3;margin:14px 0;}}
</style></head><body>
<div class="actionbar"><div class="actionbar-inner"><div class="small">{escape(LIST_METADATA[list_name]["label"])} · Review {position + 1} of {len(indices)} · Dataset index {index}</div>
<div class="actions"><a class="button-link alt" href="/">Home</a><a class="button-link alt" href="/list/exemplars">Exemplar Set</a>
<form method="get" action="/jump" style="display:flex;gap:8px;align-items:center;"><input type="hidden" name="list" value="{list_name}"><input type="number" name="index" value="{index}" style="width:96px;border-radius:999px;"><button type="submit" class="alt">Go to sample</button></form>{save_controls}</div></div></div>
<div class="shell"><div class="card"><h1>{escape(LIST_METADATA[list_name]["label"])}</h1><div class="small">{escape(LIST_METADATA[list_name]["description"])}</div>
<div class="meta"><span class="pill">Question ID: {escape(row.get("QuestionId", ""))}</span><span class="pill">Dataset index: {index}</span><span class="pill">Bucket: {escape(row.get("SelectionBucket", ""))}</span><span class="pill">Reason: {escape(row.get("SelectionReason", ""))}</span></div>
<div class="progress"><div></div></div><div class="small">{completed} completed in this workflow · {len(indices)-completed} remaining</div></div>
<div class="card"><h3>Review guidance</h3><div class="small">Focus on adjudication, not re-running the benchmark by hand:<br>1. Is this sample trustworthy as an evaluation item?<br>2. If something looks wrong, is it more likely a dataset defect or a true model failure?<br>3. Did the automated pipeline diagnose the sample correctly?<br><br>Use the sample summary for the overall judgment. Add structured findings only when there is a concrete issue worth recording.<br><br>Human review autosaves to <code>{escape(str(store.review_csv_path))}</code>.</div></div>
{form_open}<div class="card"><h3>Most Recent Auto-Evaluation</h3>{render_auto_evaluation(row)}</div>
<details open><summary>Prompt</summary><pre>{escape(row.get("Prompt", ""))}</pre></details>
<details><summary>Starter Code</summary><pre>{escape(row.get("StarterCode", ""))}</pre></details>
<details><summary>Ideal Response</summary><pre>{escape(row.get("IdealResponse", ""))}</pre></details>
<div class="card">{render_form_fields(row, readonly=readonly)}</div>{form_close}</div></body></html>"""


def render_home_page(store: ReviewStore) -> str:
    completed = store.completion_count("exemplars")
    total = len(store.ordered_indices("exemplars"))
    reviewed = store.reviewed_indices()
    meta = LIST_METADATA["exemplars"]
    cards = [
        f'<div class="list-card"><h3>{escape(meta["label"])}</h3><div class="small">{escape(meta["description"])}</div><div class="small" style="margin-top:8px;">{len(store.ordered_indices("exemplars"))} samples</div><div style="margin-top:14px;"><a class="button-link" href="/list/exemplars">Open</a></div></div>'
    ]
    return f"""<!doctype html><html lang="en"><head><meta charset="utf-8"><title>Stage 4 Review</title>
<style>body{{margin:0;font-family:Georgia,'Times New Roman',serif;background:linear-gradient(180deg,#f8f4ec 0%,#f2efe8 100%);color:#1f2937;}} .shell{{max-width:980px;margin:56px auto;padding:24px;}} .card{{background:rgba(255,255,255,.9);border:1px solid #dccfb9;border-radius:18px;padding:22px 24px;box-shadow:0 8px 30px rgba(80,58,28,.08);margin-bottom:18px;}} .grid{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:16px;}} .list-card{{background:#fffdf8;border:1px solid #e0d3bf;border-radius:16px;padding:18px;}} .button-link,button{{border:0;border-radius:999px;padding:11px 18px;font:inherit;cursor:pointer;background:#244c5a;color:white;text-decoration:none;}} button.alt{{background:#7b5b36;}} .actions{{display:flex;gap:12px;margin-top:20px;flex-wrap:wrap;align-items:center;}} .small{{font-size:14px;color:#5b6470;line-height:1.5;}}</style></head>
<body><div class="shell"><div class="card"><h1>Stage 4 Human Review</h1><p class="small">Saved human-review rows: {len(reviewed)}. Curated exemplar set: {completed} of {total} completed.</p><p class="small">Open the curated exemplar set below, or jump directly to any dataset index. Ad hoc sample review is allowed and will be saved to the canonical human-review CSV.</p>
<form method="get" action="/jump" style="display:flex;gap:12px;flex-wrap:wrap;align-items:center;margin-top:16px;"><input type="hidden" name="list" value="exemplars"><input type="number" name="index" placeholder="Dataset index" style="width:140px;border:1px solid #cbb89a;border-radius:999px;padding:11px 18px;font:inherit;background:white;"><button type="submit" class="alt">Open sample by index</button></form></div>
<div class="grid">{''.join(cards)}</div></div></body></html>"""


def render_form_fields(row: dict[str, Any], *, readonly: bool) -> str:
    chunks: list[str] = []
    if readonly:
        chunks.append('<div class="readonly-note">Read-only mode is active. You can inspect existing human eval here, but editing and saving are disabled.</div>')
    chunks.append('<div class="section-title">Sample Summary</div>')
    for field in SUMMARY_FIELDS:
        chunks.append(render_field_block(field, row, readonly))
    chunks.append('<div class="section-title">Optional Structured Findings</div>')
    for slot in FINDING_SLOTS:
        chunks.append(f'<div class="finding-block"><div class="small"><strong>Finding {slot}</strong>: add only if there is a concrete issue worth recording.</div><div class="finding-grid">')
        for field in (f"Finding{slot}DefectType", f"Finding{slot}Severity", f"Finding{slot}Confidence"):
            chunks.append(render_field_block(field, row, readonly))
        chunks.append("</div>")
        chunks.append(render_field_block(f"Notes-Finding{slot}", row, readonly))
        chunks.append("</div>")
    return "".join(chunks)


def render_field_block(field: str, row: dict[str, Any], readonly: bool) -> str:
    label = FIELD_LABELS.get(field, field)
    control = render_control(field, str(row.get(field, "")), readonly=readonly)
    glossary = render_option_glossary(field)
    tooltip = render_tooltip(field)
    value = str(row.get(field, "") or "").strip()
    classes = ["field"]
    if readonly:
        classes.append("readonly-filled" if value else "readonly-empty")
    return (
        f'<div class="{" ".join(classes)}"><div class="label-row"><label>{escape(label)}</label>{tooltip}</div>'
        f'<div class="small" style="margin-bottom:6px;">{escape(FIELD_HELP.get(field, ""))}</div>{control}{glossary}</div>'
    )


def render_control(field: str, value: str, *, readonly: bool) -> str:
    disabled = " disabled" if readonly else ""
    if field in SELECT_OPTIONS:
        options = "".join(
            f'<option value="{escape(option)}"{" selected" if option == value else ""}>{escape(option or "—")}</option>'
            for option in SELECT_OPTIONS[field]
        )
        return f'<select name="{field}"{disabled}>{options}</select>'
    return f'<textarea name="{field}"{disabled}>{escape(value)}</textarea>'


def render_failed_tests(row: dict[str, Any]) -> str:
    failed_tests = row.get("ObservedFailedTests", []) or []
    if not failed_tests:
        return ""
    return '<ul class="evidence-list">' + "".join(f"<li>{escape(item)}</li>" for item in failed_tests) + "</ul>"


def render_auto_evaluation(row: dict[str, Any]) -> str:
    summary_line = (
        f"Stage 1: Prompt {escape(row.get('Stage1Prompt', 'N/A'))} · Ideal {escape(row.get('Stage1IdealResponse', 'N/A'))} · Tests {escape(row.get('Stage1TestCases', 'N/A'))} | "
        f"Stage 2: {escape(row.get('Stage2EfficacyLabel', 'N/A'))} · {escape(row.get('Stage2FailureCategory', 'N/A'))} · best pass {escape(row.get('WinnerCombinedPassRate', 'N/A'))}"
    )
    stage1_html = (
        f'<div class="small auto-section"><strong>Stage 1 summary</strong><br>'
        f'Prompt {escape(row.get("Stage1Prompt", "N/A"))} · '
        f'Ideal {escape(row.get("Stage1IdealResponse", "N/A"))} · '
        f'Tests {escape(row.get("Stage1TestCases", "N/A"))}</div>'
    )
    checks = row.get("Stage1ImperfectChecks", []) or []
    if checks:
        items = []
        for item in checks:
            note = str(item.get("note", "") or "").strip()
            note_html = f'<div class="small" style="margin-top:4px;">{escape(note)}</div>' if note else ""
            items.append(f'<li><strong>{escape(item.get("name", ""))}</strong> · {escape(item.get("status", ""))}{note_html}</li>')
        stage1_html += '<div class="small" style="margin-top:10px;"><strong>Stage 1 imperfect checks</strong></div><ul class="evidence-list">' + "".join(items) + "</ul>"
    observed = str(row.get("ObservedTestEvidence", "") or "").strip()
    stage2_summary = (
        '<hr class="auto-divider">'
        f'<div class="small auto-section"><strong>Stage 2 summary</strong><br>{escape(row.get("Stage2EfficacyLabel", "N/A"))} · '
        f'{escape(row.get("Stage2FailureCategory", "N/A"))} · best pass {escape(row.get("WinnerCombinedPassRate", "N/A"))}</div>'
    )
    if observed:
        stage2_summary += f'<div class="small" style="margin-top:10px;"><strong>Stage 2 failed-test summary</strong><br>{escape(observed)}</div>'
    failed = render_failed_tests(row)
    if failed:
        stage2_summary += '<div class="small" style="margin-top:10px;"><strong>Stage 2 imperfect tests</strong></div>' + failed
    context_html = f'<div class="small" style="margin-top:10px;"><strong>Pipeline synthesis</strong><br>{escape(row.get("ReviewContext", ""))}</div>'
    return f'<details><summary>{summary_line}</summary>{stage1_html}{stage2_summary}{context_html}</details>'


def render_tooltip(field: str) -> str:
    option_help = option_help_for_field(field)
    if not option_help:
        return ""
    text = "<br>".join(f"<strong>{escape(option)}</strong>: {escape(description)}" for option, description in option_help.items())
    return f'<span class="tooltip">?<span class="tooltip-text">{text}</span></span>'


def render_option_glossary(field: str) -> str:
    option_help = option_help_for_field(field)
    if not option_help:
        return ""
    items = "".join(f"<li><strong>{escape(option)}</strong>: {escape(description)}</li>" for option, description in option_help.items())
    return f'<ul class="evidence-list">{items}</ul>'


def option_help_for_field(field: str) -> dict[str, str]:
    if field.startswith("Finding") and field.endswith("DefectType"):
        return FIELD_OPTION_HELP["FindingDefectType"]
    if field.startswith("Finding") and field.endswith("Severity"):
        return FIELD_OPTION_HELP["FindingSeverity"]
    if field == "SummaryConfidence" or (field.startswith("Finding") and field.endswith("Confidence")):
        return FIELD_OPTION_HELP["Confidence"]
    return FIELD_OPTION_HELP.get(field, {})


def escape(value: Any) -> str:
    return html.escape(str(value or ""))
