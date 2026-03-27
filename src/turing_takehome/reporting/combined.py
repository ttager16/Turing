from __future__ import annotations

import csv
import json
import re
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .notes import (
    generate_notes_for_requests,
    split_cached_note_requests,
    update_note_cache,
)


PROJECT_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_REPORT_DIR = PROJECT_ROOT / "outputs" / "reports"
STABLE_COMBINED_XLSX = PROJECT_ROOT / "outputs" / "dataset_analysis.xlsx"
STABLE_COMBINED_JSON = PROJECT_ROOT / "outputs" / "dataset_analysis.json"
_ILLEGAL_XLSX_CHARS = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


def export_combined_report(
    *,
    stage1_workbook_path: Path,
    stage2_output_root: Path,
    stage3_output_root: Path | None = None,
    stage4_output_root: Path | None = None,
    output_dir: Path | None = None,
    timestamp: str | None = None,
) -> tuple[Path, Path]:
    report_dir = output_dir or DEFAULT_REPORT_DIR
    report_dir.mkdir(parents=True, exist_ok=True)
    stamp = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")

    stage1_detailed_rows, stage1_summary_rows = _load_stage1_rows(stage1_workbook_path)
    stage2_context = _load_stage2_context(stage2_output_root)
    stage2_detailed_rows, test_columns = _build_stage2_detailed_rows(stage2_context)
    stage2_aggregate_rows = _build_stage2_aggregate_rows(stage2_context)
    stage3_context = _load_stage3_context(stage3_output_root) if stage3_output_root else None
    stage3_detailed_rows, stage3_summary_rows, stage3_test_columns = _build_stage3_rows(
        stage3_context,
        stage2_context,
    )
    stage4_context = _load_stage4_context(stage4_output_root) if stage4_output_root else None
    stage4_detailed_rows, stage4_summary_rows, stage4_test_columns = _build_stage4_rows(stage4_context)

    workbook_path = report_dir / f"combined_analysis_{stamp}.xlsx"
    json_path = report_dir / f"combined_analysis_{stamp}.json"

    _write_combined_workbook(
        workbook_path,
        stage1_detailed_rows=stage1_detailed_rows,
        stage1_summary_rows=stage1_summary_rows,
        stage2_detailed_rows=stage2_detailed_rows,
        stage2_aggregate_rows=stage2_aggregate_rows,
        test_columns=test_columns,
        stage3_detailed_rows=stage3_detailed_rows,
        stage3_summary_rows=stage3_summary_rows,
        stage3_test_columns=stage3_test_columns,
        stage4_detailed_rows=stage4_detailed_rows,
        stage4_summary_rows=stage4_summary_rows,
        stage4_test_columns=stage4_test_columns,
    )
    _write_combined_json(
        json_path,
        timestamp=stamp,
        stage1_detailed_rows=stage1_detailed_rows,
        stage1_summary_rows=stage1_summary_rows,
        stage2_detailed_rows=stage2_detailed_rows,
        stage2_aggregate_rows=stage2_aggregate_rows,
        stage2_context=stage2_context,
        test_columns=test_columns,
        stage3_context=stage3_context,
        stage3_detailed_rows=stage3_detailed_rows,
        stage3_summary_rows=stage3_summary_rows,
        stage3_test_columns=stage3_test_columns,
        stage4_context=stage4_context,
        stage4_detailed_rows=stage4_detailed_rows,
        stage4_summary_rows=stage4_summary_rows,
        stage4_test_columns=stage4_test_columns,
    )
    _write_stable_combined_aliases(workbook_path, json_path)
    _clean_old_timestamped_reports(report_dir, keep_stamp=stamp)
    return workbook_path, json_path


def _write_stable_combined_aliases(workbook_path: Path, json_path: Path) -> None:
    STABLE_COMBINED_XLSX.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(workbook_path, STABLE_COMBINED_XLSX)
    shutil.copy2(json_path, STABLE_COMBINED_JSON)


def _clean_old_timestamped_reports(report_dir: Path, *, keep_stamp: str) -> None:
    keep_names = {
        f"combined_analysis_{keep_stamp}.xlsx",
        f"combined_analysis_{keep_stamp}.json",
    }
    for path in report_dir.glob("combined_analysis_*.*"):
        if path.name not in keep_names and path.suffix.lower() in {".xlsx", ".json"}:
            path.unlink(missing_ok=True)


def _load_stage1_rows(workbook_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    detailed_rows = _read_sheet_rows(wb, "Detailed")
    summary_rows = _read_sheet_rows(wb, "Summary")
    return detailed_rows, summary_rows


def _load_stage2_context(stage2_output_root: Path) -> dict[str, Any]:
    sample_rows = _read_csv_rows(stage2_output_root / "sample_results.csv")
    model_rows = _read_csv_rows(stage2_output_root / "sample_model_results.csv")
    attempt_rows = _read_jsonl_rows(stage2_output_root / "model_attempts.jsonl")
    per_test_rows = _read_jsonl_rows(stage2_output_root / "per_test_results.jsonl")
    manifest = json.loads((stage2_output_root / "run_manifest.json").read_text(encoding="utf-8"))
    return {
        "output_root": stage2_output_root,
        "sample_rows": sample_rows,
        "model_rows": model_rows,
        "attempt_rows": attempt_rows,
        "per_test_rows": per_test_rows,
        "manifest": manifest,
    }


def _load_stage3_context(stage3_output_root: Path) -> dict[str, Any] | None:
    if not stage3_output_root.exists():
        return None
    payload_path = stage3_output_root / "dataset_analysis.json"
    if not payload_path.exists():
        return None
    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    return payload.get("stage3")


def _load_stage4_context(stage4_output_root: Path) -> dict[str, Any] | None:
    if not stage4_output_root.exists():
        return None
    payload_path = stage4_output_root / "manual_audit.json"
    if not payload_path.exists():
        return None
    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    return payload.get("stage4")


def _build_stage3_rows(
    stage3_context: dict[str, Any] | None,
    stage2_context: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    if not stage3_context:
        return [], [], []
    detailed_rows = stage3_context.get("detailed", [])
    summary_rows = stage3_context.get("summary", [])
    test_columns = stage3_context.get("test_columns", [])
    return detailed_rows, summary_rows, test_columns


def _build_stage4_rows(
    stage4_context: dict[str, Any] | None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    if not stage4_context:
        return [], [], []
    detailed_rows = stage4_context.get("detailed", [])
    summary_rows = stage4_context.get("summary", [])
    test_columns = stage4_context.get("test_columns", [])
    return detailed_rows, summary_rows, test_columns


def _best_model_rows_by_sample(model_rows: list[dict[str, Any]]) -> dict[int, dict[str, Any]]:
    grouped: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in model_rows:
        grouped[int(row["Index"])].append(row)
    result: dict[int, dict[str, Any]] = {}
    for sample_index, rows in grouped.items():
        result[sample_index] = sorted(
            rows,
            key=lambda row: (
                -(_to_float(row.get("BestCombinedPassRate")) or 0.0),
                -(_to_float(row.get("BestProvidedPassRate")) or 0.0),
                str(row.get("TargetName", "")),
            ),
        )[0]
    return result


def _build_stage2_detailed_rows(context: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str]]:
    best_attempt_rows = {
        (
            int(row["Index"]),
            str(row["TargetName"]),
            int(row["BestAttemptIndex"]),
        ): _normalize_attempt_row(
            _find_attempt_row(
                context["attempt_rows"],
                index=int(row["Index"]),
                target_name=str(row["TargetName"]),
                attempt_index=int(row["BestAttemptIndex"]),
            )
        )
        for row in context["model_rows"]
    }

    per_test_lookup = defaultdict(dict)
    per_test_detail_lookup = defaultdict(dict)
    all_test_columns: set[str] = set()
    for row in context["per_test_rows"]:
        if row.get("source") != "model_candidate":
            continue
        key = (
            int(row["sample_index"]),
            str(row["target_name"]),
            int(row["attempt_index"]),
        )
        column_name = _test_column_name(str(row.get("visibility", "")), int(row.get("case_index", 0)))
        per_test_lookup[key][column_name] = _test_cell_value(row)
        per_test_detail_lookup[key][column_name] = row
        all_test_columns.add(column_name)

    ordered_test_columns = sorted(all_test_columns, key=_test_column_sort_key)
    note_cache_path = Path(context["output_root"]) / "detailed_test_notes.json"
    note_requests: list[dict[str, Any]] = []
    ordered_model_rows = sorted(
        context["model_rows"],
        key=lambda row: (int(row["Index"]), str(row["TargetName"])),
    )
    for model_row in ordered_model_rows:
        key = (
            int(model_row["Index"]),
            str(model_row["TargetName"]),
            int(model_row["BestAttemptIndex"]),
        )
        request = _build_stage2_note_request(
            model_row=model_row,
            attempt_row=best_attempt_rows[key],
            test_values=per_test_lookup.get(key, {}),
            test_details=per_test_detail_lookup.get(key, {}),
        )
        if request:
            note_requests.append(request)
    notes_by_request, missing_note_requests = split_cached_note_requests(
        "sample-efficacy-analysis",
        note_requests,
        note_cache_path,
    )
    if missing_note_requests:
        fresh_notes = generate_notes_for_requests(
            "sample-efficacy-analysis",
            missing_note_requests,
            max_concurrency=8,
        )
        notes_by_request.update(fresh_notes)
        update_note_cache(
            "sample-efficacy-analysis",
            missing_note_requests,
            fresh_notes,
            note_cache_path,
        )

    detailed_rows: list[dict[str, Any]] = []
    for model_row in ordered_model_rows:
        key = (
            int(model_row["Index"]),
            str(model_row["TargetName"]),
            int(model_row["BestAttemptIndex"]),
        )
        attempt_row = best_attempt_rows[key]
        test_values = per_test_lookup.get(key, {})
        request_id = _stage2_note_request_id(model_row)
        row_notes = notes_by_request.get(request_id, {})
        detailed_row = {
            "Index": int(model_row["Index"]),
            "QuestionId": model_row["QuestionId"],
            "QuestionTitle": model_row["QuestionTitle"],
            "Difficulty": model_row["Difficulty"],
            "TargetName": model_row["TargetName"],
            "ModelLabel": model_row["ModelLabel"],
            "Attempts": int(model_row["Attempts"]),
            "BestAttemptIndex": int(model_row["BestAttemptIndex"]),
            "OraclePassRate": _to_float(model_row.get("OraclePassRate")),
            "PublicPassRate": _to_float(attempt_row.get("PublicPassRate")),
            "PrivatePassRate": _to_float(attempt_row.get("PrivatePassRate")),
            "GeneratedPassRate": _to_float(model_row.get("BestGeneratedPassRate")),
            "ProvidedPassRate": _to_float(model_row.get("BestProvidedPassRate")),
            "CombinedPassRate": _to_float(model_row.get("BestCombinedPassRate")),
            "CombinedPass@1": _to_float(model_row.get("CombinedPass@1")),
            "CombinedPass@2": _to_float(model_row.get("CombinedPass@2")),
            "DifficultyEstimate": model_row.get("DifficultyEstimate", ""),
            "FailureCategory": model_row.get("FailureCategory", ""),
            "BenchmarkQualitySignal": model_row.get("BenchmarkQualitySignal", ""),
            "EfficacyLabel": model_row.get("EfficacyLabel", ""),
            "Suspicious": _to_bool(model_row.get("Suspicious")),
            "NeedsAudit": _to_bool(model_row.get("NeedsAudit")),
        }
        for column_name in ordered_test_columns:
            detailed_row[column_name] = test_values.get(column_name, "")
            detailed_row[f"Notes-{column_name}"] = row_notes.get(column_name, "")
        detailed_rows.append(detailed_row)
    return detailed_rows, ordered_test_columns


def _build_stage2_aggregate_rows(context: dict[str, Any]) -> list[dict[str, Any]]:
    model_rows_by_sample: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in context["model_rows"]:
        model_rows_by_sample[int(row["Index"])].append(row)

    attempt_lookup = {
        (
            int(row["Index"]),
            str(row["TargetName"]),
            int(row["AttemptIndex"]),
        ): _normalize_attempt_row(row)
        for row in context["attempt_rows"]
    }

    aggregate_rows: list[dict[str, Any]] = []
    for sample_row in sorted(context["sample_rows"], key=lambda row: int(row["Index"])):
        sample_index = int(sample_row["Index"])
        model_rows = sorted(
            model_rows_by_sample.get(sample_index, []),
            key=lambda row: (
                -_to_float(row.get("BestCombinedPassRate")),
                -_to_float(row.get("BestProvidedPassRate")),
                str(row.get("TargetName", "")),
            ),
        )
        best_row = model_rows[0] if model_rows else {}
        second_row = model_rows[1] if len(model_rows) > 1 else {}
        best_attempt = attempt_lookup.get(
            (
                sample_index,
                str(best_row.get("TargetName", "")),
                int(best_row.get("BestAttemptIndex", 0) or 0),
            ),
            {},
        )
        aggregate_rows.append(
            {
                "Index": sample_index,
                "QuestionId": sample_row["QuestionId"],
                "QuestionTitle": sample_row["QuestionTitle"],
                "Difficulty": sample_row["Difficulty"],
                "ComparedModels": sample_row.get("ComparedModels", ""),
                "OracleProbeStatus": sample_row.get("OracleProbeStatus", ""),
                "OraclePassRate": _to_float(sample_row.get("OraclePassRate")),
                "GeneratedTests": int(sample_row.get("GeneratedTests", 0) or 0),
                "Winner": sample_row.get("Winner", ""),
                "WinnerModelLabel": best_row.get("ModelLabel", ""),
                "WinnerPublicPassRate": _to_float(best_attempt.get("PublicPassRate")),
                "WinnerPrivatePassRate": _to_float(best_attempt.get("PrivatePassRate")),
                "WinnerGeneratedPassRate": _to_float(best_row.get("BestGeneratedPassRate")),
                "WinnerProvidedPassRate": _to_float(best_row.get("BestProvidedPassRate")),
                "WinnerCombinedPassRate": _to_float(best_row.get("BestCombinedPassRate")),
                "WinnerCombinedPass@1": _to_float(best_row.get("CombinedPass@1")),
                "WinnerCombinedPass@2": _to_float(best_row.get("CombinedPass@2")),
                "AverageCombinedPassRateAcrossModels": round(
                    sum(_to_float(row.get("BestCombinedPassRate")) for row in model_rows) / len(model_rows),
                    4,
                )
                if model_rows
                else None,
                "CombinedPassRateDelta": round(
                    _to_float(best_row.get("BestCombinedPassRate"))
                    - _to_float(second_row.get("BestCombinedPassRate")),
                    4,
                )
                if second_row
                else None,
                "GeneratedPassRateDelta": round(
                    _to_float(best_row.get("BestGeneratedPassRate"))
                    - _to_float(second_row.get("BestGeneratedPassRate")),
                    4,
                )
                if second_row
                else None,
                "DifficultyEstimate": best_row.get("DifficultyEstimate", ""),
                "FailureCategory": best_row.get("FailureCategory", ""),
                "BenchmarkQualitySignal": best_row.get("BenchmarkQualitySignal", ""),
                "EfficacyLabel": best_row.get("EfficacyLabel", ""),
                "Suspicious": _to_bool(sample_row.get("Suspicious")),
                "NeedsAudit": _to_bool(sample_row.get("NeedsAudit")),
                "ComparisonNote": sample_row.get("ComparisonNote", ""),
            }
        )
    return aggregate_rows


def _write_combined_workbook(
    output_path: Path,
    *,
    stage1_detailed_rows: list[dict[str, Any]],
    stage1_summary_rows: list[dict[str, Any]],
    stage2_detailed_rows: list[dict[str, Any]],
    stage2_aggregate_rows: list[dict[str, Any]],
    test_columns: list[str],
    stage3_detailed_rows: list[dict[str, Any]],
    stage3_summary_rows: list[dict[str, Any]],
    stage3_test_columns: list[str],
    stage4_detailed_rows: list[dict[str, Any]],
    stage4_summary_rows: list[dict[str, Any]],
    stage4_test_columns: list[str],
) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Stage1_Detailed"
    _write_sheet(ws1, stage1_detailed_rows, percentage_columns=set())

    ws2 = wb.create_sheet("Stage1_Summary")
    _write_sheet(ws2, stage1_summary_rows, percentage_columns=set())

    ws3 = wb.create_sheet("Stage2_Detailed")
    percentage_columns = {
        "OraclePassRate",
        "PublicPassRate",
        "PrivatePassRate",
        "GeneratedPassRate",
        "ProvidedPassRate",
        "CombinedPassRate",
        "CombinedPass@1",
        "CombinedPass@2",
    }
    _write_sheet(ws3, stage2_detailed_rows, percentage_columns=percentage_columns, test_columns=set(test_columns))

    ws4 = wb.create_sheet("Stage2_Summary")
    aggregate_percentage_columns = {
        "OraclePassRate",
        "WinnerPublicPassRate",
        "WinnerPrivatePassRate",
        "WinnerGeneratedPassRate",
        "WinnerProvidedPassRate",
        "WinnerCombinedPassRate",
        "WinnerCombinedPass@1",
        "WinnerCombinedPass@2",
        "AverageCombinedPassRateAcrossModels",
        "CombinedPassRateDelta",
        "GeneratedPassRateDelta",
    }
    _write_sheet(
        ws4,
        stage2_aggregate_rows,
        percentage_columns=aggregate_percentage_columns,
    )
    if stage3_detailed_rows or stage3_summary_rows:
        ws5 = wb.create_sheet("Stage3_Detailed")
        stage3_percentage_columns = {
            "ClosestNeighborSimilarity",
        }
        _write_sheet(
            ws5,
            stage3_detailed_rows,
            percentage_columns=stage3_percentage_columns,
            test_columns=set(stage3_test_columns),
        )

        ws6 = wb.create_sheet("Stage3_Summary")
        _write_sheet(
            ws6,
            stage3_summary_rows,
            percentage_columns=set(),
        )
    if stage4_detailed_rows or stage4_summary_rows:
        ws7 = wb.create_sheet("Stage4_Detailed")
        stage4_percentage_columns = {"WinnerCombinedPassRate", "OraclePassRate"}
        _write_sheet(
            ws7,
            stage4_detailed_rows,
            percentage_columns=stage4_percentage_columns,
            test_columns=set(stage4_test_columns),
        )
        ws8 = wb.create_sheet("Stage4_Summary")
        _write_sheet(
            ws8,
            stage4_summary_rows,
            percentage_columns=set(),
        )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _write_combined_json(
    output_path: Path,
    *,
    timestamp: str,
    stage1_detailed_rows: list[dict[str, Any]],
    stage1_summary_rows: list[dict[str, Any]],
    stage2_detailed_rows: list[dict[str, Any]],
    stage2_aggregate_rows: list[dict[str, Any]],
    stage2_context: dict[str, Any],
    test_columns: list[str],
    stage3_context: dict[str, Any] | None,
    stage3_detailed_rows: list[dict[str, Any]],
    stage3_summary_rows: list[dict[str, Any]],
    stage3_test_columns: list[str],
    stage4_context: dict[str, Any] | None,
    stage4_detailed_rows: list[dict[str, Any]],
    stage4_summary_rows: list[dict[str, Any]],
    stage4_test_columns: list[str],
) -> None:
    payload = {
        "timestamp": timestamp,
        "stage1": {
            "detailed": stage1_detailed_rows,
            "summary": stage1_summary_rows,
        },
        "stage2": {
            "detailed": stage2_detailed_rows,
            "summary": stage2_aggregate_rows,
            "test_columns": test_columns,
            "sample_results": stage2_context["sample_rows"],
            "sample_model_results": stage2_context["model_rows"],
            "model_attempts": stage2_context["attempt_rows"],
            "per_test_results": stage2_context["per_test_rows"],
            "run_manifest": stage2_context["manifest"],
        },
    }
    if stage3_context:
        payload["stage3"] = {
            "detailed": stage3_detailed_rows,
            "summary": stage3_summary_rows,
            "test_columns": stage3_test_columns,
            "dataset_summary": stage3_context.get("dataset_summary", {}),
            "audit_queues": stage3_context.get("audit_queues", []),
            "duplicate_pairs": stage3_context.get("duplicate_pairs", []),
            "relationship_analysis": stage3_context.get("relationship_rules", []),
            "run_manifest": stage3_context.get("run_manifest", {}),
        }
    if stage4_context:
        payload["stage4"] = {
            "detailed": stage4_detailed_rows,
            "summary": stage4_summary_rows,
            "test_columns": stage4_test_columns,
            "run_manifest": stage4_context.get("run_manifest", {}),
        }
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _write_sheet(
    ws,
    rows: list[dict[str, Any]],
    *,
    percentage_columns: set[str],
    test_columns: set[str] | None = None,
) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in rows:
        ws.append([_excel_safe(row.get(header)) for header in headers])
    ws.freeze_panes = "A2"
    for idx, header in enumerate(headers, start=1):
        if header in percentage_columns:
            for col in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for cell in col:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.0%"
        if header in {"Prompt", "Ideal_Response", "Test_Cases", "EfficacyLabel"}:
            for col in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for cell in col:
                    _apply_label_fill(cell)
        if test_columns and header in test_columns:
            for col in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for cell in col:
                    _apply_test_fill(cell)
            ws.column_dimensions[get_column_letter(idx)].width = 14
        else:
            ws.column_dimensions[get_column_letter(idx)].width = max(14, min(44, len(header) + 2))


def _excel_safe(value: Any) -> Any:
    if isinstance(value, str):
        return _ILLEGAL_XLSX_CHARS.sub("", value)
    return value


def _stage2_note_request_id(model_row: dict[str, Any]) -> str:
    return (
        f"stage2:{int(model_row['Index'])}:"
        f"{str(model_row['TargetName'])}:{int(model_row['BestAttemptIndex'])}"
    )


def _build_stage2_note_request(
    *,
    model_row: dict[str, Any],
    attempt_row: dict[str, Any],
    test_values: dict[str, str],
    test_details: dict[str, dict[str, Any]],
) -> dict[str, Any] | None:
    flagged_columns = [
        column_name
        for column_name, value in sorted(test_values.items(), key=lambda item: _test_column_sort_key(item[0]))
        if value and value != "PASS"
    ]
    if not flagged_columns:
        return None
    lines = [
        "Write terse per-test notes for a Stage 2 benchmark audit row.",
        "Return notes only for the allowed test columns listed below.",
        f"Allowed columns: {', '.join(flagged_columns)}",
        "A note should point a human to the likely failure location or issue type, not fully explain the sample.",
        "",
        f"Sample index: {int(model_row['Index'])}",
        f"Question title: {model_row.get('QuestionTitle', '')}",
        f"Target: {model_row.get('TargetName', '')}",
        f"Efficacy label: {model_row.get('EfficacyLabel', '')}",
        f"Failure category: {model_row.get('FailureCategory', '')}",
        f"Execution probe status: {attempt_row.get('ExecutionProbeStatus', '')}",
        "",
        "Flagged tests:",
    ]
    for column_name in flagged_columns:
        lines.append(_format_stage2_test_note_line(column_name, test_values[column_name], test_details.get(column_name, {})))
    return {
        "request_id": _stage2_note_request_id(model_row),
        "prompt": "\n".join(lines),
        "allowed_columns": flagged_columns,
    }


def _format_stage2_test_note_line(column_name: str, value: str, detail: dict[str, Any]) -> str:
    parts = [f"- {column_name}: result={value}"]
    focus = str(detail.get("focus", "")).strip()
    failure_type = str(detail.get("failure_type", "")).strip()
    exception_type = str(detail.get("exception_type", "")).strip()
    exception_message = str(detail.get("exception_message", "")).strip()
    if failure_type:
        parts.append(f"failure_type={failure_type}")
    if exception_type:
        parts.append(f"exception_type={exception_type}")
    if focus:
        parts.append(f"focus={_shorten_text(focus, 160)}")
    if exception_message:
        parts.append(f"exception_message={_shorten_text(exception_message, 160)}")
    actual = _compact_json(detail.get("actual"))
    expected = _compact_json(detail.get("expected"))
    if expected:
        parts.append(f"expected={expected}")
    if actual:
        parts.append(f"actual={actual}")
    return "; ".join(parts)


def _compact_json(value: Any, limit: int = 160) -> str:
    if value is None or value == "" or value == {}:
        return ""
    try:
        rendered = json.dumps(value, ensure_ascii=False, sort_keys=True)
    except TypeError:
        rendered = str(value)
    return _shorten_text(rendered, limit)


def _shorten_text(value: str, limit: int) -> str:
    text = " ".join(str(value).split())
    if len(text) <= limit:
        return text
    return text[: limit - 3].rstrip() + "..."




def _apply_label_fill(cell) -> None:
    value = str(cell.value or "")
    if value in {"Usable", "High Efficacy"}:
        cell.fill = PatternFill("solid", fgColor="C6E0B4")
    elif value in {"Needs Fixing", "Moderate Efficacy"}:
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
    elif value in {"Unusable", "Low Efficacy", "Suspicious (Needs Audit)"}:
        cell.fill = PatternFill("solid", fgColor="F4CCCC")
    elif value == "Inconclusive":
        cell.fill = PatternFill("solid", fgColor="D9E2F3")


def _apply_test_fill(cell) -> None:
    value = str(cell.value or "")
    if value == "PASS":
        cell.fill = PatternFill("solid", fgColor="C6E0B4")
    elif value.startswith("FAIL"):
        cell.fill = PatternFill("solid", fgColor="F4CCCC")
    elif value.startswith("ERROR") or value.startswith("TIMEOUT"):
        cell.fill = PatternFill("solid", fgColor="FCE4D6")


def _read_sheet_rows(workbook, sheet_name: str) -> list[dict[str, Any]]:
    ws = workbook[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value) for value in rows[0]]
    payload: list[dict[str, Any]] = []
    for values in rows[1:]:
        payload.append(dict(zip(headers, values)))
    return payload


def _read_csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _read_jsonl_rows(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open(encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def _find_attempt_row(
    attempt_rows: list[dict[str, Any]],
    *,
    index: int,
    target_name: str,
    attempt_index: int,
) -> dict[str, Any]:
    for row in attempt_rows:
        if (
            int(row["Index"]) == index
            and str(row["TargetName"]) == target_name
            and int(row["AttemptIndex"]) == attempt_index
        ):
            return row
    return {}


def _normalize_attempt_row(row: dict[str, Any]) -> dict[str, Any]:
    if not row:
        return {}
    normalized = dict(row)
    for key in [
        "PublicPassRate",
        "PrivatePassRate",
        "GeneratedPassRate",
        "ProvidedPassRate",
        "CombinedPassRate",
    ]:
        normalized[key] = _to_float(row.get(key))
    return normalized


def _test_column_name(visibility: str, case_index: int) -> str:
    label = visibility.capitalize() if visibility else "Test"
    return f"{label}_{case_index + 1:02d}"


def _test_column_sort_key(label: str) -> tuple[int, int]:
    visibility, _, suffix = label.partition("_")
    order = {"Public": 0, "Private": 1, "Generated": 2}.get(visibility, 9)
    try:
        index = int(suffix)
    except ValueError:
        index = 999
    return (order, index)


def _test_cell_value(row: dict[str, Any]) -> str:
    status = str(row.get("status", "")).lower()
    if status == "pass":
        return "PASS"
    if status == "fail":
        failure_type = str(row.get("failure_type", "")).strip() or "incorrect_output"
        return f"FAIL:{failure_type}"
    if status == "timeout":
        return "TIMEOUT"
    if status == "error":
        failure_type = str(row.get("failure_type", "")).strip()
        exception_type = str(row.get("exception_type", "")).strip()
        detail = failure_type or exception_type or "error"
        return f"ERROR:{detail}"
    return status.upper()


def _to_float(value: Any) -> float | None:
    if value in {"", None}:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() == "true"
