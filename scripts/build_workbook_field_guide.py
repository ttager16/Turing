from __future__ import annotations

import ast
import csv
import re
import sys
import textwrap
from functools import lru_cache
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_ROOT = PROJECT_ROOT / "src"
if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))

from turing_takehome.stages.sample_requirements_analysis.audit_core import requirements as stage1_requirements
from turing_takehome.stages.sample_requirements_analysis.audit_core import schema as stage1_schema


OUTPUT_PATH = PROJECT_ROOT / "outputs" / "workbook_field_guide.xlsx"
STAGE1_WORKBOOK = PROJECT_ROOT / "outputs" / "sample_requirements_analysis" / "guideline_audit.xlsx"
STAGE2_RESULTS = PROJECT_ROOT / "outputs" / "sample_efficacy_analysis" / "sample_results.csv"
STAGE2_MODEL_RESULTS = PROJECT_ROOT / "outputs" / "sample_efficacy_analysis" / "sample_model_results.csv"
STAGE3_WORKBOOK = PROJECT_ROOT / "outputs" / "dataset_analysis" / "dataset_analysis.xlsx"


HEADERS = [
    "Field",
    "Intent",
    "How It's Computed",
    "Decision Logic",
    "LLM Prompt",
    "Class Values",
    "Code Location",
    "Subjectivity of Test",
]


GENERIC_VERDICTS = {
    "PASS": "PASS: clear evidence the requirement is satisfied.",
    "PARTIAL": "PARTIAL: some evidence of compliance, but incomplete or weak.",
    "FAIL": "FAIL: clear evidence the requirement is violated.",
    "UNCLEAR": "UNCLEAR: the available evidence is ambiguous or insufficient.",
    "NA": "NA: not applicable for this sample.",
}

SUBJECTIVITY_LABELS = {
    0: "0 -- objective test with objective classes",
    1: "1 -- objective test with subjectivity in class differentiation",
    2: "2 -- subjectivity inherent in the test",
}

USING_LLM = "Using an LLM"
DIRECT_EVALUATION = "Direct evaluation."
SEE_CODE_LOCATION = "See code location"

LLM_FONT = Font(color="1F4E78", italic=True)
DIRECT_FONT = Font(color="2F6F3E", italic=True)
SEE_CODE_FONT = Font(color="7F7F7F", italic=True)


def rel(path: str) -> str:
    return path.replace("\\", "/")


def format_class_values(lines: list[str]) -> str:
    return "\n".join(lines)


def generic_requirement_class_values(values: tuple[str, ...]) -> str:
    return format_class_values([GENERIC_VERDICTS[value] for value in values])


def metadata_row(field: str, intent: str, *, code_location: str) -> dict[str, str]:
    return {
        "Field": field,
        "Intent": intent,
        "How It's Computed": "",
        "Decision Logic": "",
        "LLM Prompt": "",
        "Class Values": "",
        "Code Location": code_location,
        "Subjectivity of Test": "",
    }


def subjectivity_label(level: int | None) -> str:
    if level is None:
        return ""
    return SUBJECTIVITY_LABELS.get(level, "")


def shorten_prompt(prompt: str) -> str:
    prompt = textwrap.dedent(prompt).strip()
    prompt = re.sub(r"\n{3,}", "\n\n", prompt)
    return prompt


def render_ast_text(source: str, node: ast.AST) -> str:
    if isinstance(node, ast.Constant) and isinstance(node.value, str):
        return node.value
    if isinstance(node, ast.JoinedStr):
        parts: list[str] = []
        for item in node.values:
            if isinstance(item, ast.Constant) and isinstance(item.value, str):
                parts.append(item.value)
            else:
                expr = ast.get_source_segment(source, item)
                expr = expr or "..."
                expr = expr.replace("{", "").replace("}", "")
                parts.append("{" + expr + "}")
        return "".join(parts)
    return ""


def parse_docstring_sections(docstring: str) -> dict[str, str]:
    sections: dict[str, str] = {}
    if not docstring:
        return sections
    current: str | None = None
    buffer: list[str] = []
    for raw_line in textwrap.dedent(docstring).splitlines():
        line = raw_line.rstrip()
        header_match = re.match(r"^([A-Za-z0-9 /_-]+):\s*$", line)
        if header_match:
            if current:
                sections[current] = "\n".join(buffer).strip()
            current = header_match.group(1).strip()
            buffer = []
            continue
        if current:
            buffer.append(line)
    if current:
        sections[current] = "\n".join(buffer).strip()
    return sections


@lru_cache(maxsize=None)
def stage1_module_info(module_path_str: str) -> dict[str, str | bool]:
    path = PROJECT_ROOT / module_path_str
    source = path.read_text(encoding="utf-8")
    tree = ast.parse(source)
    evaluate_source = ""
    for node in tree.body:
        if isinstance(node, ast.FunctionDef) and node.name == "evaluate":
            evaluate_source = ast.get_source_segment(source, node) or ""
            break
    prompt_text = ""
    for node in ast.walk(tree):
        if isinstance(node, ast.Assign):
            target_names = [target.id for target in node.targets if isinstance(target, ast.Name)]
            if "prompt" in target_names:
                prompt_text = render_ast_text(source, node.value)
                if prompt_text:
                    break
    doc = ast.get_docstring(tree) or ""
    sections = parse_docstring_sections(doc)
    ratio_match = re.search(r"verdict_from_ratio\(\s*ratio\s*,\s*([0-9.]+)\s*,\s*([0-9.]+)\s*\)", source)
    emitted_verdicts = [name for name in ["PASS", "PARTIAL", "FAIL", "UNCLEAR", "NA"] if name in evaluate_source]
    return {
        "source": source,
        "evaluate_source": evaluate_source,
        "uses_llm": "llm_judge(" in source,
        "prompt_text": shorten_prompt(prompt_text) if prompt_text else "",
        "evaluation_logic": " ".join(sections.get("Evaluation logic", "").split()),
        "subjectivity_notes": " ".join(sections.get("Subjectivity and failure modes", "").split()),
        "ratio_pass": ratio_match.group(1) if ratio_match else "",
        "ratio_partial": ratio_match.group(2) if ratio_match else "",
        "emitted_verdicts": ",".join(emitted_verdicts),
    }


def stage1_field_specific_values(requirement: stage1_requirements.Requirement, *, star_level: int, module_info: dict[str, str | bool]) -> str:
    emitted_raw = str(module_info.get("emitted_verdicts", "")).strip()
    emitted_values = tuple(value for value in emitted_raw.split(",") if value)
    values = emitted_values or requirement.values
    description = requirement.description.rstrip(".")
    ratio_pass = str(module_info.get("ratio_pass", "")).strip()
    ratio_partial = str(module_info.get("ratio_partial", "")).strip()
    if ratio_pass and ratio_partial:
        lines: list[str] = []
        if "PASS" in values:
            lines.append(f"PASS: {description} in essentially all inspected evidence ({ratio_pass}).")
        if "PARTIAL" in values:
            lines.append(f"PARTIAL: {description} in many but not all inspected cases (at least {ratio_partial}).")
        if "FAIL" in values:
            lines.append("FAIL: this requirement fails often enough to undermine the artifact.")
        if "UNCLEAR" in values:
            lines.append("UNCLEAR: the relevant artifact could not be inspected reliably.")
        if "NA" in values:
            lines.append("NA: this requirement does not apply to the sample.")
        return format_class_values(lines)
    if star_level == 0:
        if values in {("PASS", "FAIL"), ("PASS", "FAIL", "NA"), ("PASS", "FAIL", "UNCLEAR"), ("PASS", "FAIL", "UNCLEAR", "NA")}:
            lines = []
            if "PASS" in values:
                lines.append(f"PASS: {description}.")
            if "FAIL" in values:
                lines.append("FAIL: this requirement is not satisfied.")
            if "UNCLEAR" in values:
                lines.append("UNCLEAR: the evaluator could not establish the condition cleanly.")
            if "NA" in values:
                lines.append("NA: this requirement does not apply to the sample.")
            return format_class_values(lines)
        lines = []
        if "PASS" in values:
            lines.append(f"PASS: {description}.")
        if "PARTIAL" in values:
            lines.append(f"PARTIAL: {description}, but only for part of the inspected artifact.")
        if "FAIL" in values:
            lines.append("FAIL: this requirement is not satisfied.")
        if "UNCLEAR" in values:
            lines.append("UNCLEAR: the evaluator could not establish the condition cleanly.")
        if "NA" in values:
            lines.append("NA: this requirement does not apply to the sample.")
        return format_class_values(lines)
    lines = []
    if "PASS" in values:
        lines.append(f"PASS: {description}.")
    if "PARTIAL" in values:
        lines.append(f"PARTIAL: {description}, but incompletely or with caveats.")
    if "FAIL" in values:
        lines.append("FAIL: this requirement is not satisfied.")
    if "UNCLEAR" in values:
        lines.append("UNCLEAR: the evidence is too ambiguous to score confidently.")
    if "NA" in values:
        lines.append("NA: this requirement does not apply to the sample.")
    return format_class_values(lines)


def stage1_how_computed(requirement: stage1_requirements.Requirement, module_info: dict[str, str | bool]) -> str:
    if bool(module_info.get("uses_llm")):
        return USING_LLM
    ratio_pass = str(module_info.get("ratio_pass", "")).strip()
    if ratio_pass:
        category = requirement.section.lower()
        if category == "tests":
            return "Counts how often the supplied tests satisfy this requirement, then converts that ratio to a verdict."
        return "Measures compliance across the inspected artifact, then converts that ratio to a verdict."
    return DIRECT_EVALUATION


STAGE1_DECISION_OVERRIDES = {
    "p_structured_layout": "PASS when the prompt clearly separates the task contract into recognizable sections; PARTIAL when the structure is present but muddled; FAIL when the layout is largely unstructured.",
    "p_computational_limits_defined": "PASS when computational limits are stated when they matter; PARTIAL when they are hinted at but incomplete; FAIL when they appear necessary but are omitted.",
    "p_edge_cases_defined": "PASS when the prompt names the important edge cases; PARTIAL when it names only some; FAIL when edge-case behavior is left unstated.",
    "p_no_external_libs_stated": "PASS when the prompt clearly bans non-stdlib dependencies; PARTIAL when the prompt implies but does not state that constraint; FAIL when it leaves the dependency policy open.",
    "p_metadata_alignment": "PASS when the prompt, title, and function metadata point to the same task; PARTIAL when they are broadly similar but drift in emphasis; FAIL when they point at materially different tasks.",
    "p_measurable_objective": "PASS when success is concretely testable; PARTIAL when the objective is partly measurable but still loose; FAIL when success criteria stay vague.",
    "p_difficulty_balanced": "PASS when the constraints fit the intended difficulty; PARTIAL when they slightly under- or over-shoot; FAIL when they materially mis-target the benchmark difficulty.",
    "i_no_globals": "PASS when no mutable global state is used; PARTIAL when state leaks slightly but not enough to break the solution shape; FAIL when globals materially drive the solution.",
    "i_state_encapsulated": "PASS when state stays local or clearly encapsulated; PARTIAL when encapsulation is mixed; FAIL when important state management is left unscoped.",
    "i_consistent_naming_docs": "PASS when naming and docstrings are consistently clear; PARTIAL when readability is uneven; FAIL when naming or docs materially hinder understanding.",
    "i_single_entry_aligned": "PASS when exactly one prompt-aligned entry point is exposed; PARTIAL when the main entry exists but extra callable surface remains; FAIL when the main entry is missing or misaligned.",
    "i_helpers_for_repeated_logic": "PASS when repeated logic is sensibly factored; PARTIAL when factoring is uneven; FAIL when repetition materially hurts clarity or maintenance.",
    "i_no_redundant_memoization": "PASS when memoization is used once, clearly, or not needed; PARTIAL when caching is somewhat duplicated; FAIL when redundant memoization materially complicates the solution.",
    "i_no_nested_helpers": "PASS when helper logic stays properly scoped at module or class level; PARTIAL when limited nesting is present but tolerable; FAIL when nested helpers materially obscure the solution.",
    "t_single_call_per_test": "PASS when tests map cleanly to one entry-function call each; PARTIAL when a minority of tests bundle extra work; FAIL when multi-call tests are common.",
    "t_public_test1_matches_prompt_example": "PASS when public test 1 clearly matches the prompt example; PARTIAL when the match is approximate; FAIL when the sample example and first public test disagree.",
    "t_recommended_15_20_total": "PASS for roughly 15 to 20 total tests; PARTIAL when the total is close but outside that band; FAIL when the suite is far from the recommended size.",
    "t_json_escaping_valid": "PASS when object-formatted JSON strings parse cleanly; PARTIAL when escaping issues are limited; FAIL when malformed escaping is common enough to break trust.",
    "t_optional_values_included": "PASS when tests always provide optional/defaulted inputs explicitly; PARTIAL when some but not all do; FAIL when omitted optional values are common enough to change call shape.",
    "t_exception_tests_aligned": "PASS when error tests enforce prompt-defined behavior only; PARTIAL when the alignment is arguable; FAIL when tests require undeclared error behavior.",
    "t_not_large_or_redundant": "PASS when the suite is lean and non-repetitive; PARTIAL when some redundancy exists; FAIL when bloat or repetition materially reduces signal.",
    "v_coverage_confidence": "PASS when the prompt/tests appear to cover most of the intended behavior; PARTIAL when coverage is meaningful but incomplete; FAIL when major behavior appears untested.",
    "v_entry_name_consistent": "PASS when prompt, starter, tests, and ideal all use the same entry name; PARTIAL when one artifact drifts slightly; FAIL when names materially disagree.",
    "v_output_schema_aligned": "PASS when returned or asserted keys are prompt-defined; PARTIAL when some outputs are only partly described; FAIL when tests enforce undocumented output schema.",
    "s_necessary_imports": "PASS when starter imports cover the names it exposes; PARTIAL when only some needed imports are present; FAIL when missing imports would break direct use of the starter.",
    "s_only_entry_signature": "PASS when the starter is essentially just the entry signature; PARTIAL when it adds small extra scaffolding; FAIL when it exposes substantial extra code.",
}


def stage1_decision_logic(requirement_key: str, requirement: stage1_requirements.Requirement, module_info: dict[str, str | bool]) -> str:
    ratio_pass = str(module_info.get("ratio_pass", "")).strip()
    ratio_partial = str(module_info.get("ratio_partial", "")).strip()
    emitted_raw = str(module_info.get("emitted_verdicts", "")).strip()
    values = tuple(value for value in emitted_raw.split(",") if value) or requirement.values
    if requirement_key in STAGE1_DECISION_OVERRIDES:
        return STAGE1_DECISION_OVERRIDES[requirement_key]
    if ratio_pass and ratio_partial:
        return f"PASS at {ratio_pass}; PARTIAL at {ratio_partial} to below {ratio_pass}; FAIL below {ratio_partial}."
    if bool(module_info.get("uses_llm")):
        return ""
    evaluate_source = str(module_info.get("evaluate_source", ""))
    if "runtime.get('callable_found') and runtime.get('executed')" in evaluate_source:
        return "PASS when the ideal response loads and the entry function resolves; FAIL when it does not."
    if "runtime.get('total', 0) and runtime['failed'] == 0" in evaluate_source and "runtime.get('passed', 0) > 0" in evaluate_source:
        return "PASS when all provided tests pass; PARTIAL when some pass; FAIL when none pass."
    if values in {("PASS", "FAIL"), ("PASS", "FAIL", "NA"), ("PASS", "FAIL", "UNCLEAR"), ("PASS", "FAIL", "UNCLEAR", "NA")}:
        return ""
    if "PARTIAL" in values:
        return SEE_CODE_LOCATION
    return ""


def stage1_llm_prompt(module_info: dict[str, str | bool]) -> str:
    if not bool(module_info.get("uses_llm")):
        return ""
    return str(module_info.get("prompt_text", "")).strip()


def stage2_solver_prompt_text() -> str:
    return shorten_prompt(
        """
        Solve the following Python coding task.

        Requirements:
        - Return only Python code.
        - Implement the requested entry function exactly.
        - Do not include explanations before or after the code.
        - Use only the Python standard library unless the prompt explicitly permits otherwise.

        Problem Statement:
        {sample.question_content}

        Starter Code:
        ```python
        {starter_code}
        ```
        """
    )


def stage2_generated_test_prompt_text() -> str:
    return shorten_prompt(
        """
        Generate {count} additional Python function test inputs for this benchmark sample.

        Rules:
        - Return only test inputs, not expected outputs.
        - Focus on edge cases, boundary conditions, format-sensitive cases, and cases likely to separate strong from merely pattern-matching solutions.
        - Do not require external libraries.
        - Keep inputs valid under the problem statement.
        - Match the existing function-call argument shape.
        - Each input_lines item must be one JSON-encoded argument value, matching how the existing tests are stored.

        Function name: {sample.function_name}

        Problem Statement:
        {sample.question_content}

        Starter Code:
        {sample.starter_code}

        Public test examples:
        {public_examples_json}
        """
    )


def stage3_auditor_prompt_text() -> str:
    return shorten_prompt(
        """
        Judge this sample only as a dataset-audit item within a benchmark.
        Focus on benchmark utility, not coding quality.
        Return JSON only.

        Index: {Index}
        Title: {QuestionTitle}
        Difficulty: {Difficulty}
        WinnerCombinedPassRate: {WinnerCombinedPassRate}

        Choose:
        - dataset_utility_label: strong / usable / caveated / contradictory / saturated
        - primary_risk: none / redundancy / instability / benchmark_defect / threshold_fragility / outlier / contradiction
        - audit_priority: normal / medium / high / critical
        """
    )


def stage3_notes_prompt_text() -> str:
    return shorten_prompt(
        """
        System:
        You write extremely concise benchmark-audit notes. Return JSON only. Prefer a short phrase. Use a full sentence only when needed.

        User:
        Write terse per-test notes for a Stage 3 dataset-audit row.
        Return notes only for the allowed Stage 3 columns listed below.
        Allowed columns: {allowed_columns}
        Explain just enough for a human auditor to know where to inspect next.

        Sample index: {Index}
        Question title: {QuestionTitle}
        Difficulty: {Difficulty}

        Flagged Stage 3 checks:
        {flag_lines}
        """
    )


def stage2_subjectivity_for_field(field: str) -> str:
    subjective = {
        "ComparisonNote",
        "BenchmarkQualitySignal",
        "EfficacyLabel",
        "DifficultyEstimate",
        "FailureCategory",
        "Suspicious",
        "NeedsAudit",
        "Winner",
    }
    if field in subjective:
        return subjectivity_label(1)
    return subjectivity_label(0)


def stage3_subjectivity_for_field(field: str) -> str:
    level_one = {
        "PromptLengthBand",
        "StarterCodeLengthBand",
        "IdealResponseLengthBand",
        "TestCountBand",
        "RedundancyLabel",
        "PromptLengthOutlier",
        "StarterCodeOutlier",
        "IdealResponseOutlier",
        "TestCountOutlier",
        "PerformanceOutlier",
        "RedundancyStatus",
        "AttemptVarianceCheck",
        "AttemptVarianceLabel",
        "ModelDisagreementCheck",
        "ModelDisagreementLabel",
        "ThresholdSensitivityCheck",
        "ThresholdSensitivityLabel",
        "DifficultySignalRegime",
        "ContradictionCheck",
        "BenchmarkDefectCandidate",
        "TrivialityCheck",
        "ExemplarCheck",
        "AuditPriority",
    }
    level_two = {"ModelDisagreementSource"}
    if field in level_two:
        return subjectivity_label(2)
    if field in level_one:
        return subjectivity_label(1)
    return subjectivity_label(0)


def stage2_field_definition(field: str) -> dict[str, str]:
    if field in STAGE2_FIELD_DEFS:
        definition = dict(STAGE2_FIELD_DEFS[field])
    else:
        comparison_match = re.match(r"Model([AB])(.*)", field)
        if comparison_match and comparison_match.group(2) in STAGE2_FIELD_DEFS:
            base_field = comparison_match.group(2)
            base = STAGE2_FIELD_DEFS[base_field]
            definition = dict(base)
            definition["Intent"] = f"Mirror {comparison_match.group(1)}'s `{base_field}` value into the comparison export."
            definition["How It's Computed"] = "Copied from the corresponding per-model Stage 2 aggregate row."
            definition["Decision Logic"] = ""
            definition["Code Location"] = rel("src/turing_takehome/stages/sample_efficacy_analysis/runner.py :: _build_comparison_row")
        else:
            definition = {}
    return definition


def stage2_llm_prompt_for_field(field: str) -> str:
    generated_prompt_fields = {
        "GeneratedTests",
        "BestGeneratedPassRate",
        "ModelABestGeneratedPassRate",
    }
    solver_prompt_fields = {
        "Attempts",
        "BestAttemptIndex",
        "BestProvidedPassRate",
        "BestGeneratedPassRate",
        "BestCombinedPassRate",
        "AverageCombinedPassRate",
        "ProvidedSuccesses",
        "CombinedSuccesses",
        "ProvidedPass@1",
        "CombinedPass@1",
        "ProvidedPass@2",
        "CombinedPass@2",
        "DifficultyEstimate",
        "FailureCategory",
        "BenchmarkQualitySignal",
        "EfficacyLabel",
        "Suspicious",
        "NeedsAudit",
        "Winner",
        "ComparisonNote",
        "ModelABestCombinedPassRate",
        "ModelABestProvidedPassRate",
        "ModelAEfficacyLabel",
        "ModelABenchmarkQualitySignal",
        "ModelACombinedPass@1",
        "ModelACombinedPass@2",
    }
    if field in generated_prompt_fields:
        return stage2_generated_test_prompt_text()
    if field in solver_prompt_fields:
        return stage2_solver_prompt_text()
    return ""


def clean_definition_cell(value: str | None, *, allow_blank: bool = True) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    low_value_patterns = (
        "computed while building",
        "computed while aggregating",
        "comparison-surface field",
        "model-summary field",
        "descriptive only",
        "see the code location for the exact export logic",
        "see the field values in the comparison export",
        "see the exported field values",
    )
    if any(pattern in text.lower() for pattern in low_value_patterns):
        return "" if allow_blank else SEE_CODE_LOCATION
    return text


def normalize_llm_prompt(value: str | None, *, field: str, stage: str) -> str:
    text = str(value or "").strip()
    if stage == "stage2":
        return stage2_llm_prompt_for_field(field)
    if stage == "stage3":
        if field.startswith("Notes-"):
            return stage3_notes_prompt_text()
        if field in {"ModelDisagreementSource", "Stage3AuditorUtilityLabels", "Stage3AuditorPrimaryRisks", "Stage3AuditorPriorities"}:
            return stage3_auditor_prompt_text()
        if "embedding" in field.lower():
            return ""
    return text


def stage2_class_values(field: str, raw: str | None, *, surface: str) -> str:
    text = clean_definition_cell(raw)
    mirrored = surface == "Comparison"
    if text:
        if text == "Fraction from 0.0 to 1.0.":
            if "Provided" in field:
                if field == "BestProvidedPassRate":
                    return "0.0 to 1.0 best pass fraction over provided tests."
                if field == "ProvidedPass@1":
                    return "0.0 to 1.0 estimated probability that at least one attempt passes the provided tests."
                if field == "ProvidedPass@2":
                    return "0.0 to 1.0 estimated probability that at least one of two attempts passes the provided tests."
                return "0.0 to 1.0 fraction over provided tests." if not mirrored else "0.0 to 1.0 provided-test pass fraction mirrored into the comparison row."
            if "Generated" in field:
                return "0.0 to 1.0 fraction over LLM-generated tests." if not mirrored else "0.0 to 1.0 generated-test pass fraction mirrored into the comparison row."
            if "Combined" in field:
                if field == "AverageCombinedPassRate":
                    return "0.0 to 1.0 average combined pass fraction across attempts."
                if "Pass@" in field:
                    return f"0.0 to 1.0 estimated {field.replace('Combined', '').lower()} success probability over provided plus generated tests."
                return "0.0 to 1.0 fraction over provided plus generated tests." if not mirrored else "0.0 to 1.0 combined pass fraction mirrored into the comparison row."
        if text == "Nonnegative integer count.":
            if field == "GeneratedTests":
                return "Count of accepted generated tests added in Stage 2." if not mirrored else "Count of generated tests used by the mirrored comparison row."
            if field == "ProvidedSuccesses":
                return "Count of attempts that fully passed the provided tests."
            if field == "CombinedSuccesses":
                return "Count of attempts that fully passed provided plus generated tests."
        if text == "Positive integer count." and field == "Attempts":
            return "Count of candidate attempts evaluated for this model/sample pair."
        if field.endswith("OraclePassRate") or field == "OraclePassRate":
            return "0.0 to 1.0 oracle pass fraction under the Stage 2 harness." if not mirrored else "0.0 to 1.0 oracle pass fraction surfaced in the comparison row."
        if field.endswith("BestCombinedPassRate") or field == "BestCombinedPassRate":
            return "0.0 to 1.0 best combined pass fraction for this model/sample pair." if not mirrored else "0.0 to 1.0 best combined pass fraction mirrored into the comparison row."
        if field.endswith("GeneratedTests") or field == "GeneratedTests":
            return "Count of accepted generated tests used for this model/sample pair." if not mirrored else "Count of accepted generated tests used in the comparison row."
        if field.endswith("Suspicious") or field == "Suspicious":
            return "False: this model/sample result has no Stage 2 suspicion flag.\nTrue: this model/sample result should be treated as caveated." if not mirrored else "False: the comparison row has no Stage 2 suspicion flag.\nTrue: the comparison row should be treated as caveated."
        if field.endswith("NeedsAudit") or field == "NeedsAudit":
            return "False: no immediate Stage 2 audit request for this model/sample row.\nTrue: escalate this model/sample row to later review." if not mirrored else "False: no immediate Stage 2 audit request for the comparison row.\nTrue: escalate the comparison row to later review."
        if field.endswith("EfficacyLabel") or field == "EfficacyLabel":
            return text if not mirrored else text + "\nThis is the mirrored Stage 2 efficacy label used in the comparison row."
        if field.endswith("BenchmarkQualitySignal") or field == "BenchmarkQualitySignal":
            return text if not mirrored else text + "\nThis is the mirrored benchmark-quality signal used in the comparison row."
        return text
    return ""


def stage2_decision_logic(field: str, raw: str | None, *, surface: str) -> str:
    text = clean_definition_cell(raw)
    if text:
        return text
    if surface == "Comparison" and field.startswith("ModelA"):
        return "Copied directly from the corresponding per-model Stage 2 aggregate row."
    if field == "ComparedModels":
        return "Reported directly as the model targets present in the comparison row."
    if field == "GeneratedTests":
        return "Reported directly as the accepted generated-test count."
    if field == "Attempts":
        return "Reported directly as the number of recorded attempts."
    if field == "AverageCombinedPassRate":
        return "Arithmetic mean across the attempt-level combined pass rates."
    return ""


def stage3_class_values(field: str, raw: str | None) -> str:
    text = clean_definition_cell(raw)
    if text == "PASS\nFLAG":
        mapping = {
            "PromptLengthOutlier": "PASS: prompt length is within the normal dataset range.\nFLAG: prompt length is an outlier.",
            "StarterCodeOutlier": "PASS: starter code length is within the normal dataset range.\nFLAG: starter code length is an outlier.",
            "IdealResponseOutlier": "PASS: ideal-response length is within the normal dataset range.\nFLAG: ideal-response length is an outlier.",
            "TestCountOutlier": "PASS: test-count volume is within the normal dataset range.\nFLAG: test-count volume is an outlier.",
            "PerformanceOutlier": "PASS: behavioral performance is within the normal dataset range.\nFLAG: behavioral performance is an outlier.",
            "RedundancyStatus": "PASS: no material near-duplicate signal.\nFLAG: redundancy score is high enough to warrant deduplication review.",
            "AttemptVarianceCheck": "PASS: repeated attempts are behaviorally stable.\nFLAG: repeated attempts vary enough to weaken confidence.",
            "ModelDisagreementCheck": "PASS: model or proxy views are aligned.\nFLAG: disagreement is material enough to inspect.",
            "ThresholdSensitivityCheck": "PASS: interpretation is stable under small threshold shifts.\nFLAG: interpretation sits close to a threshold boundary.",
            "BenchmarkDefectCandidate": "PASS: no strong benchmark-defect signal.\nFLAG: Stage 3 thinks the benchmark artifact may be broken or misaligned.",
            "TrivialityCheck": "PASS: sample is not obviously saturated.\nFLAG: sample looks trivial or too easy to add much signal.",
            "ExemplarCheck": "PASS: not a clean exemplar candidate.\nFLAG: sample looks like a strong exemplar of intended benchmark behavior.",
        }
        return mapping.get(field, text)
    if text == "Similarity score from 0.0 to 1.0.":
        if field == "ClosestNeighborSimilarity":
            return "0.0 to 1.0 combined redundancy similarity score."
        if field == "ClosestNeighborLexicalSimilarity":
            return "0.0 to 1.0 lexical overlap score."
        if field == "ClosestNeighborEmbeddingSimilarity":
            return "0.0 to 1.0 embedding similarity score."
    if text == "Positive integer count." and field == "RedundancyClusterSize":
        return "Number of samples in the same redundancy cluster."
    if text == "INFO descriptive distribution.":
        if field == "Summary::prompt_length_distribution":
            return "Distribution over prompt-length bands."
        if field == "Summary::test_count_distribution":
            return "Distribution over test-count bands."
    return text


def stage3_decision_logic(field: str, raw: str | None) -> str:
    text = clean_definition_cell(raw)
    if text:
        return text
    if field == "AuditReason":
        return "Synthesizes the strongest signals that pushed the row into its current audit-priority bucket."
    if field == "ModelDisagreementSource":
        return "single_model when only one model is available, stage2_models when real model comparison exists, and stage3_auditors when the disagreement proxy came from the Stage 3 auditor pair."
    return ""


def load_csv_headers(path: Path) -> list[str]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return next(csv.reader(handle))


def load_stage3_summary_test_names(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Summary"]
    rows = list(ws.iter_rows(values_only=True))
    names: list[str] = []
    for row in rows[1:]:
        test_name = str(row[1] or "").strip()
        if test_name and test_name not in names:
            names.append(test_name)
    return names


def stage1_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    rows.extend(
        [
            metadata_row(
                "Detailed::Index",
                "Zero-based sample index in the source JSONL.",
                code_location=rel("src/turing_takehome/stages/sample_requirements_analysis/runner.py :: run_cli"),
            ),
            metadata_row(
                "Detailed::Question_Id",
                "Source question identifier.",
                code_location=rel("src/turing_takehome/stages/sample_requirements_analysis/audit_core/context.py :: collect_contexts"),
            ),
            metadata_row(
                "Detailed::Question_Title",
                "Source title or coarse task label.",
                code_location=rel("src/turing_takehome/stages/sample_requirements_analysis/audit_core/context.py :: collect_contexts"),
            ),
            metadata_row(
                "Detailed::Difficulty",
                "Source difficulty label from the dataset payload.",
                code_location=rel("src/turing_takehome/stages/sample_requirements_analysis/audit_core/context.py :: collect_contexts"),
            ),
            metadata_row(
                "Detailed::Function_Name",
                "Entry-point function name inferred from the sample metadata.",
                code_location=rel("src/turing_takehome/stages/sample_requirements_analysis/audit_core/context.py :: collect_contexts"),
            ),
            {
                "Field": "Detailed::Runtime_Pass_Rate",
                "Intent": "Report how the provided ideal response performed on the supplied tests during Stage 1 dry-run validation.",
                "How It's Computed": "Dry-runs the provided ideal response against the supplied tests and records passed/total.",
                "Decision Logic": "Reported directly as a fraction; Stage 1 summary logic uses it when classifying the ideal response.",
                "LLM Prompt": "",
                "Class Values": "Fraction from 0.0 to 1.0.\nHigher is better.\n1.0 means the ideal response passed all supplied tests in the Stage 1 dry run.",
                "Code Location": rel("src/turing_takehome/stages/sample_requirements_analysis/audit_core/context.py :: ideal_runtime"),
                "Subjectivity of Test": subjectivity_label(0),
            },
            {
                "Field": "Summary::Prompt",
                "Intent": "Roll up all prompt-centric Stage 1 evidence into a single evaluator-facing status.",
                "How It's Computed": "Counts prompt-related detailed failures, partials, and unclear verdicts, then applies hard-stop prompt invalidity rules.",
                "Decision Logic": "Unusable if the function signature is missing, or if multiple foundational prompt-alignment failures co-occur. Needs Fixing if any prompt-related detailed check is FAIL, PARTIAL, or UNCLEAR. Otherwise Usable.",
                "LLM Prompt": "",
                "Class Values": "Usable: no prompt-side detailed issues remain.\nNeeds Fixing: prompt-side evidence is mixed, incomplete, or caveated.\nUnusable: the prompt contract is too broken to trust as a benchmark artifact.",
                "Code Location": rel("src/turing_takehome/stages/sample_requirements_analysis/audit_core/workbook.py :: classify_prompt"),
                "Subjectivity of Test": subjectivity_label(1),
            },
            {
                "Field": "Summary::Ideal_Response",
                "Intent": "Roll up all ideal-solution evidence into a single evaluator-facing status.",
                "How It's Computed": "Combines ideal-response detailed checks with dry-run execution and pass-rate evidence.",
                "Decision Logic": "Unusable if the ideal response does not execute, fails its internal tests, or the dry-run pass rate is below 0.5. Needs Fixing if any ideal-related detailed check is FAIL, PARTIAL, or UNCLEAR, or if runtime pass rate is below 1.0. Otherwise Usable.",
                "LLM Prompt": "",
                "Class Values": "Usable: the ideal response is structurally clean and fully validated by the supplied tests.\nNeeds Fixing: the ideal response is directionally serviceable but has structural or validation gaps.\nUnusable: the ideal response is too broken or too misaligned to serve as trustworthy ground truth.",
                "Code Location": rel("src/turing_takehome/stages/sample_requirements_analysis/audit_core/workbook.py :: classify_ideal"),
                "Subjectivity of Test": subjectivity_label(1),
            },
            {
                "Field": "Summary::Test_Cases",
                "Intent": "Roll up all test-suite evidence into a single evaluator-facing status.",
                "How It's Computed": "Counts test-related detailed failures, partials, and unclear verdicts, then applies hard-stop invalidity checks for malformed or misaligned tests.",
                "Decision Logic": "Unusable if the tests are not valid JSON containers, do not use the expected string-field format, inject extra parameters, or fail multiple core alignment checks together. Needs Fixing if any test-related detailed check is FAIL, PARTIAL, or UNCLEAR. Otherwise Usable.",
                "LLM Prompt": "",
                "Class Values": "Usable: the supplied tests are structurally sound and directionally aligned with the benchmark contract.\nNeeds Fixing: the tests are parseable and partly useful, but contain coverage or alignment weaknesses.\nUnusable: the tests are malformed or too misaligned to trust as benchmark evidence.",
                "Code Location": rel("src/turing_takehome/stages/sample_requirements_analysis/audit_core/workbook.py :: classify_tests"),
                "Subjectivity of Test": subjectivity_label(1),
            },
        ]
    )

    requirement_lookup = {requirement.key: requirement for requirement in stage1_requirements.REQUIREMENTS}
    for key in stage1_schema.DETAILED_KEYS:
        requirement = requirement_lookup[key]
        display_name = stage1_schema.display_name(key)
        section_number, _, _ = stage1_schema.SECTION_MAP[key]
        section_dir = stage1_schema.section_folder_name(section_number)
        relative_module_path = rel(f"src/turing_takehome/stages/sample_requirements_analysis/{section_dir}/{display_name}.py")
        module_path = f"{relative_module_path} :: evaluate"
        star_level = stage1_schema.STAR_LEVELS.get(key, 0)
        module_info = stage1_module_info(relative_module_path)
        rows.append(
            {
                "Field": f"Detailed::{display_name}",
                "Intent": requirement.description,
                "How It's Computed": stage1_how_computed(requirement, module_info),
                "Decision Logic": stage1_decision_logic(key, requirement, module_info),
                "LLM Prompt": stage1_llm_prompt(module_info),
                "Class Values": stage1_field_specific_values(requirement, star_level=star_level, module_info=module_info),
                "Code Location": module_path,
                "Subjectivity of Test": subjectivity_label(star_level),
            }
        )
    return rows


STAGE2_FIELD_DEFS: dict[str, dict[str, str]] = {
    "ComparedModels": {
        "Intent": "Record which model targets were compared in the sample-level comparison row.",
        "How It's Computed": "Joins the ordered model target names present in the comparison surface.",
        "Decision Logic": "Descriptive only.",
        "Class Values": "Comma-separated target names.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/runner.py :: _build_comparison_row"),
    },
    "OracleProbeStatus": {
        "Intent": "Show whether the provided ideal response executed cleanly during Stage 2 oracle validation.",
        "How It's Computed": "Runs the ideal response through the same execution harness used for candidate evaluation.",
        "Decision Logic": "If the oracle fails, Stage 2 marks the sample suspicious rather than treating downstream failures as clean model evidence.",
        "Class Values": "ok: oracle execution succeeded.\nnot_run: oracle was skipped.\nexecution_failure or timeout-like status: oracle validation did not complete cleanly.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/runner.py :: _oracle_test_rows"),
    },
    "OraclePassRate": {
        "Intent": "Measure how fully the provided ideal response passed the supplied tests under the Stage 2 harness.",
        "How It's Computed": "Pass fraction from the oracle run over provided tests.",
        "Decision Logic": "Values below 1.0 are treated as benchmark suspicion in Stage 2 classification.",
        "Class Values": "Fraction from 0.0 to 1.0.\n1.0 is the expected clean value for a trustworthy sample.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/runner.py :: _oracle_test_rows"),
    },
    "GeneratedTests": {
        "Intent": "Record how many LLM-generated extra tests were used to supplement the provided tests.",
        "How It's Computed": "Counts generated test cases accepted into the Stage 2 run.",
        "Decision Logic": "Descriptive only, but several comparison notes and quality signals refer to generated-test behavior.",
        "LLM Prompt": "Stage 2 generated-tests prompt; see artifacts/audit/stage2/.",
        "Class Values": "Nonnegative integer count.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/runner.py :: _build_generated_tests"),
    },
    "Winner": {
        "Intent": "Identify the stronger target in the sample-level comparison row.",
        "How It's Computed": "Compares best combined pass rate first and generated-test pass rate second.",
        "Decision Logic": "If only one model is present, that model is the effective winner by default.",
        "Class Values": "Target name or comparison outcome.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/runner.py :: _winner"),
    },
    "ComparisonNote": {
        "Intent": "Provide a terse human-readable summary of what mattered in the comparison row.",
        "How It's Computed": "Synthesizes winner, suspiciousness, and pass-rate differences into a short note.",
        "Decision Logic": "Used as reviewer context rather than as a formal label.",
        "Class Values": "Short free-text note.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/runner.py :: _comparison_note"),
    },
    "Suspicious": {
        "Intent": "Flag rows whose behavior suggests benchmark or grader problems rather than clean model capability signal.",
        "How It's Computed": "Inherited from Stage 2 classification or aggregated across compared models.",
        "Decision Logic": "True when oracle behavior is broken, public/private results diverge sharply, or another benchmark-quality caveat is triggered.",
        "Class Values": "False: no Stage 2 suspicion flag.\nTrue: sample should be treated as caveated and likely audited further.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/labeling.py :: classify_sample"),
    },
    "NeedsAudit": {
        "Intent": "Mark rows that should be routed to tighter inspection rather than trusted at face value.",
        "How It's Computed": "Inherited from Stage 2 classification or aggregated across compared models.",
        "Decision Logic": "Usually tracks benchmark-quality ambiguity or suspiciousness.",
        "Class Values": "False: no immediate Stage 2 audit request.\nTrue: escalate to later audit or human review.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/labeling.py :: classify_sample"),
    },
    "Attempts": {
        "Intent": "Count how many candidate attempts were evaluated for a target on this sample.",
        "How It's Computed": "Length of the target's attempt record list.",
        "Decision Logic": "Descriptive only.",
        "Class Values": "Positive integer count.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/runner.py :: _aggregate_model_attempts"),
    },
    "BestAttemptIndex": {
        "Intent": "Show which attempt was selected as the representative best attempt.",
        "How It's Computed": "Chooses the attempt with the highest combined pass rate, then provided pass rate, then generated pass rate.",
        "Decision Logic": "Tie-breaking is deterministic within the attempt aggregator.",
        "Class Values": "Zero-based attempt index.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/runner.py :: _aggregate_model_attempts"),
    },
    "BestProvidedPassRate": {
        "Intent": "Measure the best attempt's success on supplied public and private tests only.",
        "How It's Computed": "Best attempt's provided-test pass fraction.",
        "Decision Logic": "Higher is better; used in tie-breaking and reviewer interpretation.",
        "Class Values": "Fraction from 0.0 to 1.0.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/runner.py :: _aggregate_model_attempts"),
    },
    "BestGeneratedPassRate": {
        "Intent": "Measure the best attempt's success on LLM-generated tests.",
        "How It's Computed": "Best attempt's generated-test pass fraction.",
        "Decision Logic": "Higher is better; can expose edge-case weaknesses even when provided tests pass cleanly.",
        "LLM Prompt": "Stage 2 generated-tests prompt; see artifacts/audit/stage2/.",
        "Class Values": "Fraction from 0.0 to 1.0.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/runner.py :: _aggregate_model_attempts"),
    },
    "BestCombinedPassRate": {
        "Intent": "Provide the main Stage 2 behavioral success metric for a target on a sample.",
        "How It's Computed": "Best attempt pass fraction over provided plus generated tests.",
        "Decision Logic": "Feeds the Stage 2 difficulty, efficacy, and comparison logic.",
        "Class Values": "Fraction from 0.0 to 1.0.\nHigher means the sample was easier for the evaluated target.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/runner.py :: _aggregate_model_attempts"),
    },
    "AverageCombinedPassRate": {
        "Intent": "Show average behavioral performance across attempts rather than only the best attempt.",
        "How It's Computed": "Mean combined pass rate across attempts for the target.",
        "Decision Logic": "Descriptive only, but useful when reading attempt variance later in Stage 3.",
        "Class Values": "Fraction from 0.0 to 1.0.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/runner.py :: _aggregate_model_attempts"),
    },
    "ProvidedSuccesses": {
        "Intent": "Count how many attempts fully passed the provided tests.",
        "How It's Computed": "Counts attempts with ok execution and provided pass rate at or near 1.0.",
        "Decision Logic": "Used to compute pass@k over provided tests.",
        "Class Values": "Nonnegative integer count.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/runner.py :: _aggregate_model_attempts"),
    },
    "CombinedSuccesses": {
        "Intent": "Count how many attempts fully passed both provided and generated tests.",
        "How It's Computed": "Counts attempts with ok execution and combined pass rate at or near 1.0.",
        "Decision Logic": "Used to compute combined pass@k.",
        "Class Values": "Nonnegative integer count.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/runner.py :: _aggregate_model_attempts"),
    },
    "DifficultyEstimate": {
        "Intent": "Give a coarse difficulty regime inferred from behavioral results.",
        "How It's Computed": "Derived from generation status, execution status, oracle quality, and pass-rate profile.",
        "Decision Logic": "Trivial when pass rate is very high, extreme when pass rate is zero, reasonable for the middle band, unknown when execution never produced a valid signal.",
        "Class Values": "trivial: candidate passed almost everything.\nreasonable: candidate exposed informative failures without looking broken.\nextreme: candidate failed everything after a clean run.\nunknown: generation, execution, or oracle conditions prevented clean interpretation.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/labeling.py :: classify_sample"),
    },
    "FailureCategory": {
        "Intent": "Explain the dominant way a sample failed to produce clean signal in Stage 2.",
        "How It's Computed": "Derived from generation status, execution status, oracle status, divergence patterns, and pass-rate regime.",
        "Decision Logic": "Generation and execution failures are handled before behavioral interpretation. Clean behavioral failures are separated from likely benchmark-suspicion cases.",
        "Class Values": "generation_failure: no trustworthy candidate was produced.\nexecution_failure: candidate generation succeeded but execution did not.\nbenchmark_suspicion: oracle or public/private behavior suggests the benchmark itself may be compromised.\nlogical_or_edge_failure: model behavior failed on substantive tests.\nlogical_or_test_failure: zero-pass outcome after a clean run.\nclean_pass: near-perfect behavioral success.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/labeling.py :: classify_sample"),
    },
    "BenchmarkQualitySignal": {
        "Intent": "Describe whether Stage 2 thinks the sample looks like clean benchmark signal or a caveated artifact.",
        "How It's Computed": "Assigned alongside efficacy labels from execution, oracle, and pass-rate behavior.",
        "Decision Logic": "Oracle problems and public/private divergence escalate the label away from clean_evaluation.",
        "Class Values": "clean_evaluation: no immediate benchmark-quality warning from Stage 2.\nmisaligned_or_broken: the oracle itself appears broken or misaligned.\nhidden_or_underspecified_requirements: public/private behavior suggests hidden requirements.\ngenerated_tests_expose_weakness: provided tests looked easy, but generated tests found real gaps.\nambiguous_or_brittle: behavior looks caveated, often because formatting mismatches dominate.\npublic_private_divergence: public and private behavior diverge too sharply to trust cleanly.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/labeling.py :: classify_sample"),
    },
    "EfficacyLabel": {
        "Intent": "Summarize how informative the sample was at eliciting model failure after Stage 2 execution.",
        "How It's Computed": "Bucketed from pass rate plus benchmark-quality overrides.",
        "Decision Logic": "Inconclusive if generation or execution never completed. Suspicious when oracle or divergence logic dominates. Low Efficacy when pass rate is at least 0.95. High Efficacy when generated tests expose weakness or pass rate is between 0.25 and 0.8. Moderate Efficacy otherwise.",
        "Class Values": "Inconclusive: no clean candidate execution signal.\nSuspicious (Needs Audit): Stage 2 thinks the benchmark artifact may be broken or underspecified.\nLow Efficacy: sample looks too easy or saturated for the evaluated target.\nModerate Efficacy: sample provides some useful failure signal but not the sharpest available.\nHigh Efficacy: sample cleanly exposes meaningful model weakness.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/labeling.py :: classify_sample"),
    },
    "DominantFailureType": {
        "Intent": "Surface the most common concrete test-failure mode in the combined test run.",
        "How It's Computed": "Takes the modal failure type among failed combined tests.",
        "Decision Logic": "Descriptive only.",
        "Class Values": "Failure-type label such as format_mismatch, assertion_mismatch, runtime_error, or clean.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/runner.py :: _dominant_failure_type"),
    },
    "ProvidedPass@1": {
        "Intent": "Estimate probability of solving the provided tests within one attempt.",
        "How It's Computed": "Pass@k estimate over attempts using provided-test successes.",
        "Decision Logic": "Higher is better. Equals 1.0 if at least one provided-test success exists and k covers the attempt count.",
        "Class Values": "Fraction from 0.0 to 1.0.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/labeling.py :: estimate_pass_at_k"),
    },
    "ProvidedPass@2": {
        "Intent": "Estimate probability of solving the provided tests within two attempts.",
        "How It's Computed": "Pass@k estimate over attempts using provided-test successes.",
        "Decision Logic": "Higher is better.",
        "Class Values": "Fraction from 0.0 to 1.0.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/labeling.py :: estimate_pass_at_k"),
    },
    "CombinedPass@1": {
        "Intent": "Estimate probability of solving provided plus generated tests within one attempt.",
        "How It's Computed": "Pass@k estimate over attempts using combined-test successes.",
        "Decision Logic": "Higher is better.",
        "Class Values": "Fraction from 0.0 to 1.0.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/labeling.py :: estimate_pass_at_k"),
    },
    "CombinedPass@2": {
        "Intent": "Estimate probability of solving provided plus generated tests within two attempts.",
        "How It's Computed": "Pass@k estimate over attempts using combined-test successes.",
        "Decision Logic": "Higher is better.",
        "Class Values": "Fraction from 0.0 to 1.0.",
        "Code Location": rel("src/turing_takehome/stages/sample_efficacy_analysis/labeling.py :: estimate_pass_at_k"),
    },
}


def stage2_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    comparison_headers = load_csv_headers(STAGE2_RESULTS)
    model_headers = load_csv_headers(STAGE2_MODEL_RESULTS)

    for header in comparison_headers:
        if header in {"Index", "QuestionId", "QuestionTitle", "Difficulty"}:
            rows.append(
                metadata_row(
                    f"Comparison::{header}",
                    f"Comparison-row metadata field `{header}`.",
                    code_location=rel("src/turing_takehome/stages/sample_efficacy_analysis/runner.py :: _build_comparison_row"),
                )
            )
            continue
        definition = stage2_field_definition(header)
        rows.append(
            {
                "Field": f"Comparison::{header}",
                "Intent": definition.get("Intent", f"Per-sample comparison export field `{header}`."),
                "How It's Computed": clean_definition_cell(definition.get("How It's Computed")),
                "Decision Logic": stage2_decision_logic(header, definition.get("Decision Logic"), surface="Comparison"),
                "LLM Prompt": normalize_llm_prompt(definition.get("LLM Prompt"), field=header, stage="stage2"),
                "Class Values": stage2_class_values(header, definition.get("Class Values"), surface="Comparison"),
                "Code Location": definition.get("Code Location", rel("src/turing_takehome/stages/sample_efficacy_analysis/runner.py :: _build_comparison_row")),
                "Subjectivity of Test": stage2_subjectivity_for_field(header),
            }
        )

    for header in model_headers:
        if header in {"Index", "QuestionId", "QuestionTitle", "Difficulty", "TargetName", "ModelLabel"}:
            rows.append(
                metadata_row(
                    f"ModelSummary::{header}",
                    f"Per-model Stage 2 export field `{header}`.",
                    code_location=rel("src/turing_takehome/stages/sample_efficacy_analysis/runner.py :: _aggregate_model_attempts"),
                )
            )
            continue
        definition = stage2_field_definition(header)
        rows.append(
            {
                "Field": f"ModelSummary::{header}",
                "Intent": definition.get("Intent", f"Per-model Stage 2 export field `{header}`."),
                "How It's Computed": clean_definition_cell(definition.get("How It's Computed")),
                "Decision Logic": stage2_decision_logic(header, definition.get("Decision Logic"), surface="ModelSummary"),
                "LLM Prompt": normalize_llm_prompt(definition.get("LLM Prompt"), field=header, stage="stage2"),
                "Class Values": stage2_class_values(header, definition.get("Class Values"), surface="ModelSummary"),
                "Code Location": definition.get("Code Location", rel("src/turing_takehome/stages/sample_efficacy_analysis/runner.py :: _aggregate_model_attempts")),
                "Subjectivity of Test": stage2_subjectivity_for_field(header),
            }
        )
    return rows


STAGE3_SUMMARY_DEFS = {
    "dataset_size": ("Report the sample count available for dataset-level reasoning.", "Counts joined rows after Stage 1 and Stage 2 aggregation.", "PASS when the slice is large enough for structural reasoning; otherwise REVIEW.", "PASS or REVIEW, depending on dataset size."),
    "efficacy_distribution": ("Summarize the Stage 2 efficacy mix across the dataset.", "Counts Stage 2 efficacy labels across joined rows.", "Always descriptive; Result reflects whether the dataset spans multiple observed regimes.", "Distribution summary plus PASS/REVIEW result."),
    "template_overlap_scan": ("Highlight likely redundancy and template recycling in the dataset.", "Uses lexical plus embedding similarity to count materially redundant samples.", "REVIEW when redundancy candidates are present at a meaningful level.", "REVIEW-oriented structural redundancy signal."),
    "attempt_variance_scan": ("Surface samples whose repeated attempts are unstable.", "Counts stable, moderate, and volatile attempt-variance labels from Stage 3 features.", "REVIEW when nontrivial instability is present.", "Stable / moderate / volatile distribution with REVIEW-oriented result."),
    "model_disagreement_scan": ("Surface disagreement across models or the Stage 3 proxy disagreement layer.", "Counts aligned versus disagreement labels across rows.", "REVIEW when material disagreement exists.", "Aligned / moderate_disagreement / strong_disagreement distribution with REVIEW-oriented result."),
    "threshold_sensitivity_scan": ("Surface rows that flip interpretation under small threshold changes.", "Measures distance to Stage 2 efficacy cutoffs.", "REVIEW when nontrivial threshold-fragile rows exist.", "Stable / high / caveated threshold-sensitivity distribution."),
    "length_and_test_count_outliers": ("Count structural outliers in prompt length, starter length, ideal length, test count, and performance.", "Aggregates Stage 3 outlier flags derived from IQR-style checks.", "REVIEW when outliers exist.", "REVIEW-oriented outlier summary."),
    "high_static_low_dynamic_contradictions": ("Count rows that look strong statically but weak dynamically.", "Uses contradiction logic comparing Stage 1 strength with Stage 2 outcomes.", "REVIEW when contradictions are present.", "Contradiction-candidate count with REVIEW-oriented result."),
    "strongest_upstream_caveat": ("Show the strongest descriptive Stage 1 to Stage 2 relationship signal.", "Chooses the top descriptive relationship row by strength and support.", "INFO only; this is context, not a policy decision.", "INFO contextual signal."),
    "next_step": ("Record one recommended action arising from the Stage 3 synthesis.", "Generated from Stage 3 recommendation rules.", "ACTION rows are prescriptive follow-up guidance rather than measurements.", "ACTION recommendation row."),
    "prompt_length_distribution": ("Summarize prompt-length distribution for context.", "Uses prompt word counts across joined rows.", "INFO only.", "INFO descriptive distribution."),
    "test_count_distribution": ("Summarize test-count distribution for context.", "Uses total test counts across joined rows.", "INFO only.", "INFO descriptive distribution."),
}


STAGE3_FIELD_DEFS: dict[str, dict[str, str]] = {
    "PromptLengthBand": {"Intent": "Bucket prompt size for reviewer-friendly structural interpretation.", "How It's Computed": "Maps prompt word count into short / medium / long bands.", "Decision Logic": "short if under 150 words, medium if 150 to 500, otherwise long.", "Class Values": "short\nmedium\nlong", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _length_band")},
    "StarterCodeLengthBand": {"Intent": "Bucket starter-code size.", "How It's Computed": "Maps starter line count into minimal / moderate / large bands.", "Decision Logic": "minimal if up to 5 lines, moderate if 6 to 20, otherwise large.", "Class Values": "minimal\nmoderate\nlarge", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _length_band")},
    "IdealResponseLengthBand": {"Intent": "Bucket ideal-solution size.", "How It's Computed": "Maps ideal-response line count into compact / moderate / large bands.", "Decision Logic": "compact if up to 80 lines, moderate if 81 to 250, otherwise large.", "Class Values": "compact\nmoderate\nlarge", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _length_band")},
    "TestCountBand": {"Intent": "Bucket test-suite size for review and outlier interpretation.", "How It's Computed": "Maps total test count into sparse / recommended / heavy.", "Decision Logic": "sparse if under 10 tests, recommended if 10 to 20, otherwise heavy.", "Class Values": "sparse\nrecommended\nheavy", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _test_count_band")},
    "RedundancyLabel": {"Intent": "Describe whether a sample appears unique or near-duplicate relative to another sample.", "How It's Computed": "Derived from duplicate-pair analysis that combines lexical and embedding similarity.", "Decision Logic": "Unique rows stay unique; materially similar rows inherit the duplicate-pair label.", "Class Values": "Unique or a duplicate/redundancy label.", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _duplicate_detail_maps")},
    "ClosestNeighborIndex": {"Intent": "Point to the nearest detected redundancy neighbor.", "How It's Computed": "Taken from the highest-similarity duplicate-pair map.", "Decision Logic": "Blank when no meaningful neighbor exists.", "Class Values": "Sample index or blank.", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _duplicate_detail_maps")},
    "ClosestNeighborSimilarity": {"Intent": "Show the combined similarity score used for redundancy review.", "How It's Computed": "Combined lexical and embedding similarity score from duplicate analysis.", "Decision Logic": "Higher means more redundant.", "Class Values": "Similarity score from 0.0 to 1.0.", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _similarity_pairs")},
    "ClosestNeighborLexicalSimilarity": {"Intent": "Show the lexical portion of the redundancy similarity signal.", "How It's Computed": "Character- and token-based overlap score.", "Decision Logic": "Higher means more textually similar.", "Class Values": "Similarity score from 0.0 to 1.0.", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _similarity_pairs")},
    "ClosestNeighborEmbeddingSimilarity": {"Intent": "Show the embedding portion of the redundancy similarity signal.", "How It's Computed": "Cosine similarity between embedding vectors for prompt-centric text.", "Decision Logic": "Higher means more semantically similar.", "LLM Prompt": "Embedding model pass; see artifacts/audit/stage3/.", "Class Values": "Similarity score from 0.0 to 1.0.", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _attach_embedding_vectors")},
    "RedundancyClusterId": {"Intent": "Group redundancy-connected samples into reviewer-friendly clusters.", "How It's Computed": "Clustered from similarity pairs above the configured threshold.", "Decision Logic": "Blank or singleton-like behavior means no material redundancy cluster.", "Class Values": "Cluster identifier or blank-like singleton value.", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _cluster_pairs")},
    "RedundancyClusterSize": {"Intent": "Show how many samples are in the same redundancy cluster.", "How It's Computed": "Cluster cardinality from redundancy clustering.", "Decision Logic": "Larger clusters suggest more material template recycling.", "Class Values": "Positive integer count.", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _cluster_pairs")},
    "PromptLengthOutlier": {"Intent": "Flag prompts whose length is unusually extreme relative to the dataset.", "How It's Computed": "Derived from Stage 3 outlier flags.", "Decision Logic": "FLAG when the prompt-length outlier flag is present, else PASS.", "Class Values": "PASS\nFLAG", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _compute_outlier_flags")},
    "StarterCodeOutlier": {"Intent": "Flag unusually large or small starter code.", "How It's Computed": "Derived from Stage 3 outlier flags.", "Decision Logic": "FLAG when the starter-code outlier flag is present, else PASS.", "Class Values": "PASS\nFLAG", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _compute_outlier_flags")},
    "IdealResponseOutlier": {"Intent": "Flag unusually large or small ideal responses.", "How It's Computed": "Derived from Stage 3 outlier flags.", "Decision Logic": "FLAG when the ideal-response outlier flag is present, else PASS.", "Class Values": "PASS\nFLAG", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _compute_outlier_flags")},
    "TestCountOutlier": {"Intent": "Flag unusually sparse or heavy test suites.", "How It's Computed": "Derived from Stage 3 outlier flags.", "Decision Logic": "FLAG when the test-count outlier flag is present, else PASS.", "Class Values": "PASS\nFLAG", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _compute_outlier_flags")},
    "PerformanceOutlier": {"Intent": "Flag pass-rate behavior that is extreme relative to the dataset.", "How It's Computed": "Derived from Stage 3 outlier flags over behavioral performance.", "Decision Logic": "FLAG when the pass-rate outlier flag is present, else PASS.", "Class Values": "PASS\nFLAG", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _compute_outlier_flags")},
    "RedundancyStatus": {"Intent": "Provide a direct PASS/FLAG redundancy gate for the sample.", "How It's Computed": "Flags rows whose redundancy score is at least 0.55.", "Decision Logic": "FLAG if redundancy score >= 0.55, else PASS.", "Class Values": "PASS\nFLAG", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _build_stage3_detailed_rows")},
    "AttemptVarianceCheck": {"Intent": "Show whether repeated attempts were behaviorally stable.", "How It's Computed": "Flags the attempt-variance label when it is not stable.", "Decision Logic": "PASS if stable, FLAG if moderate or volatile.", "Class Values": "PASS\nFLAG", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _attempt_variance_features")},
    "AttemptVarianceLabel": {"Intent": "Classify repeated-attempt stability.", "How It's Computed": "Uses attempt pass-rate range, standard deviation, and execution success rate.", "Decision Logic": "stable for small variation, moderate for noticeable but not extreme variation, volatile for failed executions or large swings, unknown if no attempts.", "Class Values": "stable\nmoderate\nvolatile\nunknown", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _attempt_variance_features")},
    "AttemptCombinedPassRange": {"Intent": "Quantify the best-minus-worst combined pass-rate spread across attempts.", "How It's Computed": "Best combined pass rate minus worst combined pass rate.", "Decision Logic": "Higher means more instability.", "Class Values": "Range from 0.0 upward.", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _attempt_variance_features")},
    "AttemptCombinedPassStdDev": {"Intent": "Quantify the spread of combined pass rates across attempts.", "How It's Computed": "Population standard deviation over attempt combined pass rates.", "Decision Logic": "Higher means more instability.", "Class Values": "Standard deviation from 0.0 upward.", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _attempt_variance_features")},
    "AttemptExecutionSuccessRate": {"Intent": "Quantify how many attempts executed cleanly.", "How It's Computed": "Fraction of attempts with execution probe status ok.", "Decision Logic": "Less than 1.0 contributes to volatile labeling.", "Class Values": "Fraction from 0.0 to 1.0.", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _attempt_variance_features")},
    "ModelDisagreementCheck": {"Intent": "Provide a direct PASS/FLAG gate for cross-model or proxy disagreement.", "How It's Computed": "Flags rows whose model-disagreement label is not aligned or single_model.", "Decision Logic": "PASS if aligned or single_model, FLAG otherwise.", "Class Values": "PASS\nFLAG", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _model_disagreement_features")},
    "ModelDisagreementLabel": {"Intent": "Classify cross-model disagreement strength.", "How It's Computed": "Uses gaps in best combined pass rate plus disagreements in efficacy, benchmark-quality, and suspicion labels.", "Decision Logic": "single_model if only one model is available, aligned if no material disagreement, moderate_disagreement for noticeable but not severe gaps, strong_disagreement for large gaps or suspicion disagreement.", "Class Values": "single_model\naligned\nmoderate_disagreement\nstrong_disagreement\nunknown", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _model_disagreement_features")},
    "ModelDisagreementSource": {"Intent": "Record where the disagreement signal came from when a lightweight Stage 3 auditor was used.", "How It's Computed": "Taken from the disagreement-source metadata attached during Stage 3 analysis.", "Decision Logic": "Descriptive only.", "LLM Prompt": "Stage 3 auditor prompt; see artifacts/audit/stage3/.", "Class Values": "Source label or blank.", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _stage3_auditor_disagreement")},
    "ModelCount": {"Intent": "Show how many models or disagreement sources were compared for this row.", "How It's Computed": "Count of model rows in the disagreement feature builder.", "Decision Logic": "Descriptive only.", "Class Values": "Positive integer count.", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _model_disagreement_features")},
    "ModelBestCombinedPassGap": {"Intent": "Quantify the gap between the strongest and weakest model in the disagreement surface.", "How It's Computed": "Max minus min best combined pass rate across model rows.", "Decision Logic": "Higher gaps push the row toward stronger disagreement labels.", "Class Values": "Gap from 0.0 upward.", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _model_disagreement_features")},
    "ThresholdSensitivityCheck": {"Intent": "Provide a direct PASS/FLAG gate for threshold fragility.", "How It's Computed": "Flags rows whose threshold-sensitivity label is not stable.", "Decision Logic": "PASS if stable, FLAG if high or caveated.", "Class Values": "PASS\nFLAG", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _threshold_sensitivity_features")},
    "ThresholdSensitivityLabel": {"Intent": "Classify how fragile the Stage 2 interpretation is to small pass-rate threshold shifts.", "How It's Computed": "Measures distance from the nearest key Stage 2 efficacy threshold.", "Decision Logic": "caveated if the row is already suspicious, high when very close to a threshold, stable otherwise.", "Class Values": "stable\nhigh\ncaveated", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _threshold_sensitivity_features")},
    "ThresholdNearestBoundary": {"Intent": "Show which Stage 2 efficacy boundary the row is closest to.", "How It's Computed": "Nearest among the configured pass-rate thresholds.", "Decision Logic": "Descriptive only.", "Class Values": "Boundary label such as high_efficacy_floor, moderate_high_split, or saturation_boundary.", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _threshold_sensitivity_features")},
    "ThresholdDistance": {"Intent": "Show how far the row sits from its nearest key threshold.", "How It's Computed": "Absolute distance from the nearest configured threshold.", "Decision Logic": "Smaller means more fragile.", "Class Values": "Distance from 0.0 upward.", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _threshold_sensitivity_features")},
    "DifficultySignalRegime": {"Intent": "Summarize the overall interpretive regime of the sample as benchmark signal.", "How It's Computed": "Combines suspiciousness, threshold fragility, disagreement, volatility, and combined pass rate.", "Decision Logic": "caveated if suspicious, threshold_fragile if near a threshold, unstable if variance/disagreement dominates, saturated if pass rate >= 0.95, extreme_or_blocked if pass rate <= 0.1, otherwise discriminative.", "Class Values": "caveated\nthreshold_fragile\nunstable\nsaturated\nextreme_or_blocked\ndiscriminative", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _difficulty_signal_regime")},
    "ContradictionCheck": {"Intent": "Flag samples whose static artifact quality and dynamic behavior sharply disagree.", "How It's Computed": "Compares Stage 1 score against Stage 2 suspiciousness, pass rate, and efficacy label.", "Decision Logic": "high_static_low_dynamic when static quality is high but dynamic behavior is weak or suspicious; low_static_high_dynamic when static quality is low but dynamic behavior is very strong; none otherwise.", "Class Values": "none\nhigh_static_low_dynamic\nlow_static_high_dynamic", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _contradiction_label")},
    "BenchmarkDefectCandidate": {"Intent": "Give a direct PASS/FLAG signal for likely broken benchmark artifacts.", "How It's Computed": "Flags rows that are both critical-priority and benchmark-defect candidates.", "Decision Logic": "FLAG when audit priority is critical and the row is suspicious or has oracle pass rate below 1.0.", "Class Values": "PASS\nFLAG", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _is_benchmark_defect_candidate")},
    "TrivialityCheck": {"Intent": "Flag samples that are likely too easy or saturated to add much benchmark signal.", "How It's Computed": "Checks whether the row is already Low Efficacy or near-perfect in winner combined pass rate.", "Decision Logic": "FLAG if the row is trivially easy or saturated, else PASS.", "Class Values": "PASS\nFLAG", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _is_trivial_candidate")},
    "ExemplarCheck": {"Intent": "Flag rows that look like strong clean exemplars of the benchmark's intended signal.", "How It's Computed": "Combines strong Stage 1 quality, informative Stage 2 efficacy, stability, lack of suspicion, and low redundancy.", "Decision Logic": "FLAG when all exemplar criteria are satisfied, else PASS.", "Class Values": "PASS\nFLAG", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _is_exemplar_candidate")},
    "AuditPriority": {"Intent": "Rank how urgently a sample should be reviewed or adjudicated.", "How It's Computed": "Combines suspiciousness, oracle failures, contradiction signals, redundancy, outliers, variance, disagreement, and threshold fragility.", "Decision Logic": "critical for suspicious/oracle-broken/contradictory severe rows, high for strong redundancy or contradiction, medium for softer caveats, normal otherwise.", "Class Values": "normal\nmedium\nhigh\ncritical", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _audit_priority")},
    "AuditReason": {"Intent": "Provide a concise reviewer-facing explanation for the audit-priority assignment.", "How It's Computed": "Synthesizes the dominant reasons that pushed the row into the queue.", "Decision Logic": "Descriptive only.", "Class Values": "Short free-text reason summary.", "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _audit_reason_summary")},
}


def stage3_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    summary_names = load_stage3_summary_test_names(STAGE3_WORKBOOK)
    for name in summary_names:
        intent, computed, logic, values = STAGE3_SUMMARY_DEFS.get(
            name,
            (
                f"Stage 3 summary row `{name}`.",
                SEE_CODE_LOCATION,
                SEE_CODE_LOCATION,
                "",
            ),
        )
        rows.append(
            {
                "Field": f"Summary::{name}",
                "Intent": intent,
                "How It's Computed": clean_definition_cell(computed, allow_blank=False),
                "Decision Logic": clean_definition_cell(logic, allow_blank=False),
                "LLM Prompt": "",
                "Class Values": stage3_class_values(f"Summary::{name}", values),
                "Code Location": rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _build_summary"),
                "Subjectivity of Test": subjectivity_label(1),
            }
        )

    wb = load_workbook(STAGE3_WORKBOOK, read_only=True, data_only=True)
    ws = wb["Detailed"]
    detailed_headers = [cell for cell in next(ws.iter_rows(values_only=True))]
    for header in detailed_headers:
        if header in {"Index", "QuestionId", "QuestionTitle", "Difficulty", "FunctionName"}:
            rows.append(metadata_row(f"Detailed::{header}", f"Stage 3 detailed metadata field `{header}`.", code_location=rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _build_stage3_detailed_rows")))
            continue
        if str(header).startswith("Notes-"):
            continue
        definition = STAGE3_FIELD_DEFS.get(header, {})
        rows.append(
            {
                "Field": f"Detailed::{header}",
                "Intent": definition.get("Intent", f"Stage 3 detailed field `{header}`."),
                "How It's Computed": clean_definition_cell(definition.get("How It's Computed")),
                "Decision Logic": stage3_decision_logic(str(header), definition.get("Decision Logic")),
                "LLM Prompt": normalize_llm_prompt(definition.get("LLM Prompt", ""), field=str(header), stage="stage3"),
                "Class Values": stage3_class_values(str(header), definition.get("Class Values")),
                "Code Location": definition.get("Code Location", rel("src/turing_takehome/stages/dataset_analysis/runner.py :: _build_stage3_detailed_rows")),
                "Subjectivity of Test": stage3_subjectivity_for_field(str(header)),
            }
        )
    return rows


def write_sheet(ws, rows: list[dict[str, str]]) -> None:
    ws.append(HEADERS)
    for row in rows:
        ws.append([row.get(header, "") for header in HEADERS])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
            value = str(cell.value or "").strip()
            if value == USING_LLM:
                cell.font = LLM_FONT
            elif value == DIRECT_EVALUATION:
                cell.font = DIRECT_FONT
            elif value == SEE_CODE_LOCATION:
                cell.font = SEE_CODE_FONT
    ws.freeze_panes = "A2"
    widths = {"A": 34, "B": 42, "C": 40, "D": 40, "E": 52, "F": 44, "G": 44, "H": 30}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def main() -> int:
    workbook = Workbook()
    stage1_ws = workbook.active
    stage1_ws.title = "Stage 1"
    write_sheet(stage1_ws, stage1_rows())
    stage2_ws = workbook.create_sheet("Stage 2")
    write_sheet(stage2_ws, stage2_rows())
    stage3_ws = workbook.create_sheet("Stage 3")
    write_sheet(stage3_ws, stage3_rows())
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    print(f"Wrote workbook field guide to {OUTPUT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
