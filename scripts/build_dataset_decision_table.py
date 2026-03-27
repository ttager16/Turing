from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
STAGE1_WORKBOOK = PROJECT_ROOT / "outputs" / "sample_requirements_analysis" / "guideline_audit.xlsx"
STAGE2_RESULTS = PROJECT_ROOT / "outputs" / "sample_efficacy_analysis" / "sample_results.csv"
STAGE3_DETAILED = PROJECT_ROOT / "outputs" / "dataset_analysis" / "enriched_samples.csv"
OUTPUT_PATH = PROJECT_ROOT / "private" / "dataset_decision_table_preview.csv"


def load_stage1_summary(path: Path) -> dict[int, dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Summary"]
    rows = ws.iter_rows(values_only=True)
    headers = [str(cell) for cell in next(rows)]
    by_index: dict[int, dict[str, str]] = {}
    for row in rows:
        payload = {header: ("" if value is None else str(value).strip()) for header, value in zip(headers, row)}
        if not payload.get("Index"):
            continue
        by_index[int(payload["Index"])] = payload
    return by_index


def load_csv_by_index(path: Path) -> dict[int, dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return {
            int(str(row["Index"]).strip()): {key: ("" if value is None else str(value).strip()) for key, value in row.items()}
            for row in reader
        }


def recommended_action(stage1: dict[str, str], stage2: dict[str, str], stage3: dict[str, str]) -> tuple[str, str, str, str]:
    stage1_values = [stage1.get("Prompt", ""), stage1.get("Ideal_Response", ""), stage1.get("Test_Cases", "")]
    unusable_count = sum(value == "Unusable" for value in stage1_values)
    needs_fixing = any(value == "Needs Fixing" for value in stage1_values)
    suspicious = stage2.get("Suspicious", "").lower() == "true"
    efficacy = stage2.get("ModelAEfficacyLabel", "")
    benchmark_quality = stage2.get("ModelABenchmarkQualitySignal", "")
    contradiction = stage3.get("ContradictionCheck", "")
    benchmark_defect = stage3.get("BenchmarkDefectCandidate", "")
    redundancy = stage3.get("RedundancyStatus", "")
    triviality = stage3.get("TrivialityCheck", "")
    exemplar = stage3.get("ExemplarCheck", "")
    priority = stage3.get("AuditPriority", "normal") or "normal"

    reasons: list[str] = []
    evidence: list[str] = []

    if benchmark_defect == "FLAG" or suspicious:
        action = "FIX_OR_REMOVE"
        fix_cost = "medium"
        reasons.append("Stage 2 or Stage 3 flagged the sample as likely benchmark-defective.")
        evidence.extend([
            "outputs/sample_efficacy_analysis/sample_results.csv::Suspicious",
            "outputs/dataset_analysis/enriched_samples.csv::BenchmarkDefectCandidate",
        ])
    elif unusable_count >= 2:
        action = "FIX"
        fix_cost = "high"
        reasons.append("Stage 1 marked multiple artifact layers as unusable.")
        evidence.append("outputs/sample_requirements_analysis/guideline_audit.xlsx::Summary")
    elif redundancy == "FLAG":
        action = "DEDUPE_OR_DOWNWEIGHT"
        fix_cost = "low"
        reasons.append("Stage 3 found a material redundancy signal.")
        evidence.append("outputs/dataset_analysis/enriched_samples.csv::RedundancyStatus")
    elif triviality == "FLAG":
        action = "DEPRIORITIZE"
        fix_cost = "low"
        reasons.append("Stage 3 marked the sample as trivial or saturated.")
        evidence.append("outputs/dataset_analysis/enriched_samples.csv::TrivialityCheck")
    elif exemplar == "FLAG":
        action = "KEEP"
        fix_cost = "low"
        reasons.append("Stage 3 surfaced the sample as a strong exemplar candidate.")
        evidence.append("outputs/dataset_analysis/enriched_samples.csv::ExemplarCheck")
    elif needs_fixing:
        action = "FIX"
        fix_cost = "medium"
        reasons.append("Stage 1 found repairable artifact-quality issues.")
        evidence.append("outputs/sample_requirements_analysis/guideline_audit.xlsx::Summary")
    else:
        action = "KEEP"
        fix_cost = "low"
        reasons.append("No strong structural or behavioral defect signal was raised in Stage 1 to Stage 3.")
        evidence.extend([
            "outputs/sample_requirements_analysis/guideline_audit.xlsx::Summary",
            "outputs/sample_efficacy_analysis/sample_results.csv::ModelAEfficacyLabel",
        ])

    if contradiction and contradiction != "none":
        reasons.append(f"Stage 3 contradiction signal: {contradiction}.")
        evidence.append("outputs/dataset_analysis/enriched_samples.csv::ContradictionCheck")
    if benchmark_quality and benchmark_quality != "clean_evaluation":
        reasons.append(f"Stage 2 benchmark-quality caveat: {benchmark_quality}.")
        evidence.append("outputs/sample_efficacy_analysis/sample_results.csv::ModelABenchmarkQualitySignal")
    if efficacy:
        reasons.append(f"Stage 2 efficacy label: {efficacy}.")

    return action, priority, " ".join(reasons), "; ".join(dict.fromkeys(evidence)), fix_cost


def build_rows() -> list[dict[str, str]]:
    stage1 = load_stage1_summary(STAGE1_WORKBOOK)
    stage2 = load_csv_by_index(STAGE2_RESULTS)
    stage3 = load_csv_by_index(STAGE3_DETAILED)
    rows: list[dict[str, str]] = []
    for index in sorted(stage3):
        s1 = stage1[index]
        s2 = stage2[index]
        s3 = stage3[index]
        action, priority, rationale, evidence_source, fix_cost = recommended_action(s1, s2, s3)
        current_status = f"Stage1={s1['Prompt']}/{s1['Ideal_Response']}/{s1['Test_Cases']}; Stage2={s2['ModelAEfficacyLabel']}; Stage3Priority={s3['AuditPriority']}"
        rows.append(
            {
                "sample_id": str(index),
                "question_id": s1["Question_Id"],
                "question_title": s1["Question_Title"],
                "current_status": current_status,
                "recommended_action": action,
                "priority": priority,
                "rationale": rationale,
                "evidence_source": evidence_source,
                "fix_cost_estimate": fix_cost,
                "notes": "Heuristic preview generated from Stage 1-3 outputs only; not a human-adjudicated decision.",
            }
        )
    return rows


def main() -> int:
    rows = build_rows()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT_PATH.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "sample_id",
                "question_id",
                "question_title",
                "current_status",
                "recommended_action",
                "priority",
                "rationale",
                "evidence_source",
                "fix_cost_estimate",
                "notes",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)
    print(f"Wrote dataset decision table to {OUTPUT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
